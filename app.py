from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file, Response, jsonify
import os
import requests
import base64
from werkzeug.utils import secure_filename
from datetime import datetime, date, timedelta
from openpyxl import load_workbook
from PyPDF2 import PdfReader, PdfWriter
from PyPDF2.generic import BooleanObject, NameObject, NumberObject, DictionaryObject
# Reportlab: overlay de texto compatible con visores web
try:
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.utils import simpleSplit
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False
import mimetypes
import io
import uuid
import json
import pandas as pd
import unicodedata
import secrets
import re
import urllib.parse
# Las importaciones específicas para Google Drive API han sido eliminadas.


app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "clave_super_segura_cardiohome_2025")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
ANTHROPIC_MODEL   = "claude-sonnet-4-6"
ALLOWED_EXTENSIONS = {'pdf', 'docx', 'doc', 'xls', 'xlsx', 'csv'}

# Define los PDFs base para cada tipo de formulario
# Asegúrate de que estos archivos PDF existan en la misma carpeta que app.py
PDF_BASE_NEUROLOGIA = 'FORMULARIO TIPO NEUROLOGIA INFANTIL EDITABLE.pdf'
PDF_BASE_FAMILIAR = 'formulario_familiar.pdf' 
PDF_BASE_INFORME_NEURO = 'INFORME_NEUROLOGICO_BASE.pdf'
PDF_BASES_NEUROLOGIA_DIR = 'pdf_bases_doctoras_neurologia'
# Directorio para PDFs personalizados de medicina familiar por doctora
PDF_BASES_FAMILIAR_DIR = 'pdf_bases_doctoras_familiar'


# -------------------- Supabase Configuration --------------------
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://rbzxolreglwndvsrxhmg.supabase.co")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InJienhvbHJlZ2x3bmR2c3J4aG1nIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NDc1NDE3ODcsImV4cCI6MjA2MzExNzc4N30.BbzsUhed1Y_dJYWFKLAHqtV4cXdvjF_ihGdQ_Bpov3Y")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "eyJhbGciOiJIUzI1NiIsInR5cCI6IlNJUDU4IiwicmVmIjoiYnhzbnFmZml4d2pkcWl2eGJrZXkiLCJyb2xlIjoic2VydmljZV9yb2xlIiwiaWF0IjoxNzE5Mjg3MzI1LCJleHAiOjE3NTA4MjMzMjV9.qNlSg_p4_u1O5xQ9s6bNN0K2Z0f0v_N9s8k0k0k0k0k") # ASEGÚRATE DE USAR TU SERVICE_KEY REAL

SUPABASE_HEADERS = {
    "apikey": SUPABASE_KEY,
    "Authorization": f"Bearer {SUPABASE_KEY}",
    "Content-Type": "application/json",
    "Accept": "application/json" 
}
SUPABASE_SERVICE_HEADERS = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
    "Accept": "application/json",
    "Prefer": "count=exact"
}

# Configuración de SendGrid
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
SENDGRID_FROM = os.getenv("SENDGRID_FROM_EMAIL", 'your_sendgrid_email@example.com')
SENDGRID_TO = os.getenv("SENDGRID_ADMIN_EMAIL", 'destination_admin_email@example.com')


# -------------------- Utilidades --------------------
def format_rut_python(rut):
    """
    Formatea un RUT chileno (ej: 12345678-9) a un formato con puntos y guion (ej: 12.345.678-9).
    Acepta RUTs con o sin puntos y guiones.
    """
    if not rut:
        return ""
    
    # Asegurarse de que el RUT sea una cadena y limpiar puntos y guiones existentes
    rut = str(rut).replace('.', '').replace('-', '').strip().upper() 

    if not rut:
        return ""

    # Separar cuerpo y dígito verificador
    body = rut[:-1]
    dv = rut[-1]

    # Formatear el cuerpo con puntos
    formatted_body = ""
    for i, digit in enumerate(reversed(body)):
        if i > 0 and i % 3 == 0:
            formatted_body = "." + formatted_body
        formatted_body = digit + formatted_body

    return f"{formatted_body}-{dv}"

# app-30.py (VERSIÓN FINAL CON PREFER HEADER Y MANEJO DE FILTROS)

# Asegúrate de que 'import uuid' esté al inicio de app.py

# Reemplaza la función get_supabase_count en app.py por esta versión:

# --- Asegúrate de que esta línea esté al inicio de app.py:
# from datetime import datetime, date
# import requests 
# import uuid
# ---

# app-31.py (Reemplazo de la función get_supabase_count)
# app.py (Reemplazo de la función get_supabase_count)

# app.py (Reemplazo de la función get_supabase_count)
# Asegúrate de tener: 
# import urllib.parse 
# from datetime import datetime 
# import requests 
# ...

# app.py (Reemplazo de la función get_supabase_count)
# app.py (Reemplazo de la función get_supabase_count)
# Necesita: from datetime import datetime, import requests

# app.py (Reemplazo de la función get_supabase_count)
# Necesita: from datetime import datetime, import requests

# app.py (Reemplazo de la función get_supabase_count)

# app.py (Reemplazo de la función get_supabase_count)
# Asegúrate de tener import requests y from datetime import datetime

# --- UTILIDAD ROBUSTA DE CONTEO: get_supabase_count ---
def get_supabase_count(filter_params=""):
    """
    Retorna un entero con el conteo de filas coincidentes en estudiantes_nomina.
    Usa select=id y lee Content-Range. filter_params debe ser algo como:
      "nomina_id=eq.<uuid>&evaluado_flag=eq.true"
    """
    # Normalizar filtro (no empezar con &)
    if filter_params and filter_params.startswith('&'):
        filter_params = filter_params[1:]

    url = f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?select=id"
    if filter_params:
        url = f"{url}&{filter_params}"

    try:
        # Usamos SERVICE headers (tienes Prefer: count=exact ahí)
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        res.raise_for_status()

        # Content-Range ejemplo: "0-9/24" → total = 24
        content_range = res.headers.get("Content-Range")
        if content_range and '/' in content_range:
            total_str = content_range.split('/')[-1]
            try:
                return int(total_str)
            except ValueError:
                pass

        # Si no hay Content-Range, fallback a len(json)
        data = res.json()
        return len(data) if isinstance(data, list) else 0

    except Exception as e:
        print(f"❌ ERROR en get_supabase_count. URL: {url}. Error: {e}")
        return 0


def get_counts_para_nominas(nomina_ids):
    """
    Versión rápida: obtiene total y evaluados para TODAS las nóminas
    en solo 1 llamada a Supabase en lugar de 2×N llamadas individuales.
    Con 20 nóminas: 40 requests → 1 request.

    Retorna dict: { nomina_id: {"total": int, "evaluados": int} }
    """
    if not nomina_ids:
        return {}

    ids_str   = ','.join(str(i) for i in nomina_ids)
    resultado = {str(nid): {"total": 0, "evaluados": 0} for nid in nomina_ids}

    try:
        url = (
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=in.({ids_str})"
            f"&estado_asistencia=in.(activo,extra)"
            f"&select=nomina_id,evaluado_flag"
        )
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        if res.ok:
            for row in res.json():
                nid = str(row.get('nomina_id', ''))
                if nid in resultado:
                    resultado[nid]["total"] += 1
                    if row.get('evaluado_flag') is True:
                        resultado[nid]["evaluados"] += 1
    except Exception as e:
        print(f"❌ ERROR en get_counts_para_nominas: {e}")

    return resultado

# app-30.py (Función get_assigned_nomina_ids - CORREGIDA)

def get_assigned_nomina_ids(user_id):
    """Obtiene la lista de IDs de nóminas asignadas directamente al user_id."""
    
    # 🟢 CORRECCIÓN CLAVE: Usamos el nombre de columna real 'coord_general_id'
    columna_asignacion = "coord_general_id" 
    
    url_asignaciones = (
        f"{SUPABASE_URL}/rest/v1/nominas_medicas"
        f"?{columna_asignacion}=eq.{user_id}" 
        f"&select=id"
    )
    
    try:
        print(f"DEBUG: Consultando nóminas asignadas: {url_asignaciones}") 
        
        # Usa SERVICE HEADERS para asegurar el acceso a la tabla de nóminas
        res = requests.get(url_asignaciones, headers=SUPABASE_SERVICE_HEADERS)
        res.raise_for_status()
        
        nomina_ids = [item['id'] for item in res.json()]
        print(f"DEBUG: Nóminas asignadas a {user_id}: {len(nomina_ids)} nóminas encontradas.")
        return nomina_ids
        
    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR al obtener nóminas asignadas para {user_id}: {e}")
        return []
        
def permitido(filename):
    """Verifica si la extensión del archivo está permitida."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def calculate_age(birth_date, fecha_ref=None):
    """
    Calcula la edad exacta en años y meses a partir de una fecha de nacimiento.
    - fecha_ref: fecha de referencia (fecha_evaluacion). Si no se pasa, usa hoy.
    - Considera el día del mes para no adelantar el cumpleaños.
      Ej: nacido 30/04/2019, evaluado 04/04/2026 → 6 años con 11 meses (NO 7 años).
    """
    ref = fecha_ref if fecha_ref else date.today()
    years  = ref.year  - birth_date.year
    months = ref.month - birth_date.month
    # Si el día de la ref es menor al día de nacimiento, el mes aún no se cumplió
    if ref.day < birth_date.day:
        months -= 1
    if months < 0:
        years  -= 1
        months += 12
    if years < 0:
        return "0 meses"
    if years == 0:
        return f"{months} meses"
    return f"{years} años con {months} meses"

def guess_gender(name):
    """
    Intenta adivinar el género basado en el nombre (heurística simple).
    Retorna 'M', 'F' o None si no puede adivinar.
    """
    name_lower = name.lower().strip()
    first_word = name_lower.split(' ')[0]

    nombres_masculinos = ["juan", "pedro", "luis", "carlos", "jose", "manuel", "alejandro", "ignacio", "felipe", "vicente", "emilio", "cristobal", "mauricio", "diego", "jean", "agustin", "joaquin", "thomas", "martin", "angel", "alonso"]
    nombres_femeninos = ["maria", "ana", "sofia", "laura", "paula", "trinidad", "mariana", "lizeth", "alexandra", "lisset"] 

    if first_word in nombres_masculinos:
        return 'M'
    elif first_word in nombres_femeninos:
        return 'F'
    
    return None # Retorna None si no puede adivinar con certeza
        
def normalizar(texto):
    """Normaliza texto: quita espacios, minúsculas, tildes y reemplaza espacios por guiones bajos."""
    if not isinstance(texto, str):
        return ""
    texto = texto.strip().lower()
    texto = unicodedata.normalize('NFKD', texto).encode('ascii', 'ignore').decode('utf-8')
    texto = texto.replace(" ", "_")
    return texto

def enviar_correo_sendgrid(asunto, cuerpo, adjuntos=None):
    """Envía un correo electrónico usando la API de SendGrid."""
    if not SENDGRID_API_KEY:
        print("Falta SENDGRID_API_KEY en variables de entorno. No se enviará correo.")
        return

    data = {
        "personalizations": [{"to": [{"email": SENDGRID_TO}]}],
        "from": {"email": SENDGRID_FROM},
        "subject": asunto,
        "content": [{"type": "text/plain", "value": cuerpo}]
    }

    if adjuntos:
        data["attachments"] = [
            {
                "content": adj["content"],
                "filename": adj["filename"],
                "type": "application/octet-stream", 
                "disposition": "attachment"
            } for adj in adjuntos
        ]

    try:
        response = requests.post(
            "https://api.sendgrid.com/v3/mail/send",
            headers={
                "Authorization": f"Bearer {SENDGRID_API_KEY}",
                "Content-Type": "application/json"
            },
            json=data
        )
        print(f"Correo enviado, status: {response.status_code}")
        if response.status_code >= 400:
            print(f"Error SendGrid Response: {response.text}")
    except Exception as e:
        print(f"Error al enviar correo con SendGrid: {e}")

# app-30.py (Función generate_and_upload_pdf corregida)

# Asegúrate de que 'os' y 'requests' estén importados al inicio de app-30.py
# import os
# import requests
# ...
# import io
# from PyPDF2 import PdfReader, PdfWriter
# ...

# app-30.py (Función auxiliar generate_and_upload_pdf COMPLETA)

def generate_and_upload_pdf(estudiante_id, nomina_id, doctora_id, form_type, datos_actualizacion):
    """Genera el PDF rellenado y lo sube al almacenamiento de Supabase.
       Incluye lógica para seleccionar plantilla específica por doctora.
    """
    
    # ---------------------------------------------------------------------
    # 1. OBTENCIÓN DEL ID PARA EL FORMULARIO (CORRECCIÓN CLAVE)
    # ---------------------------------------------------------------------
    pdf_template_path = None
    
    # Usamos el ID de la Doctora LOGUEADA como la clave para la plantilla.
    # El ID de la sesión es la única fuente de verdad fiable en este punto.
    current_doctora_id = session.get('usuario_id')
    
    if form_type == 'neurologia' and current_doctora_id:

        # 1.1. Definir la ruta del formulario específico usando el ID LOGUEADO
        doctora_pdf_filename = f"FORMULARIO TIPO NEUROLOGIA_{current_doctora_id}.pdf"
        # La carpeta donde deben estar tus archivos personalizados
        doctora_pdf_path = os.path.join(PDF_BASES_NEUROLOGIA_DIR, doctora_pdf_filename)

        # 1.2. Comprobar si existe el archivo PDF específico
        if os.path.exists(doctora_pdf_path):
            pdf_template_path = doctora_pdf_path
            print(f"DEBUG: Usando PDF específico de Doctora LOGUEADA: {pdf_template_path}")
        else:
            # Fallback: Si no existe el archivo específico, usar el genérico
            pdf_template_path = PDF_BASE_NEUROLOGIA
            print(f"DEBUG: Usando PDF genérico de Neurología.")
            
    elif form_type == 'medicina_familiar':
        pdf_template_path = PDF_BASE_FAMILIAR
    
    if not pdf_template_path or not os.path.exists(pdf_template_path):
        message = f"ERROR: Plantilla PDF no encontrada para el tipo de formulario '{form_type}' en la ruta: {pdf_template_path}"
        print(message)
        return {"success": False, "message": message}

    # ---------------------------------------------------------------------
    # 2. CONTINUACIÓN DE LA GENERACIÓN Y RELLENO DEL PDF
    # ---------------------------------------------------------------------
    try:
        # 2.1. Recuperar información completa del estudiante (ya que update_data solo tiene lo nuevo)
        url_estudiante_completo = (
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?id=eq.{estudiante_id}"
            f"&select=nombre,rut,fecha_nacimiento,edad,nacionalidad,sexo,estado_general,diagnostico,derivaciones,fecha_evaluacion,fecha_reevaluacion"
        )
        res_est = requests.get(url_estudiante_completo, headers=SUPABASE_SERVICE_HEADERS)
        estudiante_data = res_est.json()[0] if res_est.ok and res_est.json() else {}

        # 2.2. Recuperar información de la Doctora (Firma)
        url_doctora = f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{doctora_id}&select=nombre"
        res_doc = requests.get(url_doctora, headers=SUPABASE_SERVICE_HEADERS)
        doctora_nombre = res_doc.json()[0]['nombre'] if res_doc.ok and res_doc.json() else 'Doctora Asignada'
        
        # --- MAPEO DE CAMPOS (ADAPTAR ESTO A TU FORMATO PDF REAL) ---
        campos_pdf = {} 
        # Campos comunes
        campos_pdf['NOMBRE_ESTUDIANTE'] = estudiante_data.get('nombre', '')
        campos_pdf['RUT'] = estudiante_data.get('rut', '')
        campos_pdf['FECHA_EVALUACION'] = estudiante_data.get('fecha_evaluacion', '')
        campos_pdf['FECHA_REVALUACION'] = estudiante_data.get('fecha_reevaluacion', '')
        campos_pdf['DOCTORA_NOMBRE'] = doctora_nombre
        # Si form_type es neurologia (solo un ejemplo)
        if form_type == 'neurologia':
            campos_pdf['ESTADO_GENERAL'] = estudiante_data.get('estado_general', '')
            campos_pdf['DIAGNOSTICO'] = estudiante_data.get('diagnostico', '')
            campos_pdf['DERIVACIONES'] = estudiante_data.get('derivaciones', '')
        # Si form_type es medicina_familiar (solo un ejemplo)
        elif form_type == 'medicina_familiar':
            # Aquí iría el mapeo de todos los campos específicos de medicina familiar
            pass
        # --------------------------------------------------------------
        
        # 2.3. Rellenar PDF
        # Nota: Asegúrate de que PdfReader y PdfWriter estén importados
        with open(pdf_template_path, 'rb') as file:
            reader = PdfReader(file)
            writer = PdfWriter()
            page = reader.pages[0]
            writer.add_page(page)

            # Rellenar los campos
            if reader.get_form_text_fields():
                writer.update_page_form_field_values(writer.pages[0], campos_pdf)
                print(f"DEBUG: Campos PDF rellenados: {list(campos_pdf.keys())}")
            else:
                print("ADVERTENCIA: No se encontraron campos de formulario en el PDF.")

            # Auto-size: evita que texto largo se corte o salga del campo
            aplicar_autosize_campos(writer)

            # Crear un buffer en memoria para el PDF
            output_buffer = io.BytesIO()
            writer.write(output_buffer)
            output_buffer.seek(0)
            
            # 2.4. Subir a Supabase Storage
            rut_estudiante = estudiante_data.get('rut', 'SinRut')
            # Usamos UUID para asegurar unicidad del archivo
            unique_id = str(uuid.uuid4())
            filename = f"Evaluacion_{form_type}_{rut_estudiante}_{unique_id}.pdf"
            
            # Ruta de Storage: formularios_completados/nomina_id/nombre_archivo.pdf
            supabase_path = f"formularios_completados/{nomina_id}/{filename}"
            supabase_url = f"{SUPABASE_URL}/storage/v1/object/formularios_completados/{nomina_id}/{filename}"
            
            # Usar requests.put para subir el archivo
            upload_res = requests.put(
                supabase_url,
                data=output_buffer.getvalue(),
                headers={
                    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}", # Usar SERVICE KEY para Storage
                    "Content-Type": "application/pdf",
                    "x-upsert": "true" 
                }
            )
            upload_res.raise_for_status()
            print(f"DEBUG: PDF subido con éxito a: {supabase_path}")
            
            return {"success": True, "path": supabase_path}

    except requests.exceptions.RequestException as e:
        message = f"Error de conexión al generar/subir PDF: {str(e)}"
        print(f"❌ ERROR: {message}")
        if 'upload_res' in locals() and upload_res.text:
             print(f"ERROR DETAIL: {upload_res.text}")
        return {"success": False, "message": message}
    except Exception as e:
        message = f"Error inesperado al generar/subir PDF: {str(e)}"
        print(f"❌ ERROR: {message}")
        return {"success": False, "message": message}
        
# Helper function to get form field values, converting None to empty string
def get_form_field_value(field_name, form_data, return_none_if_empty=False):
    """
    Retrieves a form field value from form_data.
    If return_none_if_empty is True, returns None for empty strings.
    Otherwise, returns an empty string for empty values.
    """
    value = form_data.get(field_name)
    if value is None:
        return None # If field is not present in form data at all
    
    stripped_value = value.strip()
    if not stripped_value: # If it's an empty string after stripping
        return None if return_none_if_empty else '' # Return None for dates/numeric, empty string for text/select
    return stripped_value


def aplicar_autosize_campos(writer):
    """
    Recorre todos los campos de texto del PDF y les aplica:
    - Tamaño de fuente automático (DA con tamaño 0 = auto-fit)
    - Flag Multiline activado para campos de texto largos (indicaciones, diagnóstico, etc.)
    - Flag DoNotScroll desactivado para permitir scroll dentro del campo
    Evita que el texto se corte o salga del campo cuando el contenido es largo.
    """
    # Campos que sabemos son multilinea (texto largo)
    CAMPOS_MULTILINEA = {
        'indicaciones', 'derivaciones', 'estado_general', 'diagnostico',
        'diagnostico_1', 'diagnostico_2', 'observaciones', 'observacion_neurologia',
        'motivo_consulta', 'observacion_1', 'observacion_2', 'observacion_3',
        'observacion_4', 'observacion_5', 'observacion_6', 'observacion_7',
    }
    # Flags de campo PDF
    FF_MULTILINE    = 1 << 12   # bit 13 — activa texto multilinea
    FF_DO_NOT_SCROLL = 1 << 23  # bit 24 — desactiva scroll (lo quitamos)

    try:
        if "/AcroForm" not in writer._root_object:
            return

        acroform = writer._root_object["/AcroForm"]
        if "/Fields" not in acroform:
            return

        def _procesar_campo(field_ref):
            try:
                field = field_ref.get_object()
                ft = field.get("/FT")

                if ft == "/Tx":
                    # Tamaño 0 = auto-size en lectores PDF compatibles
                    da_actual = str(field.get("/DA", "/Helv 0 Tf 0 g"))
                    import re
                    da_nuevo = re.sub(r'(\d+\.?\d*)\s+Tf', '0 Tf', da_actual)
                    if "Tf" not in da_nuevo:
                        da_nuevo = "/Helv 0 Tf 0 g"
                    field.update({NameObject("/DA"): NameObject(da_nuevo)})
                    if "/DS" in field:
                        del field[NameObject("/DS")]

                    # Obtener nombre del campo para detectar si es multilinea
                    nombre_campo = str(field.get("/T", "")).lower()
                    es_multilinea = any(m in nombre_campo for m in CAMPOS_MULTILINEA)

                    if es_multilinea:
                        # Leer flags actuales
                        ff_actual = int(field.get("/Ff", 0))
                        # Activar Multiline, desactivar DoNotScroll
                        ff_nuevo = (ff_actual | FF_MULTILINE) & ~FF_DO_NOT_SCROLL
                        field.update({NameObject("/Ff"): NumberObject(ff_nuevo)})

                # Procesar hijos (campos agrupados)
                if "/Kids" in field:
                    for kid in field["/Kids"]:
                        _procesar_campo(kid)

            except Exception as e:
                print(f"  [autosize] Warning en campo: {e}")

        fields = acroform["/Fields"]
        for field_ref in fields:
            _procesar_campo(field_ref)

        writer._root_object["/AcroForm"].update({
            NameObject("/NeedAppearances"): BooleanObject(True)
        })
        print("✅ aplicar_autosize_campos: auto-size + multiline aplicado.")

    except Exception as e:
        print(f"⚠️  aplicar_autosize_campos: error general — {e}")





def aplicar_overlay_texto_largo(pdf_bytes, campos_valores):
    """
    Solución definitiva para visores web (Chrome, Edge, iOS):
    1. Lee el PDF ya rellenado con PyPDF2
    2. VACIA el valor visible de los campos largos en AcroForm (evita texto doble)
    3. Obtiene las coordenadas exactas de cada campo largo
    4. Dibuja el texto con reportlab como gráfico fijo en esas coordenadas
    5. Fusiona el overlay con el PDF
    Si reportlab no está disponible, retorna el PDF sin cambios.
    """
    if not REPORTLAB_OK:
        return pdf_bytes

    CAMPOS_LARGOS = {
        'indicaciones', 'derivaciones', 'estado_general', 'diagnostico',
        'diagnostico_1', 'diagnostico_2', 'observaciones', 'observacion_neurologia',
        'motivo_consulta', 'observacion_1', 'observacion_2', 'observacion_3',
        'observacion_4', 'observacion_5', 'observacion_6', 'observacion_7',
    }

    def _get_rects_y_limpiar(reader):
        """Lee coords de campos largos y vacía su valor visible en AcroForm."""
        rects = {}
        try:
            from pypdf.generic import create_string_object
            root = reader.trailer["/Root"].get_object()
            acro = root.get("/AcroForm")
            if not acro:
                return rects
            ao = acro.get_object() if hasattr(acro, 'get_object') else acro

            def _proc(fref):
                try:
                    f = fref.get_object()
                    nombre = str(f.get("/T", "")).lower().strip()
                    if f.get("/FT") == "/Tx" and any(c in nombre for c in CAMPOS_LARGOS):
                        # Buscar valor en campos_valores
                        valor = None
                        for k, v in campos_valores.items():
                            if k.lower() == nombre:
                                valor = v; break
                        if not valor:
                            for k, v in campos_valores.items():
                                if nombre.startswith(k.lower()) or k.lower().startswith(nombre):
                                    valor = v; break
                        if valor and str(valor).strip():
                            rect = f.get("/Rect")
                            if rect:
                                rects[nombre] = (
                                    float(rect[0]), float(rect[1]),
                                    float(rect[2]), float(rect[3]),
                                    str(valor).strip()
                                )
                            # Vaciar valor visible del campo para evitar texto doble
                            try:
                                from pypdf.generic import create_string_object
                                f["/V"] = create_string_object("")
                                if "/AP" in f:
                                    del f["/AP"]
                            except Exception:
                                pass
                    if "/Kids" in f:
                        for kid in f["/Kids"]: _proc(kid)
                except Exception:
                    pass

            for fr in ao.get("/Fields", []):
                _proc(fr)
        except Exception as ex:
            print(f"  [overlay] get_rects: {ex}")
        return rects

    def _crear_overlay(rects, pw, ph):
        """Crea página PDF con texto dibujado en posiciones exactas."""
        buf = io.BytesIO()
        c = rl_canvas.Canvas(buf, pagesize=(pw, ph))
        for nombre, (x0, y0, x1, y1, texto) in rects.items():
            w, h = x1 - x0, y1 - y0
            margin, fs = 3, 9.0
            lh = fs * 1.35
            lines = []
            for par in texto.splitlines():
                if not par.strip():
                    lines.append('')
                else:
                    lines.extend(simpleSplit(par, "Helvetica", fs, w - 2*margin) or [''])
            # Reducir fuente si no cabe verticalmente
            while lines and len(lines) * lh > h - 2*margin and fs > 6:
                fs -= 0.5; lh = fs * 1.35
                lines = []
                for par in texto.splitlines():
                    if not par.strip():
                        lines.append('')
                    else:
                        lines.extend(simpleSplit(par, "Helvetica", fs, w - 2*margin) or [''])
            c.setFont("Helvetica", fs)
            c.setFillColorRGB(0, 0, 0)
            y_pos = y1 - margin - fs
            for line in lines:
                if y_pos < y0 + margin:
                    break
                try:
                    c.drawString(x0 + margin, y_pos, line)
                except Exception:
                    c.drawString(x0 + margin, y_pos,
                                 line.encode('latin-1', 'replace').decode('latin-1'))
                y_pos -= lh
        c.save()
        buf.seek(0)
        return buf

    try:
        from pypdf import PdfReader as _PR, PdfWriter as _PW
        reader = _PR(io.BytesIO(pdf_bytes))

        # Obtener coordenadas Y limpiar valores visibles de campos largos
        rects = _get_rects_y_limpiar(reader)
        if not rects:
            return pdf_bytes

        pg0 = reader.pages[0]
        pw, ph = float(pg0.mediabox.width), float(pg0.mediabox.height)

        # Crear overlay con el texto
        ov_buf = _crear_overlay(rects, pw, ph)
        ov_reader = _PR(ov_buf)

        # Fusionar: primero el overlay (texto limpio), luego el resto del PDF
        writer_out = _PW()
        for i, pg in enumerate(reader.pages):
            if i == 0 and ov_reader.pages:
                pg.merge_page(ov_reader.pages[0])
            writer_out.add_page(pg)

        out = io.BytesIO()
        writer_out.write(out)
        out.seek(0)
        print("✅ Overlay de texto aplicado — compatible con visores web.")
        return out.read()

    except Exception as e:
        print(f"⚠️  overlay: {e} — PDF sin overlay.")
        return pdf_bytes


def _es_masculino(sexo_val):
    """Detecta masculino sin importar cómo esté guardado en la BD."""
    v = (sexo_val or '').strip().upper()
    return v in ('M', 'MASCULINO', 'MALE', 'H', 'HOMBRE')

def _es_femenino(sexo_val):
    """Detecta femenino sin importar cómo esté guardado en la BD."""
    v = (sexo_val or '').strip().upper()
    return v in ('F', 'FEMENINO', 'FEMALE', 'MUJER')


def generar_pdf_neurologia_overlay(pdf_base_path, campos):
    """
    Exclusivo para NEUROLOGÍA.
    Usa ReportLab para texto largo + pikepdf para checkboxes de sexo.
    Los checkboxes usan Helvetica del AcroForm inyectada como recurso de página.
    """
    if not REPORTLAB_OK:
        return None

    COORDS_NEURO = {
        'nombre':             (43.0,  718.0, 346.4, 734.0),
        'rut':                (454.9, 712.5, 553.7, 734.0),
        'fecha_nacimiento':   (40.6,  690.0, 142.8, 712.0),
        'edad':               (143.9, 689.5, 245.5, 711.0),
        'nacionalidad':       (247.0, 691.5, 348.4, 712.0),
        'fecha_evaluacion':   (249.0, 568.8, 399.7, 589.7),
        'fecha_reevaluacion': (403.1, 568.2, 553.7, 591.0),
        'diagnostico_1':      (393.4, 653.0, 565.0, 671.0),
        'diagnostico_2':      (41.9,  294.3, 520.5, 316.3),
        'estado_general':     (43.2,  359.7, 557.1, 514.6),
        'derivaciones':       (41.9,  122.8, 552.5, 247.6),
    }
    COORDS_SEXO = {
        'sexo_f': (360.7, 716.5, 379.0, 731.7),
        'sexo_m': (406.5, 715.2, 426.2, 731.9),
    }
    CAMPOS_MULTILINEA = {'estado_general', 'derivaciones', 'diagnostico_2'}

    try:
        import pikepdf

        reader_base = PdfReader(pdf_base_path)
        pg0 = reader_base.pages[0]
        pw = float(pg0.mediabox.width)
        ph = float(pg0.mediabox.height)

        # ── PASO 1: ReportLab para todos los campos de TEXTO ─────────────────
        ov_buf = io.BytesIO()
        c = rl_canvas.Canvas(ov_buf, pagesize=(pw, ph))

        for campo, (x0, y0, x1, y1) in COORDS_NEURO.items():
            valor = campos.get(campo, '')
            if not valor or not str(valor).strip():
                continue
            texto = str(valor).strip()
            w = x1 - x0
            h = y1 - y0
            margin = 3

            if campo in CAMPOS_MULTILINEA:
                fs = 9.0
                lh = fs * 1.35
                lines = []
                for par in texto.splitlines():
                    if not par.strip():
                        lines.append('')
                    else:
                        lines.extend(simpleSplit(par, "Helvetica", fs, w - 2*margin) or [''])
                while lines and len(lines) * lh > h - 2*margin and fs > 6:
                    fs -= 0.5
                    lh = fs * 1.35
                    lines = []
                    for par in texto.splitlines():
                        if not par.strip():
                            lines.append('')
                        else:
                            lines.extend(simpleSplit(par, "Helvetica", fs, w - 2*margin) or [''])
                c.setFont("Helvetica", fs)
                c.setFillColorRGB(0, 0, 0)
                y_pos = y1 - margin - fs
                for line in lines:
                    if y_pos < y0 + margin:
                        break
                    try:
                        c.drawString(x0 + margin, y_pos, line)
                    except Exception:
                        c.drawString(x0 + margin, y_pos,
                                     line.encode('latin-1', 'replace').decode('latin-1'))
                    y_pos -= lh
            else:
                fs = 10.0
                while fs > 6 and c.stringWidth(texto, "Helvetica", fs) > w - 2*margin:
                    fs -= 0.5
                c.setFont("Helvetica", fs)
                c.setFillColorRGB(0, 0, 0)
                y_pos = y0 + (h - fs) / 2
                try:
                    c.drawString(x0 + margin, y_pos, texto)
                except Exception:
                    c.drawString(x0 + margin, y_pos,
                                 texto.encode('latin-1', 'replace').decode('latin-1'))

        # Dibujar X del sexo en el mismo canvas (coordenadas exactas del AP stream del PDF base)
        COORDS_SEXO_REAL = {
            'sexo_f': (361.8, 718.7, 377.8, 729.8),   # del AP stream de sexo_f
            'sexo_m': (407.6, 718.0, 424.3, 728.8),   # del AP stream de sexo_m
        }
        for fname, (x0, y0, x1, y1) in COORDS_SEXO_REAL.items():
            if not campos.get(fname, '').strip():
                continue
            w = x1 - x0
            h = y1 - y0
            fs = 9.0
            # Rectángulo blanco para tapar el fondo blanco del /Square annotation
            c.setFillColorRGB(1, 1, 1)
            c.rect(x0, y0, w, h, fill=1, stroke=0)
            # X centrada
            c.setFont("Helvetica-Bold", fs)
            c.setFillColorRGB(0, 0, 0)
            xt = x0 + (w - c.stringWidth("X", "Helvetica-Bold", fs)) / 2
            yt = y0 + (h - fs) / 2
            c.drawString(xt, yt, "X")

        c.save()
        ov_buf.seek(0)

        # Fusionar overlay encima del PDF base
        ov_reader = PdfReader(ov_buf)
        writer_out = PdfWriter()
        for i, pg in enumerate(reader_base.pages):
            if i == 0 and ov_reader.pages:
                # overlay encima: merge_page(base) sobre el overlay
                ov_pg = ov_reader.pages[0]
                ov_pg.merge_page(pg)
                writer_out.add_page(ov_pg)
            else:
                writer_out.add_page(pg)

        dst = io.BytesIO()
        writer_out.write(dst)
        print("✅ PDF neurología listo.")
        return flatten_pdf_fields(dst.getvalue())

    except Exception as e:
        print(f"❌ generar_pdf_neurologia_overlay: {e}")
        return None


def flatten_pdf_fields(pdf_bytes):
    try:
        import pikepdf
        src = io.BytesIO(pdf_bytes)
        dst = io.BytesIO()
        with pikepdf.open(src) as pdf:
            if "/AcroForm" in pdf.Root:
                del pdf.Root["/AcroForm"]
            for page in pdf.pages:
                if "/Annots" in page:
                    # Eliminar Widget Y Square (los /Square tapan con fondo blanco)
                    page["/Annots"] = pikepdf.Array([
                        a for a in page["/Annots"]
                        if a.get("/Subtype") not in ("/Widget", "/Square")
                    ])
            pdf.save(dst)
        dst.seek(0)
        return dst.read()
    except Exception as e:
        print(f"⚠️ flatten_pdf_fields: {e}")
        return pdf_bytes


def wrap_texto_pdf(texto, chars_por_linea=85):
    """
    Inserta saltos de linea automaticos en texto largo para que no se corte
    al escribirlo en campos PDF de ancho fijo.
    Respeta saltos existentes. chars_por_linea=85 calibrado para el campo
    indicaciones del formulario neurologico.
    """
    if not texto:
        return texto or ""
    import textwrap
    lineas = []
    for linea in texto.splitlines():
        if len(linea) <= chars_por_linea:
            lineas.append(linea)
        else:
            lineas.append(textwrap.fill(linea, width=chars_por_linea,
                break_long_words=True, break_on_hyphens=True))
    return "\n".join(lineas)

# Nuevo: Función para obtener el PDF de neurología específico para una doctora
def get_doctor_specific_neurologia_pdf(doctora_id):
    """
    Intenta encontrar un PDF de neurología específico para la doctora en el directorio configurado.
    Si no lo encuentra, retorna el PDF de neurología por defecto.
    """
    # Construir la ruta completa al directorio de bases de PDF
    base_dir = os.path.dirname(os.path.abspath(__file__)) # Obtiene el directorio del script actual
    full_pdf_bases_dir_path = os.path.join(base_dir, PDF_BASES_NEUROLOGIA_DIR)

    # Corregido: Asume que los PDFs están nombrados como 'FORMULARIO TIPO NEUROLOGIA_{doctora_id}.pdf'
    specific_pdf_filename = f"FORMULARIO TIPO NEUROLOGIA_{doctora_id}.pdf"
    specific_pdf_path = os.path.join(full_pdf_bases_dir_path, specific_pdf_filename)

    print(f"DEBUG: Buscando PDF en la ruta absoluta: {specific_pdf_path}") # Añadido para depuración

    if os.path.exists(specific_pdf_path):
        print(f"DEBUG: Se encontró PDF específico para doctora {doctora_id}: {specific_pdf_path}")
        return specific_pdf_path
    else:
        print(f"ADVERTENCIA: No se encontró PDF específico para doctora {doctora_id} en {specific_pdf_path}. Usando PDF por defecto: {PDF_BASE_NEUROLOGIA}")
        # Fallback al PDF por defecto, asegurando que su ruta también sea absoluta para claridad
        default_pdf_path = os.path.join(base_dir, PDF_BASE_NEUROLOGIA)
        print(f"DEBUG: Usando PDF por defecto en la ruta absoluta: {default_pdf_path}")
        return default_pdf_path

# Coloca esta función en la sección de utilidades de app-38.py:
def map_db_value_to_x(db_value):
    """Convierte el valor True/False o string de la base de datos a 'X' o ''."""
    if db_value is True or (isinstance(db_value, str) and db_value.strip()):
        return "X"
    return ""

# -------------------- Rutas de la Aplicación --------------------

# app-30.py (Reemplaza la función relleno_formulario completa)
# app-32.py (REEMPLAZO COMPLETO DE relleno_formulario - BASADO EN tipo_nomina)

@app.route('/relleno_formulario/<string:nomina_id>', methods=['GET'])
def relleno_formulario(nomina_id):
    # Asumo que las importaciones (datetime, requests, etc.) y funciones (calculate_age, etc.) existen

    if 'usuario' not in session:
        return redirect(url_for('index'))

    user_role = session.get('usuario')
    user_id = session.get('usuario_id')
    
    # Guardamos el ID de la nómina en la sesión para que esté disponible en otras rutas
    session['current_nomina_id'] = nomina_id 

    # 1. Obtener detalles de la nómina
    url_nomina = (
        f"{SUPABASE_URL}/rest/v1/nominas_medicas"
        f"?id=eq.{nomina_id}"
        f"&select=form_type,tipo_nomina,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,nombre_nomina,doctora_id_para_formulario,token_acceso,fecha_evaluacion_fija"
    )
    
    try:
        res_nomina = requests.get(url_nomina, headers=SUPABASE_SERVICE_HEADERS) 
        res_nomina.raise_for_status()
        nomina_data = res_nomina.json()
        
        if not nomina_data:
            flash(f'Error: Nómina con ID {nomina_id} no encontrada.', 'error')
            return redirect(url_for('dashboard'))
        
        nomina = nomina_data[0]
        
        form_type = nomina['form_type']
        
        # Guardar form_type y doctora_id_para_formulario en la sesión (CRÍTICO para otras rutas)
        session['current_form_type'] = form_type
        session['doctora_id_para_formulario'] = nomina.get('doctora_id_para_formulario')
        
        # Validación de acceso — soporta nómina compartida (doctora principal O segunda doctora)
        if user_role == 'doctora':
            doctora_1 = nomina.get('doctora_id')
            doctora_2 = nomina.get('doctora_id_2')
            doctora_3 = nomina.get('doctora_id_3')
            doctora_4 = nomina.get('doctora_id_4')
            if user_id not in [doctora_1, doctora_2, doctora_3, doctora_4]:
                flash('Acceso no autorizado a esta nómina.', 'error')
                return redirect(url_for('dashboard'))

    except requests.exceptions.RequestException as e:
        error_detail = e.response.text if e.response is not None else 'Conexión Fallida'
        print(f"❌ ERROR al obtener detalles de la nómina {nomina_id}: {e}. Detalle: {error_detail}")
        flash(f'Error al cargar la nómina. Detalle: {error_detail}', 'error')
        return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"❌ ERROR Inesperado al procesar detalles de la nómina: {e}")
        flash('Error interno del servidor.', 'error')
        return redirect(url_for('dashboard'))

    # 2. Obtener la lista de estudiantes con TODOS los campos de evaluación
    url_estudiantes = (
        f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
        f"?nomina_id=eq.{nomina_id}"
        f"&select=id,nombre,rut,fecha_nacimiento,nacionalidad,sexo,estado_general,diagnostico,derivaciones,fecha_evaluacion,fecha_reevaluacion,fecha_relleno,diagnostico_1,diagnostico_2,diagnostico_complementario,clasificacion,observacion_1,observacion_2,observacion_3,observacion_4,observacion_5,observacion_6,observacion_7,check_cesarea,check_atermino,check_vaginal,check_prematuro,check_acorde,check_retrasogeneralizado,check_esquemac,check_esquemai,check_alergiano,check_alergiasi,check_cirugiano,check_cirugiasi,check_retraso,check_visionsinalteracion,check_visionrefraccion,check_audicionnormal,check_hipoacusia,check_tapondecerumen,check_sinhallazgos,check_caries,check_apinamientodental,check_retenciondental,check_frenillolingual,check_hipertrofia,altura,peso,imc,indicaciones,fecha_reevaluacion_select,motivo_consulta,observacion_neurologia,observaciones,diagnostico_sospecha,diagnostico_definitivo,estado_asistencia,motivo_ausencia" 
        f"&order=nombre.asc"
    )

    try:
        res_estudiantes = requests.get(url_estudiantes, headers=SUPABASE_SERVICE_HEADERS) 
        res_estudiantes.raise_for_status()
        estudiantes_raw = res_estudiantes.json()
        
        # 3. Preparar los datos de estudiantes para el template
        estudiantes = []
        for est_raw in estudiantes_raw:
            est = est_raw.copy()
            
            # Procesamiento de Fechas y Edad
            fecha_nacimiento_obj = None
            if est.get('fecha_nacimiento') and est['fecha_nacimiento'].strip():
                try:
                    fecha_nacimiento_obj = datetime.strptime(est['fecha_nacimiento'], '%Y-%m-%d').date()
                except:
                    pass

            edad_calculada = "N/A"
            if fecha_nacimiento_obj:
                # Usar fecha_evaluacion como referencia si existe, si no hoy
                fecha_eval_obj = None
                if est.get('fecha_evaluacion'):
                    try:
                        fecha_eval_obj = datetime.strptime(est['fecha_evaluacion'], '%Y-%m-%d').date()
                    except:
                        pass
                edad_calculada = calculate_age(fecha_nacimiento_obj, fecha_ref=fecha_eval_obj)

            est['edad'] = edad_calculada
            # 💡 Aseguramos que el formato de fecha de nacimiento siempre esté disponible:
            est['fecha_nacimiento_formato'] = fecha_nacimiento_obj.strftime("%d/%m/%Y") if fecha_nacimiento_obj else 'N/A'
            est['fecha_nacimiento'] = est.get('fecha_nacimiento', '')
            est['fecha_evaluacion'] = est.get('fecha_evaluacion', '')
            est['fecha_reevaluacion'] = est.get('fecha_reevaluacion', '')
            
            # 💡 Corrección del Mapeo de check_cirugiasi / si_2 (Medicina Familiar)
            # Aunque ya no está en el formulario, es necesario para cargar datos antiguos de la DB.
            if est.get('si_2'):
                est['check_cirugiasi'] = est.get('si_2') 
                
            # Mapeo de género (sexo) a los checkboxes (Medicina Familiar)
            sexo_db = (est.get('sexo') or "").upper()
            est['genero_f'] = (sexo_db == 'F')
            est['genero_m'] = (sexo_db == 'M')
            
            estudiantes.append(est)
        
        # 4. Obtener la doctora asignada
        doctora_asignada_id = nomina['doctora_id']
        url_doctora = f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{doctora_asignada_id}&select=nombre"
        res_doctora = requests.get(url_doctora, headers=SUPABASE_SERVICE_HEADERS) 
        doctora_nombre = res_doctora.json()[0]['nombre'] if res_doctora.ok and res_doctora.json() else 'Doctora Asignada'
        
        # Total de formularios completados
        # Solo contar evaluados que están activos o son extra (excluir no_asiste)
        total_forms_completed_for_nomina = sum(1 for est in estudiantes if est.get('fecha_relleno') is not None and est.get('estado_asistencia') in ('activo','extra',None,''))


        # 5. LÓGICA DE REDIRECCIÓN CLAVE (UNIFICADA POR form_type)
        # Guardar fecha_evaluacion_fija en sesión para que marcar_evaluado la use
        session['fecha_evaluacion_fija'] = nomina.get('fecha_evaluacion_fija') or None

        base_render_params = {
            'nomina_id': nomina_id,
            'establecimiento_nombre': nomina['nombre_nomina'],
            'form_type': form_type, 
            'estudiantes': estudiantes,
            'total_forms_completed_for_nomina': total_forms_completed_for_nomina,
            'doctora_asignada_id': doctora_asignada_id,
            'doctora_nombre': doctora_nombre,
            'usuario': user_role,
            'token_acceso': nomina.get('token_acceso') or '',
            'fecha_evaluacion_fija': nomina.get('fecha_evaluacion_fija') or None,
        }
        
        # Redirección basada en la columna 'form_type'
        if form_type == 'informe_neurologico':
            # Usa el HTML para el nuevo informe
            return render_template('formulario_informe_neurologico.html', **base_render_params)

        elif form_type == 'medicina_familiar':
            # Usa el HTML de Medicina Familiar
            return render_template('formulario_medicina_familiar.html', **base_render_params)
        
        elif form_type == 'neurologia':
            # Usa el HTML de Neurología
            return render_template('formulario_relleno.html', **base_render_params)
        
        else:
            flash(f'❌ El tipo de formulario "{form_type.capitalize()}" no se pudo mapear a un formulario conocido.', 'error')
            return redirect(url_for('dashboard'))

    except requests.exceptions.RequestException as e:
        error_detail = e.response.text if e.response is not None else 'Conexión Fallida'
        print(f"❌ ERROR al obtener estudiantes para nómina {nomina_id}: {e}. Detalle: {error_detail}")
        flash(f'Error al cargar la lista de estudiantes. Revise si todas las columnas del SELECT existen en su tabla estudiantes_nomina.', 'error')
        return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"❌ ERROR Inesperado en relleno_formulario: {e}")
        flash('Error interno del servidor. Detalle: ' + str(e), 'error')
        return redirect(url_for('dashboard'))
        

# app.py (REEMPLAZO FINAL Y CORREGIDO DE generar_pdf - INTEGRIDAD DEL PDF)
# Asegúrate de que todas las demás importaciones necesarias (flask, get_form_field_value, format_rut_python, etc.) 
# estén definidas al inicio de tu app.py

# REEMPLAZA COMPLETAMENTE ESTA FUNCIÓN EN TU app.py
@app.route('/generar_pdf', methods=['POST'])
def generar_pdf():
    if 'usuario' not in session:
        return redirect(url_for('index'))

    estudiante_id = request.form.get('estudiante_id')
    nomina_id = request.form.get('nomina_id')
    
    form_type = session.get('current_form_type', 'neurologia') 
    current_doctora_id = session.get('usuario_id')
    
    print(f"DEBUG: generar_pdf - Solicitud para generar PDF para estudiante_id={estudiante_id}, nomina_id={nomina_id}, form_type={form_type}, doctora_id={current_doctora_id}")

    if not all([estudiante_id, nomina_id]):
        flash("❌ Faltan datos esenciales del formulario para generar PDF.", 'danger')
        if 'current_nomina_id' in session:
            return redirect(url_for('relleno_formulario', nomina_id=session['current_nomina_id']))
        return redirect(url_for('dashboard'))

    # --- 1. Obtener datos de la nómina para el ID de la Doctora Firmante ---
    doctora_id_para_pdf = current_doctora_id 
    try:
        nomina_data = obtener_nomina_por_id(nomina_id)
        doctora_id_para_pdf = nomina_data.get('doctora_id_para_formulario') if nomina_data.get('doctora_id_para_formulario') else current_doctora_id
    except Exception as e:
        print(f"ADVERTENCIA: No se pudo obtener el ID de la doctora firmante. Usando ID de sesión. {e}")
        
    # --- 2. Procesamiento de Campos Comunes y Fechas ---
    nombre = get_form_field_value('nombre', request.form)
    rut = format_rut_python(get_form_field_value('rut', request.form))
    
    fecha_nac_formato = ''
    # ✅ FIX: leer 'fecha_nacimiento' primero (nombre nuevo), con fallback a 'fecha_nacimiento_original' (nombre viejo)
    fecha_nac_original_str = (
        get_form_field_value('fecha_nacimiento', request.form)
        or get_form_field_value('fecha_nacimiento_original', request.form)
    )
    if fecha_nac_original_str:
        try:
            fecha_nac_formato = datetime.strptime(fecha_nac_original_str, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            pass 

    edad = get_form_field_value('edad', request.form)
    nacionalidad = get_form_field_value('nacionalidad', request.form)

    # Funciones auxiliares
    def map_check_value(field_name):
        return get_form_field_value(field_name, request.form) or ""
    
    # Función para escribir "X" en campos de texto que actúan como casillas
    def map_check_as_text(field_name):
        val = request.form.get(field_name)
        if val and val.strip() != "":
            return "X"
        return ""

    # Mapeos de Género (como texto "X")
    sexo_f_pdf = "X" if get_form_field_value('genero_f', request.form) or get_form_field_value('sexo', request.form) == 'F' else ""
    sexo_m_pdf = "X" if get_form_field_value('genero_m', request.form) or get_form_field_value('sexo', request.form) == 'M' else ""

    # Formatos de Fecha
    fecha_evaluacion_form_value = get_form_field_value('fecha_evaluacion', request.form)
    fecha_evaluacion_formatted = ''
    if fecha_evaluacion_form_value:
        try:
            fecha_evaluacion_formatted = datetime.strptime(fecha_evaluacion_form_value, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            pass

    fecha_reevaluacion_form_value = get_form_field_value('fecha_reevaluacion', request.form)
    fecha_reeval_pdf = ''
    if fecha_reevaluacion_form_value:
        try:
            fecha_reeval_pdf = datetime.strptime(fecha_reevaluacion_form_value, '%Y-%m-%d').strftime('%d/%m/%Y')
        except ValueError:
            pass
            
    derivaciones = get_form_field_value('derivaciones', request.form)

    # --- 3. Selección de Plantilla ---
    base_dir = os.path.dirname(os.path.abspath(__file__))
    pdf_base_path = ''
    
    if form_type == 'neurologia':
        specific_pdf_filename = f"FORMULARIO TIPO NEUROLOGIA_{doctora_id_para_pdf}.pdf"
        full_pdf_bases_dir_path = os.path.join(base_dir, PDF_BASES_NEUROLOGIA_DIR)
        specific_pdf_path = os.path.join(full_pdf_bases_dir_path, specific_pdf_filename)
        pdf_base_path = specific_pdf_path if os.path.exists(specific_pdf_path) else os.path.join(base_dir, PDF_BASE_NEUROLOGIA)
            
    elif form_type == 'informe_neurologico':
        specific_pdf_filename = f"INFORME_NEUROLOGICO_BASE_{doctora_id_para_pdf}.pdf"
        pdf_bases_dir = os.path.join(base_dir, PDF_BASES_NEUROLOGIA_DIR)
        specific_pdf_path = os.path.join(pdf_bases_dir, specific_pdf_filename)
        pdf_base_path = specific_pdf_path if os.path.exists(specific_pdf_path) else os.path.join(base_dir, PDF_BASE_INFORME_NEURO)
    
    elif form_type == 'medicina_familiar':
        # Soporte para PDF personalizado por doctora — fallback al base si no existe
        specific_pdf_filename_fam = f"FORMULARIO_FAMILIAR_{doctora_id_para_pdf}.pdf"
        full_pdf_bases_familiar_dir = os.path.join(base_dir, PDF_BASES_FAMILIAR_DIR)
        specific_pdf_path_fam = os.path.join(full_pdf_bases_familiar_dir, specific_pdf_filename_fam)
        pdf_base_path = specific_pdf_path_fam if os.path.exists(specific_pdf_path_fam) else os.path.join(base_dir, PDF_BASE_FAMILIAR)

    # --- 4. Lógica de Relleno y Generación ---
    try:
        reader = PdfReader(pdf_base_path)
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)
            
        campos = {}
        
        # 🟢 NEUROLOGÍA: SE MANTIENE TU LÓGICA ORIGINAL COMPLETA
        if form_type == 'neurologia':
            print("DEBUG: Usando mapeo original de Neurología.")
            campos = {
                "nombre": nombre,
                "rut": rut,
                "fecha_nacimiento": fecha_nac_formato,
                "nacionalidad": nacionalidad,
                "edad": edad,
                "diagnostico_1": get_form_field_value('diagnostico', request.form),
                "diagnostico_2": get_form_field_value('diagnostico', request.form), 
                "estado_general": wrap_texto_pdf(get_form_field_value('estado', request.form)),
                "fecha_evaluacion": fecha_evaluacion_formatted,
                "fecha_reevaluacion": fecha_reeval_pdf,
                "derivaciones": wrap_texto_pdf(derivaciones),
                "sexo_f": sexo_f_pdf,
                "sexo_m": sexo_m_pdf,
            }
        
        # 🟢 INFORME NEUROLÓGICO: SE MANTIENE TU LÓGICA ORIGINAL COMPLETA
        elif form_type == 'informe_neurologico':
            campos = {
                "nombre": nombre, "rut": rut, "fecha_nacimiento": fecha_nac_formato, 
                "edad": edad, "genero_m": sexo_m_pdf, "genero_f": sexo_f_pdf, 
                "nacionalidad": nacionalidad,
                "motivo_consulta": wrap_texto_pdf(get_form_field_value('motivo_consulta', request.form)),
                "observaciones": wrap_texto_pdf(get_form_field_value('observaciones', request.form)),      
                "observacion_neurologia": wrap_texto_pdf(get_form_field_value('observacion_neurologia', request.form)), 
                "diagnostico": wrap_texto_pdf(get_form_field_value('diagnostico', request.form)),
                "indicaciones": wrap_texto_pdf(get_form_field_value('indicaciones', request.form)),        
                "derivaciones": wrap_texto_pdf(derivaciones), 
                "fecha_evaluacion": fecha_evaluacion_formatted,
                "fecha_reevaluacion": fecha_reeval_pdf,
            }
            
        # 🟢 MEDICINA FAMILIAR: CORRECCIÓN DE LA "X" EN CAMPOS DE TEXTO
        elif form_type == 'medicina_familiar':
            diagnostico_unificado_valor = get_form_field_value('diagnostico_unificado', request.form)
            campos = {
                "nombre": nombre, "rut": rut, "fecha_nacimiento": fecha_nac_formato, "edad": edad, "nacionalidad": nacionalidad,
                "sexo_f": sexo_f_pdf, "sexo_m": sexo_m_pdf,
                "diagnostico_1": diagnostico_unificado_valor, "diagnostico_2": diagnostico_unificado_valor, 
                "diagnostico_complementario": get_form_field_value('diagnostico_complementario', request.form),
                "clasificacion": get_form_field_value('clasificacion_imc', request.form),
                "indicaciones": get_form_field_value('indicaciones', request.form), "derivaciones": derivaciones, 
                "fecha_evaluacion": fecha_evaluacion_formatted, "fecha_reevaluacion": fecha_reeval_pdf,
                "altura": get_form_field_value('altura', request.form), "peso": get_form_field_value('peso', request.form), "imc": get_form_field_value('imc', request.form),
                "observacion_1": get_form_field_value('observacion_1', request.form), "observacion_2": get_form_field_value('observacion_2', request.form),
                "observacion_3": get_form_field_value('observacion_3', request.form), "observacion_4": get_form_field_value('observacion_4', request.form),
                "observacion_5": get_form_field_value('observacion_5', request.form), "observacion_6": get_form_field_value('observacion_6', request.form),
                "observacion_7": get_form_field_value('observacion_7', request.form),
                
                # Campos "check" que son de texto: escribimos "X" si están marcados
                "check_cesarea": map_check_as_text('check_cesarea'),
                "check_atermino": map_check_as_text('check_atermino'),
                "check_vaginal": map_check_as_text('check_vaginal'),
                "check_prematuro": map_check_as_text('check_prematuro'),
                "check_acorde": map_check_as_text('check_acorde'),
                "check_retraso": map_check_as_text('check_retraso'),
                "check_retrasogeneralizado": map_check_as_text('check_retrasogeneralizado'),
                "check_esquemac": map_check_as_text('check_esquemac'),
                "check_esquemai": map_check_as_text('check_esquemai'),
                "check_alergiano": map_check_as_text('check_alergiano'),
                "check_alergiasi": map_check_as_text('check_alergiasi'),
                "check_cirugiano": map_check_as_text('check_cirugiano'),
                "check_cirugiasi": map_check_as_text('check_cirugiasi'),
                "check_visionsinalteracion": map_check_as_text('check_visionsinalteracion'),
                "check_visionrefraccion": map_check_as_text('check_visionrefraccion'),
                "check_audicionnormal": map_check_as_text('check_audicionnormal'),
                "check_hipoacusia": map_check_as_text('check_hipoacusia'),
                "check_tapondecerumen": map_check_as_text('check_tapondecerumen'),
                "check_sinhallazgos": map_check_as_text('check_sinhallazgos'),
                "check_caries": map_check_as_text('check_caries'),
                "check_apinamientodental": map_check_as_text('check_apinamientodental'),
                "check_retenciondental": map_check_as_text('check_retenciondental'),
                "check_frenillolingual": map_check_as_text('check_frenillolingual'),
                "check_hipertrofia": map_check_as_text('check_hipertrofia'),
            }

        # Aplicar el relleno
        if "/AcroForm" not in writer._root_object:
            writer._root_object.update({NameObject("/AcroForm"): DictionaryObject()})
            
        for page in writer.pages:
            writer.update_page_form_field_values(page, campos)

        writer._root_object["/AcroForm"].update({NameObject("/NeedAppearances"): BooleanObject(True)})

        # Auto-size y apariencias visuales para compatibilidad con visores web
        aplicar_autosize_campos(writer)
        # overlay aplicado después de escribir PDF
        
        output = io.BytesIO()
        writer.write(output)
        output.seek(0)
        pdf_final = aplicar_overlay_texto_largo(output.read(), campos)


        nombre_descarga = f"{nombre.replace(' ', '_')}_{rut}_formulario_{form_type}.pdf"
        return send_file(io.BytesIO(pdf_final), as_attachment=True, download_name=nombre_descarga, mimetype='application/pdf')

    except Exception as e:
        print(f"❌ Error al generar PDF: {e}")
        flash(f"❌ Error al generar el PDF: {e}", 'error')
        return redirect(url_for('dashboard'))
        



# ─────────────────────────────────────────────────────────────────────────────
#  CORRECCIÓN EN PLATAFORMA (Admin)
#  GET  /api/admin/alumno_datos/<alumno_id>   — carga datos del alumno para editar
#  POST /api/admin/corregir_alumno            — guarda cambios y marca solicitud resuelta
# ─────────────────────────────────────────────────────────────────────────────

@app.route('/api/admin/alumno_datos/<alumno_id>', methods=['GET'])
def api_admin_alumno_datos(alumno_id):
    """Devuelve todos los campos editables del alumno + form_type de su nómina."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        url = (f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
               f"?id=eq.{alumno_id}"
               f"&select=*")
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        if not res.ok or not res.json():
            return jsonify({"success": False, "message": "Alumno no encontrado"}), 404
        alumno = res.json()[0]

        # Obtener form_type y nombre de la nómina
        nomina_id = alumno.get('nomina_id')
        form_type = None
        nom_nombre = None
        if nomina_id:
            r_n = requests.get(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}&select=form_type,nombre_nomina,nombre_colegio",
                headers=SUPABASE_SERVICE_HEADERS)
            if r_n.ok and r_n.json():
                nd = r_n.json()[0]
                form_type  = nd.get('form_type')
                nom_nombre = nd.get('nombre_colegio') or nd.get('nombre_nomina')

        return jsonify({"success": True, "alumno": alumno,
                        "form_type": form_type, "nom_nombre": nom_nombre})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/admin/corregir_alumno', methods=['POST'])
def api_admin_corregir_alumno():
    """
    Guarda las correcciones de un alumno hechas por el admin en el dashboard.
    Marca la solicitud de corrección como Aprobada.
    No genera PDF aquí — el coordinador descargará desde /descargar_pdf_alumno/<id>.
    """
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        data        = request.get_json()
        alumno_id   = data.get('alumno_id')
        solicitud_id = data.get('solicitud_id')
        campos      = data.get('campos', {})   # dict con los campos a actualizar

        if not alumno_id or not campos:
            return jsonify({"success": False, "message": "Faltan datos"}), 400

        # Campos permitidos para edición por admin (todos los editables del formulario)
        CAMPOS_PERMITIDOS = {
            # Datos personales
            'nombre','rut','fecha_nacimiento','fecha_evaluacion','fecha_reevaluacion',
            'nacionalidad','sexo','edad',
            # Neurología / Informe neurológico
            'estado_general','diagnostico','derivaciones',
            'motivo_consulta','observaciones','observacion_neurologia','indicaciones',
            'diagnostico_sospecha','diagnostico_definitivo',
            # Medicina familiar — diagnósticos
            'diagnostico_1','diagnostico_2','diagnostico_complementario',
            'clasificacion','observacion_1','observacion_2','observacion_3',
            'observacion_4','observacion_5','observacion_6','observacion_7',
            # Medicina familiar — checkboxes
            'check_cesarea','check_atermino','check_vaginal','check_prematuro',
            'check_acorde','check_retraso','check_retrasogeneralizado','check_esquemac','check_esquemai',
            'check_alergiano','check_alergiasi','check_cirugiano','check_cirugiasi',
            'check_visionsinalteracion','check_visionrefraccion','check_audicionnormal',
            'check_hipoacusia','check_tapondecerumen','check_sinhallazgos',
            'check_caries','check_apinamientodental','check_retenciondental',
            'check_frenillolingual','check_hipertrofia',
            # Medicina familiar — medidas
            'altura','peso','imc','clasificacion_imc',
            'fecha_reevaluacion_select','motivo_consulta',
            # Solo editable desde correcciones admin
            'deficit',
        }

        payload = {k: v for k, v in campos.items() if k in CAMPOS_PERMITIDOS}
        if not payload:
            return jsonify({"success": False, "message": "No hay campos válidos para actualizar"}), 400

        # PATCH en estudiantes_nomina
        # Limpiar payload: convertir strings vacíos a None, nunca guardar el string "None"
        payload = {
            k: (None if v in (None, '', 'None', 'null') else v)
            for k, v in payload.items()
        }
        r_patch = requests.patch(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{alumno_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=minimal"},
            json=payload)
        if not r_patch.ok:
            return jsonify({"success": False,
                            "message": f"Error al guardar: {r_patch.text}"}), 500

        # Marcar solicitud como Aprobada si se pasó el ID
        if solicitud_id:
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?id=eq.{solicitud_id}",
                headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=minimal"},
                json={"estado": "Aprobada", "fecha_resolucion": str(date.today()),
                      "respuesta_admin": (data.get('respuesta_admin') or '').strip() or None,
                      "notificacion_vista": False})

        return jsonify({"success": True,
                        "message": "Corrección guardada. La coordinadora ya puede ver los cambios."})

    except Exception as e:
        print(f"ERROR api_admin_corregir_alumno: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/marcar_evaluado', methods=['POST'])
def marcar_evaluado():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    estudiante_id = request.form.get('estudiante_id')
    
    # --- CORRECCIÓN CLAVE: Fallback si nomina_id no viene en el formulario (era "") ---
    nomina_id = request.form.get('nomina_id')
    if not nomina_id:
        nomina_id = session.get('current_nomina_id')
    # ----------------------------------------------------------------------------------
    
    doctora_id = session.get('usuario_id')
    form_type = session.get('current_form_type', 'neurologia') 
    
    # Nota: Debes tener get_form_field_value y date importados
    nombre = get_form_field_value('nombre', request.form)
    rut = get_form_field_value('rut', request.form)

    print(f"DEBUG: Recibida solicitud para marcar como evaluado: estudiante_id={estudiante_id}, nomina_id={nomina_id}, doctora_id={doctora_id}, form_type={form_type}")
    print(f"DEBUG: Contenido completo de request.form: {request.form.to_dict()}")

    # Validación básica de datos obligatorios
    if not all([estudiante_id, nomina_id, doctora_id]):
        print(f"ERROR: Datos faltantes en /marcar_evaluado. Estudiante ID: {estudiante_id}, Nomina ID: {nomina_id}, Doctora ID: {doctora_id}. Campos del formulario: {request.form.to_dict()}")
        return jsonify({"success": False, "message": "Faltan datos obligatorios para marcar y guardar la evaluación."}), 400

    # --- 1. DATOS BASE (Comunes a todos los formularios) ---
    # Si la nómina tiene fecha_evaluacion_fija → usarla siempre (inamovible)
    # Si no → usar la que envía el formulario (comportamiento actual)
    fecha_evaluacion_fija = session.get('fecha_evaluacion_fija')
    fecha_evaluacion_final = fecha_evaluacion_fija if fecha_evaluacion_fija else get_form_field_value('fecha_evaluacion', request.form, return_none_if_empty=True)

    # Si hay fecha fija, recalcular edad con esa fecha como referencia
    edad_final = get_form_field_value('edad', request.form)
    if fecha_evaluacion_fija:
        fn_str = get_form_field_value('fecha_nacimiento', request.form, return_none_if_empty=True)
        if fn_str:
            try:
                from datetime import date as _dt_date
                fn_obj  = _dt_date.fromisoformat(fn_str)
                ref_obj = _dt_date.fromisoformat(fecha_evaluacion_fija)
                edad_final = calculate_age(fn_obj, fecha_ref=ref_obj)
            except Exception:
                pass  # si falla, deja la edad que vino del formulario

    update_data = {
        'fecha_relleno': str(date.today()),
        'doctora_evaluadora_id': doctora_id, 
        'nombre': nombre,
        'rut': rut, 
        'fecha_nacimiento': get_form_field_value('fecha_nacimiento', request.form, return_none_if_empty=True), 
        'fecha_evaluacion': fecha_evaluacion_final,
        'fecha_reevaluacion': get_form_field_value('fecha_reevaluacion', request.form, return_none_if_empty=True),
        'edad': edad_final, 
        'nacionalidad': get_form_field_value('nacionalidad', request.form), 
        'sexo': get_form_field_value('sexo', request.form),
        'evaluado_flag': True,
    }

    # --- 2. LÓGICA PARA CAMPOS ESPECÍFICOS ---
    if form_type == 'neurologia':
        # Campos específicos de Neurología (Antiguo) se añaden a update_data
        update_data.update({
            'estado_general': get_form_field_value('estado', request.form),
            'diagnostico': get_form_field_value('diagnostico', request.form), 
            'derivaciones': get_form_field_value('derivaciones', request.form),
        })
    
    # 🟢 CORRECCIÓN: Lógica para guardar el Informe Neurológico Individual
    elif form_type == 'informe_neurologico':
         update_data.update({
             # Campos de Evaluación Confirmados (5 campos)
             'motivo_consulta': get_form_field_value('motivo_consulta', request.form),
             'observaciones': get_form_field_value('observaciones', request.form),
             'observacion_neurologia': get_form_field_value('observacion_neurologia', request.form),
             'diagnostico': get_form_field_value('diagnostico', request.form), 
             'indicaciones': get_form_field_value('indicaciones', request.form),
             
             # Nota: Los campos que no se usan en este formulario (como diagnostico_sospecha o historia_actual) se omiten o se dejan en NULL.
             # Para evitar errores en otros procesos (como la generación del PDF), 
             # rellenamos los campos relacionados si es necesario, pero solo con la data existente.
             # Si el PDF necesita 'derivaciones', debemos agregarlo al formulario HTML, pero aquí solo guardamos lo que el HTML envía.
         })
         
    elif form_type == 'medicina_familiar':
        
        # OBTENEMOS EL VALOR UNIFICADO DEL CAMPO DIAGNOSTICO
        diagnostico_unificado_valor = get_form_field_value('diagnostico_unificado', request.form)

        # FUNCIÓN AUXILIAR PARA MAPEO BOOLEANO
        # Si el campo tiene valor → True (checkbox marcado)
        # Si el campo viene vacío → False (checkbox desmarcado, JS envía '')
        # Si el campo NO viene en el form → None (no tocar en BD)
        def map_to_boolean(field_name):
            raw = request.form.get(field_name)
            if raw is None:
                return None      # campo ausente → no modificar
            if raw.strip():
                return True      # tiene valor → marcado
            return False         # string vacío → desmarcado
            
        # 💡 CORRECCIÓN 1: Capturar y mapear el género (sexo) para guardar en el campo maestro 'sexo'
        genero_f_check = get_form_field_value('genero_f', request.form)
        genero_m_check = get_form_field_value('genero_m', request.form)
        sexo_final = None
        if genero_f_check:
            sexo_final = 'F'
        elif genero_m_check:
            sexo_final = 'M'
        # Sobreescribir el campo 'sexo' en el payload base con el valor final
        update_data['sexo'] = sexo_final
        # -----------------------------------------------------------------------------------------


        # Campos específicos de Medicina Familiar
        update_data.update({
            # Diagnósticos y Derivaciones
            'diagnostico_1': diagnostico_unificado_valor,
            # Limpiar dx_previo solo si se guardó un valor válido del desplegable
            'diagnostico_sospecha': ('' if diagnostico_unificado_valor and diagnostico_unificado_valor.strip() else None),
            'diagnostico_2': diagnostico_unificado_valor,
            'diagnostico_complementario': get_form_field_value('diagnostico_complementario', request.form),
            'clasificacion': get_form_field_value('clasificacion_imc', request.form),
            'derivaciones': get_form_field_value('derivaciones', request.form),
            
            # Mapeo Corregido: Guardado del campo 'indicaciones'
            'indicaciones': get_form_field_value('indicaciones', request.form), 
            
            # 💡 CORRECCIÓN 3: Guardado del valor del SELECT de años
            'fecha_reevaluacion_select': get_form_field_value('fecha_reevaluacion_select', request.form, return_none_if_empty=True),
            
            # Observaciones
            'observacion_1': get_form_field_value('observacion_1', request.form),
            'observacion_2': get_form_field_value('observacion_2', request.form),
            'observacion_3': get_form_field_value('observacion_3', request.form),
            'observacion_4': get_form_field_value('observacion_4', request.form),
            'observacion_5': get_form_field_value('observacion_5', request.form),
            'observacion_6': get_form_field_value('observacion_6', request.form),
            'observacion_7': get_form_field_value('observacion_7', request.form),

            # Checkboxes y Numéricos - CRÍTICO: USAR map_to_boolean para booleanos
            'check_cesarea': map_to_boolean('check_cesarea'),
            'check_atermino': map_to_boolean('check_atermino'),
            'check_vaginal': map_to_boolean('check_vaginal'),
            'check_prematuro': map_to_boolean('check_prematuro'),
            'check_acorde': map_to_boolean('check_acorde'),
            'check_retraso': map_to_boolean('check_retraso'), 
            'check_retrasogeneralizado': map_to_boolean('check_retrasogeneralizado'),
            'check_esquemac': map_to_boolean('check_esquemac'),
            'check_esquemai': map_to_boolean('check_esquemai'),
            'check_alergiano': map_to_boolean('check_alergiano'),
            'check_alergiasi': map_to_boolean('check_alergiasi'),
            'check_cirugiano': map_to_boolean('check_cirugiano'),
            'check_cirugiasi': map_to_boolean('check_cirugiasi'), 
            'check_visionsinalteracion': map_to_boolean('check_visionsinalteracion'),
            'check_visionrefraccion': map_to_boolean('check_visionrefraccion'),
            'check_audicionnormal': map_to_boolean('check_audicionnormal'),
            'check_hipoacusia': map_to_boolean('check_hipoacusia'),
            'check_tapondecerumen': map_to_boolean('check_tapondecerumen'),
            'check_sinhallazgos': map_to_boolean('check_sinhallazgos'),
            'check_caries': map_to_boolean('check_caries'),
            'check_apinamientodental': map_to_boolean('check_apinamientodental'),
            'check_retenciondental': map_to_boolean('check_retenciondental'),
            'check_frenillolingual': map_to_boolean('check_frenillolingual'),
            'check_hipertrofia': map_to_boolean('check_hipertrofia'),
            'altura': get_form_field_value('altura', request.form, return_none_if_empty=True),
            'peso': get_form_field_value('peso', request.form, return_none_if_empty=True),
            'imc': get_form_field_value('imc', request.form, return_none_if_empty=True),
            'clasificacion_imc': get_form_field_value('clasificacion_imc', request.form, return_none_if_empty=True),
        })

    print(f"DEBUG: Payload final para Supabase PATCH en /marcar_evaluado: {update_data}")
    
    try:
        print(f"DEBUG: Intentando PATCH a estudiantes_nomina con ID: {estudiante_id}.")
        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{estudiante_id}",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=update_data
        )
        
        if response.status_code >= 400: 
            print(f"ERROR: Supabase PATCH falló en /marcar_evaluado.")
            print(f"ERROR: Estado HTTP: {response.status_code}")
            print(f"ERROR: Cuerpo de la respuesta de Supabase: {response.text}")
            return jsonify({"success": False, "message": f"Error al actualizar estudiante: {response.text}"}), response.status_code

        print(f"DEBUG: Estudiante {estudiante_id} marcado como evaluado y guardado en Supabase. Status: {response.status_code}")
        print(f"DEBUG: Respuesta exitosa de Supabase: {response.text}")
        return jsonify({"success": True, "message": "Estudiante marcado como evaluado y datos guardados."})

    except requests.exceptions.RequestException as e:
        print(f"ERROR: Error de conexión con Supabase: {str(e)}")
        return jsonify({"success": False, "message": f"Error de conexión con Supabase: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR: Error inesperado al marcar estudiante como evaluado: {e}")
        return jsonify({"success": False, "message": f"Error interno del servidor: {str(e)}"}), 500

        
@app.route('/')
def index():
    return render_template('login.html')

# --- INICIO MODIFICACIONES CLAVE PARA COORDINADOR DE ESCUELA ---

@app.route('/login', methods=['POST'])
def login():
    usuario_login = request.form['username']
    clave = request.form['password']
    
    url = f"{SUPABASE_URL}/rest/v1/doctoras?usuario=eq.{usuario_login}&password=eq.{clave}&select=id,rol,nombre"
    
    print(f"DEBUG: Intento de login para usuario: {usuario_login}, URL: {url}")
    try:
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS) 
        res.raise_for_status()
        data = res.json()
        print(f"DEBUG: Respuesta Supabase login (Initial): {data}")
        
        if data:
            user_data = data[0]
            role = user_data['rol']
            
            session['usuario']    = role
            session['usuario_id'] = user_data['id']
            session['nombre']     = user_data.get('nombre') or usuario_login
            
            # 2. Lógica específica para COORDINADOR DE ESCUELA
            if role == 'coordinador_escuela':
                
                # --- OBTENER COLEGIOS/NÓMINAS ASIGNADAS DESDE NOMINAS_MEDICAS ---
                url_nominas_asignadas_por_colegio = (
                    f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                    f"?coord_escuela_id=eq.{user_data['id']}" # Filtrar por el ID del coordinador
                    f"&select=nombre_colegio,token_acceso"
                )
                res_nominas = requests.get(url_nominas_asignadas_por_colegio, headers=SUPABASE_SERVICE_HEADERS)
                res_nominas.raise_for_status()
                nominas_raw = res_nominas.json()
                
                # Agrupar por nombre_colegio para tener una lista única de colegios
                colegios_asignados_data = {}
                for nom in nominas_raw:
                    nombre_colegio = nom.get('nombre_colegio')
                    
                    if nombre_colegio and nombre_colegio not in colegios_asignados_data:
                        # Usamos el nombre del colegio como ID temporal para el frontend
                        colegios_asignados_data[nombre_colegio] = {
                            'id': nombre_colegio, # Usar el nombre como ID/clave de acceso temporal
                            'nombre_colegio': nombre_colegio
                        }

                colegios_asignados_list = list(colegios_asignados_data.values())
                # Los IDs ahora son los nombres de los colegios
                establecimientos_ids = [c['id'] for c in colegios_asignados_list] 

                session['colegios_asignados_ids'] = establecimientos_ids
                session['colegios_asignados'] = colegios_asignados_list
            
            print(f"DEBUG: Sesión iniciada: usuario={session['usuario']}, usuario_id={session['usuario_id']}")
            flash(f'¡Bienvenido, {usuario_login}!', 'success')
            return redirect(url_for('dashboard'))
        
        flash('Usuario o contraseña incorrecta.', 'error')
        return redirect(url_for('index'))
        
    except requests.exceptions.RequestException as e:
        print(f"❌ Error en el login: {e} - {res.text if 'res' in locals() else ''}")
        flash('Error de conexión al intentar iniciar sesión o error de base de datos.', 'error')
        return redirect(url_for('index'))

@app.route('/admin/cargar_informe_individual', methods=['POST'])
def admin_cargar_informe_individual():
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))
    
    # 1. Obtener datos del formulario
    nombre_establecimiento = request.form.get('nombre_establecimiento_informe', '').strip()
    doctora_id_asignada = request.form.get('doctora_asignada_informe', '').strip()
    excel_file = request.files.get('excel_informe_individual')

    # Validaciones básicas
    if not all([nombre_establecimiento, doctora_id_asignada, excel_file]):
        flash('❌ Faltan campos obligatorios para cargar el informe.', 'error')
        return redirect(url_for('dashboard'))

    if not permitido(excel_file.filename):
        flash('❌ Archivo Excel o CSV no válido. Extensiones permitidas: .xls, .xlsx, .csv', 'error')
        return redirect(url_for('dashboard'))

    # Generar un ID ÚNICO para esta carga (como si fuera una "mini-nomina")
    nomina_id_individual = str(uuid.uuid4())
    excel_filename = secure_filename(excel_file.filename)
    excel_file_data = excel_file.read()

    try:
        # (Opcional) Subir el archivo a Supabase Storage (se mantiene el flujo de la nómina)
        upload_path = f"informes-individuales/{nomina_id_individual}/{excel_filename}"
        upload_url = f"{SUPABASE_URL}/storage/v1/object/{upload_path}"
        
        res_upload = requests.put(upload_url, headers=SUPABASE_SERVICE_HEADERS, data=excel_file_data)
        res_upload.raise_for_status()
        url_excel_publica = f"{SUPABASE_URL}/storage/v1/object/public/{upload_path}" 

    except requests.exceptions.RequestException as e:
        error_detail = res_upload.text if 'res_upload' in locals() else 'No response from Supabase Storage.'
        flash(f"❌ Error al subir el archivo a Supabase Storage: {error_detail}", 'error')
        return redirect(url_for('dashboard'))

    # 2. Procesar Excel y mapear datos
    excel_data_stream = io.BytesIO(excel_file_data)
    
    if excel_filename.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(excel_data_stream)
    elif excel_filename.endswith('.csv'):
        df = pd.read_csv(excel_data_stream, encoding='utf-8')
    else:
        flash('❌ Formato de archivo no soportado.', 'error')
        return redirect(url_for('dashboard'))

    # Normalizar columnas
    df.columns = [normalizar(col) for col in df.columns]

    # Mapeo estricto a las columnas requeridas (nombre, rut, fecha_nacimiento, nacionalidad)
    required_cols = {'nombre', 'rut', 'fecha_nacimiento', 'nacionalidad'}
    
    if not all(col in df.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in df.columns]
        flash(f"❌ El archivo no contiene las columnas necesarias: {', '.join(missing_cols)}.", 'error')
        return redirect(url_for('dashboard'))

    # 3. Crear la "Mini-Nómina" de Informe Individual (en la tabla nominas_medicas)
    data_nomina_individual = {
        "id": nomina_id_individual,
        "nombre_nomina": f"INF_{nombre_establecimiento}",
        "tipo_nomina": "INFORME_INDIVIDUAL_NEURO", # Nuevo tipo de nómina para filtrado
        "doctora_id": doctora_id_asignada, 
        "url_excel_original": url_excel_publica,
        "nombre_excel_original": excel_filename,
        "form_type": "informe_neurologico", # Nuevo tipo de formulario específico para routing
        "nombre_colegio": nombre_establecimiento,
        "coord_general_id": None, 
        "coord_escuela_id": None,
        "token_acceso": None,
        "establecimiento_id": None 
    }
    
    try:
        res_insert_nomina = requests.post(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=data_nomina_individual
        )
        res_insert_nomina.raise_for_status()

    except requests.exceptions.RequestException as e:
        error_detail = res_insert_nomina.text if 'res_insert_nomina' in locals() else 'No response from Supabase.'
        flash(f"❌ Error al guardar la nómina individual en DB: {error_detail}", 'error')
        return redirect(url_for('dashboard'))

    # 4. Insertar estudiantes
    estudiantes_a_insertar = []
    
    for index, row in df.iterrows():
        try:
            nombre_completo_raw = row['nombre']
            rut_raw = row['rut']
            fecha_nacimiento_raw = row['fecha_nacimiento']
            nacionalidad_raw = row['nacionalidad'] 

            if pd.isna(nombre_completo_raw) or pd.isna(rut_raw) or pd.isna(fecha_nacimiento_raw):
                continue
            
            rut_limpio = str(rut_raw).replace('.', '').replace('-', '').strip()
            
            fecha_nac_str = None
            if isinstance(fecha_nacimiento_raw, (datetime, date)):
                fecha_nac_str = fecha_nacimiento_raw.strftime('%Y-%m-%d')
            else:
                try:
                    s = str(fecha_nacimiento_raw).strip()
                    parsed_date = None
                    # Intentar formatos explícitos en orden: DD/MM/YYYY → DD-MM-YYYY → YYYY-MM-DD
                    for fmt in ('%d/%m/%Y', '%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%y', '%d-%m-%y'):
                        try:
                            parsed_date = datetime.strptime(s, fmt)
                            break
                        except ValueError:
                            continue
                    # Fallback: pd.to_datetime con dayfirst=True
                    if parsed_date is None:
                        parsed_pd = pd.to_datetime(s, dayfirst=True, errors='coerce')
                        if pd.notna(parsed_pd):
                            parsed_date = parsed_pd.to_pydatetime()
                    if parsed_date:
                        fecha_nac_str = parsed_date.strftime('%Y-%m-%d')
                    else:
                        raise ValueError("Formato de fecha no reconocido.")
                except Exception:
                    fecha_nac_str = None 

            if fecha_nac_str is None:
                continue

            sexo_adivinado = guess_gender(str(nombre_completo_raw))
            nacionalidad_valor = str(nacionalidad_raw).strip() if pd.notna(nacionalidad_raw) else 'Chilena'

            # Calcular edad para pre-rellenar el campo
            fecha_nac_obj = datetime.strptime(fecha_nac_str, '%Y-%m-%d').date()
            edad_calculada = calculate_age(fecha_nac_obj)

            estudiante = {
                "nomina_id": nomina_id_individual,
                "nombre": str(nombre_completo_raw).strip(),
                "rut": rut_limpio,
                "fecha_nacimiento": fecha_nac_str, 
                "nacionalidad": nacionalidad_valor,
                "sexo": sexo_adivinado,
                "edad": edad_calculada, # Guardamos la edad calculada
                "fecha_relleno": None,
                "evaluado_flag": False,
                "tipo_registro_individual": "INFORME_NEURO", # Flag para diferenciar
            }
            estudiantes_a_insertar.append(estudiante)
            
        except Exception as e:
            print(f"❌ Error al procesar fila {index+2} para informe individual: {e}. Datos de la fila: {row.to_dict()}")
            flash(f"Error al procesar la fila {index+2} del archivo. ({e})", 'error')
            return redirect(url_for('dashboard'))

    if not estudiantes_a_insertar:
        flash("⚠️ El archivo Excel/CSV no contiene datos válidos para informes.", 'warning')
        return redirect(url_for('dashboard'))

    try:
        res_insert_estudiantes = requests.post(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=estudiantes_a_insertar
        )
        res_insert_estudiantes.raise_for_status()

        flash(f"✅ Informe(s) individual(es) cargado(s) con éxito. Se agregaron {len(estudiantes_a_insertar)} estudiantes.", 'success')
        return redirect(url_for('dashboard'))

    except requests.exceptions.RequestException as e:
        error_detail = res_insert_estudiantes.text if 'res_insert_estudiantes' in locals() else 'No response from Supabase.'
        flash(f"❌ Error al guardar los estudiantes en la base de datos. {error_detail}", 'error')
        return redirect(url_for('dashboard'))

@app.route('/api/admin/reporte_proyecto/<project_id>')
def reporte_proyecto_detalle(project_id):
    """
    Genera los datos para el reporte ejecutivo de Cardiohome SpA.
    Incluye totales globales y lista detallada de alumnos por nómina.
    """
    if session.get('usuario') != 'admin': 
        return jsonify({"success": False, "message": "No autorizado"}), 403

    try:
        # 1. Obtener nombre del proyecto real
        proyecto_nombre = "Reporte Global"
        if project_id != 'all':
            url_p = f"{SUPABASE_URL}/rest/v1/proyectos?id=eq.{project_id}&select=nombre_proyecto"
            res_p = requests.get(url_p, headers=SUPABASE_SERVICE_HEADERS)
            if res_p.ok and res_p.json():
                proyecto_nombre = res_p.json()[0]['nombre_proyecto']

        # 2. Obtener todas las nóminas vinculadas al proyecto
        url_n = f"{SUPABASE_URL}/rest/v1/nominas_medicas?select=id,nombre_nomina"
        if project_id != 'all':
            url_n += f"&proyecto_id=eq.{project_id}"
        
        res_n = requests.get(url_n, headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_n.json() if res_n.ok else []
        
        detalles_completos = []
        global_total = 0
        global_evaluados = 0
        
        # 3. Recorrer nóminas para traer alumnos y contar estados
        for nom in nominas:
            # Traer lista de alumnos activos/extra (excluir no_asiste)
            url_e = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(activo,extra)"
                f"&select=id,nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia,fecha_relleno"
                f"&order=nombre.asc"
            )
            res_e = requests.get(url_e, headers=SUPABASE_SERVICE_HEADERS)
            alumnos = res_e.json() if res_e.ok else []

            # También traer los no_asiste para mostrarlos aparte en el reporte
            url_ausentes = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(no_asiste_reemplazado,no_asiste_sin_reemplazo)"
                f"&select=nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia"
                f"&order=nombre.asc"
            )
            res_aus = requests.get(url_ausentes, headers=SUPABASE_SERVICE_HEADERS)
            ausentes = res_aus.json() if res_aus.ok else []

            ev_count    = len([a for a in alumnos if a.get('evaluado_flag') is True])
            total_count = len(alumnos)

            global_total     += total_count
            global_evaluados += ev_count

            detalles_completos.append({
                "colegio":   nom['nombre_nomina'],
                "alumnos":   alumnos,
                "ausentes":  ausentes,
                "total":     total_count,
                "evaluados": ev_count
            })

        global_pendientes = global_total - global_evaluados
        global_porcentaje = round((global_evaluados / global_total * 100), 1) if global_total > 0 else 0

        return jsonify({
            "success": True,
            "proyecto": proyecto_nombre,
            "fecha": datetime.now().strftime("%d-%m-%Y %H:%M"),
            "resumen": {
                "total": global_total,
                "evaluados": global_evaluados,
                "pendientes": global_pendientes,
                "porcentaje": f"{global_porcentaje}%"
            },
            "data": detalles_completos
        })
    except Exception as e:
        print(f"Error generando reporte: {str(e)}")
        return jsonify({"success": False, "error": str(e)})

# ─── GET /api/admin/stats_global — KPIs para welcome screen ───
@app.route('/api/admin/stats_global', methods=['GET'])
def api_admin_stats_global():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        total_alumnos  = get_supabase_count("estado_asistencia=in.(activo,extra)")
        total_evaluados = get_supabase_count("estado_asistencia=in.(activo,extra)&evaluado_flag=eq.true")
        total_pendientes = total_alumnos - total_evaluados

        # Contar nóminas activas
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?select=id",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "count=exact", "Range": "0-0"}
        )
        total_nominas = int(res_nom.headers.get("Content-Range", "0/0").split("/")[-1]) if res_nom.ok else 0

        return jsonify({
            "success": True,
            "total_evaluados":  total_evaluados,
            "total_pendientes": total_pendientes,
            "total_alumnos":    total_alumnos,
            "total_nominas":    total_nominas,
        })
    except Exception as e:
        print(f"ERROR api_admin_stats_global: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/admin/eliminar_visita/<visita_id>', methods=['POST'])
def eliminar_visita_calendario(visita_id):
    """
    Elimina una visita/establecimiento desde el calendario del Admin.
    El JS del dashboard usa POST + esta ruta exacta.
    """
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403

    print(f"DEBUG: Eliminando visita calendario ID: {visita_id}")

    try:
        # La tabla es 'establecimientos' — mismo origen que api_admin_visitas_doctora
        res = requests.delete(
            f"{SUPABASE_URL}/rest/v1/establecimientos?id=eq.{visita_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=minimal"}
        )

        # Supabase devuelve 200 o 204 en DELETE exitoso
        if res.status_code in (200, 204):
            print(f"DEBUG: Visita {visita_id} eliminada correctamente.")
            return jsonify({"success": True, "message": "Visita eliminada correctamente"})
        else:
            print(f"ERROR eliminar_visita: status={res.status_code} body={res.text}")
            return jsonify({
                "success": False,
                "message": f"Error al eliminar en base de datos: {res.text}"
            }), 500

    except requests.exceptions.RequestException as e:
        print(f"ERROR eliminar_visita (request): {e}")
        return jsonify({"success": False, "message": f"Error de conexión: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR eliminar_visita (inesperado): {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500


# ══════════════════════════════════════════════════════════════════════
# MERCADO PÚBLICO — Proxy seguro (la API key queda en el servidor)
# ══════════════════════════════════════════════════════════════════════

MERCADO_PUBLICO_API_KEY = os.getenv("MERCADO_PUBLICO_API_KEY", "C1148555-988E-40E5-9115-809B36F23168")

@app.route('/api/mercadopublico/licitaciones', methods=['GET'])
def api_mp_licitaciones():
    """
    Proxy a la API de Mercado Público.
    Parámetros aceptados por la API: fecha (ddmmaaaa), estado, CodigoOrganismo, codigo.
    NO acepta keywords — el filtro por palabra clave se hace en Python sobre los resultados.
    """
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        keyword = request.args.get('keyword', '').strip().lower()
        estado  = request.args.get('estado', '').strip()
        fecha   = request.args.get('fecha', '').strip()
        codigo  = request.args.get('codigo', '').strip()

        # Si no viene fecha, usar hoy en formato ddmmaaaa
        if not fecha and not codigo:
            from datetime import datetime
            fecha = datetime.now().strftime('%d%m%Y')

        params = {'ticket': MERCADO_PUBLICO_API_KEY}
        if codigo: params['codigo'] = codigo
        if fecha:  params['fecha']  = fecha
        if estado: params['estado'] = estado

        res = requests.get(
            'https://api.mercadopublico.cl/servicios/v1/publico/licitaciones.json',
            params=params, timeout=25
        )

        if not res.ok:
            return jsonify({"success": False,
                            "message": f"Error API MP ({res.status_code}): {res.text[:300]}"}), 502

        data = res.json()
        items = data.get('Listado', [])

        # Filtrar por keyword en Python (sobre nombre + descripcion + organismo)
        if keyword and items:
            palabras = keyword.split()
            items_filtrados = []
            for item in items:
                texto = (
                    (item.get('Nombre') or '') + ' ' +
                    (item.get('Descripcion') or '') + ' ' +
                    (item.get('Organismo') or '') + ' ' +
                    (item.get('Rubro') or '')
                ).lower()
                if any(p in texto for p in palabras):
                    items_filtrados.append(item)
            items = items_filtrados

        data['Listado'] = items
        return jsonify({"success": True, "data": data, "total_filtrado": len(items)})

    except requests.exceptions.Timeout:
        return jsonify({"success": False,
                        "message": "Timeout al conectar con Mercado Público. Intenta de nuevo."}), 504
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/mercadopublico/ordenes', methods=['GET'])
def api_mp_ordenes():
    """
    Proxy: busca órdenes de compra por RUT proveedor.
    Parámetros aceptados: rutProveedor, estado, fecha (ddmmaaaa), codigo.
    """
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        rut    = request.args.get('rut', '77028328-0').strip()
        estado = request.args.get('estado', '').strip()
        fecha  = request.args.get('fecha', '').strip()

        # Limpiar el RUT (quitar puntos, dejar guion)
        rut_limpio = rut.replace('.', '').strip()

        params = {'ticket': MERCADO_PUBLICO_API_KEY, 'rutProveedor': rut_limpio}
        if estado: params['estado'] = estado
        if fecha:  params['fecha']  = fecha

        res = requests.get(
            'https://api.mercadopublico.cl/servicios/v1/publico/ordenesdecompra.json',
            params=params, timeout=25
        )
        if not res.ok:
            return jsonify({"success": False,
                            "message": f"Error API MP ({res.status_code}): {res.text[:300]}"}), 502

        data = res.json()
        items = data.get('Listado', [])

        # Filtrar por estado "recepcion conforme" si se pidió
        if estado and items:
            estado_lower = estado.lower().replace(' ', '')
            items = [
                oc for oc in items
                if estado_lower in (oc.get('Estado') or '').lower().replace(' ', '')
            ]
            data['Listado'] = items

        return jsonify({"success": True, "data": data, "total": len(items)})

    except requests.exceptions.Timeout:
        return jsonify({"success": False, "message": "Timeout al conectar con Mercado Público."}), 504
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════
# FIX PUNTO 2 — GESTIÓN DE USUARIOS (Admin)
# ══════════════════════════════════════════════════════════════════════

@app.route('/api/admin/roles', methods=['GET'])
def api_admin_roles():
    """Devuelve los roles únicos desde la tabla doctoras — normalizados a minúscula."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?select=rol",
            headers=SUPABASE_SERVICE_HEADERS
        )
        roles_raw = res.json() if res.ok else []

        # Normalizar: minúscula, strip, ignorar None/vacíos, eliminar duplicados
        roles_set = set()
        for r in roles_raw:
            val = r.get('rol')
            if val and str(val).strip():
                roles_set.add(str(val).strip().lower())

        roles_norm = sorted(roles_set)

        # Roles base siempre disponibles (por si la tabla está vacía)
        for rb in ['admin', 'coordinador_escuela', 'coordinador_general', 'coordinadora', 'doctora']:
            if rb not in roles_norm:
                roles_norm.append(rb)
        roles_norm.sort()

        return jsonify({"success": True, "roles": roles_norm})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/admin/listar_usuarios', methods=['GET'])
def api_admin_listar_usuarios():
    """Lista todos los usuarios del sistema desde la tabla doctoras."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        # ✅ FIX: no incluir 'email' si la columna no existe en tu tabla doctoras
        # Usamos solo las columnas seguras que vimos en la imagen de Supabase
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?select=id,nombre,usuario,rol&order=nombre.asc",
            headers=SUPABASE_SERVICE_HEADERS
        )

        if not res.ok:
            return jsonify({
                "success": False,
                "message": f"Error Supabase {res.status_code}: {res.text[:300]}"
            }), 500

        usuarios_raw = res.json() if isinstance(res.json(), list) else []

        # Normalizar: rol NULL → string vacío, nunca None en el frontend
        usuarios = []
        for u in usuarios_raw:
            usuarios.append({
                "id":      u.get("id", ""),
                "nombre":  u.get("nombre") or "",
                "usuario": u.get("usuario") or "",
                "rol":     (u.get("rol") or "sin rol").strip().lower(),
                "email":   u.get("email") or "",
            })

        return jsonify({"success": True, "usuarios": usuarios, "total": len(usuarios)})

    except Exception as e:
        print(f"ERROR api_admin_listar_usuarios: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/admin/crear_usuario', methods=['POST'])
def api_admin_crear_usuario():
    """Admin crea un nuevo usuario en la tabla doctoras."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        data     = request.get_json() or {}
        nombre   = (data.get('nombre') or '').strip()
        usuario  = (data.get('usuario') or '').strip()
        password = (data.get('password') or '').strip()
        rol      = (data.get('rol') or '').strip()
        email    = (data.get('email') or '').strip() or None

        if not all([nombre, usuario, password, rol]):
            return jsonify({"success": False,
                            "message": "Nombre, usuario, contraseña y rol son obligatorios"}), 400

        # Verificar duplicado
        res_check = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?usuario=eq.{usuario}&select=id",
            headers=SUPABASE_SERVICE_HEADERS
        )
        if res_check.ok and res_check.json():
            return jsonify({"success": False,
                            "message": f"El usuario '{usuario}' ya existe"}), 409

        new_id  = str(uuid.uuid4())
        payload = {
            "id":       new_id,
            "nombre":   nombre,
            "usuario":  usuario,
            "password": password,
            "rol":      rol,
            "email":    email,
        }

        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/doctoras",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json=payload
        )
        if not res.ok:
            return jsonify({"success": False,
                            "message": f"Error Supabase: {res.text}"}), 500

        nuevo = res.json()[0] if res.json() else {}
        return jsonify({"success": True,
                        "message": f"✅ Usuario '{usuario}' creado exitosamente",
                        "usuario": nuevo})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/admin/eliminar_usuario/<user_id>', methods=['DELETE'])
def api_admin_eliminar_usuario(user_id):
    """Admin elimina un usuario. No puede eliminarse a sí mismo."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403

    if str(session.get('usuario_id')) == str(user_id):
        return jsonify({"success": False,
                        "message": "No puedes eliminarte a ti mismo"}), 400
    try:
        # Proteger al último admin
        res_admins = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?rol=eq.admin&select=id",
            headers=SUPABASE_SERVICE_HEADERS
        )
        admins = res_admins.json() if res_admins.ok else []

        res_target = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{user_id}&select=rol,nombre",
            headers=SUPABASE_SERVICE_HEADERS
        )
        target = res_target.json()[0] if res_target.ok and res_target.json() else {}

        if target.get('rol') == 'admin' and len(admins) <= 1:
            return jsonify({"success": False,
                            "message": "No puedes eliminar al único administrador del sistema"}), 400

        res = requests.delete(
            f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{user_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=minimal"}
        )
        if res.status_code in (200, 204):
            return jsonify({"success": True,
                            "message": f"Usuario '{target.get('nombre', user_id)}' eliminado correctamente"})
        return jsonify({"success": False, "message": f"Error: {res.text}"}), 500

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500



# - Ruta nueva: Asignar segunda doctora a nómina existente
@app.route('/api/admin/nomina_asignar_doctora2', methods=['POST'])
def api_nomina_asignar_doctora2():
    """Admin asigna o quita una segunda doctora a una nómina ya existente."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        data        = request.get_json() or {}
        nomina_id   = data.get('nomina_id', '').strip()
        doctora_id2 = data.get('doctora_id_2') or None  # None = quitar segunda doctora

        if not nomina_id:
            return jsonify({"success": False, "message": "nomina_id requerido"}), 400

        # Verificar que no sea la misma que la doctora principal
        if doctora_id2:
            res_nom = requests.get(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}&select=doctora_id,nombre_nomina",
                headers=SUPABASE_SERVICE_HEADERS
            )
            nom_data = res_nom.json()[0] if res_nom.ok and res_nom.json() else {}
            if doctora_id2 == nom_data.get('doctora_id'):
                return jsonify({"success": False,
                                "message": "La segunda doctora no puede ser la misma que la doctora principal"}), 400

        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json={"doctora_id_2": doctora_id2}
        )
        if not res.ok:
            return jsonify({"success": False, "message": f"Error Supabase: {res.text}"}), 500

        accion = "asignada" if doctora_id2 else "removida"
        return jsonify({"success": True, "message": f"Segunda doctora {accion} correctamente"})

    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# - Ruta 
@app.route('/api/admin/stats/<project_id>')
def get_admin_stats(project_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403

    try:
        from collections import defaultdict

        # Filtro opcional por doctora (query param)
        doctor_filter = request.args.get('doctor_id', 'all')

        # ── 1. Obtener nóminas filtradas ────────────────────────────────
        params = "select=id,tipo_nomina,doctora_id,nombre_colegio,nombre_nomina"
        if project_id != 'all':
            params += f"&proyecto_id=eq.{project_id}"
        if doctor_filter != 'all':
            params += f"&doctora_id=eq.{doctor_filter}"

        url_nominas = f"{SUPABASE_URL}/rest/v1/nominas_medicas?{params}"
        res_n = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_n.json() if res_n.ok else []

        # ── 2. Obtener lista de doctoras para lookup ────────────────────
        url_docs = f"{SUPABASE_URL}/rest/v1/doctoras?select=id,nombre,usuario&rol=eq.doctora"
        res_docs = requests.get(url_docs, headers=SUPABASE_SERVICE_HEADERS)
        doctoras_map = {}
        if res_docs.ok:
            for d in res_docs.json():
                doctoras_map[str(d['id'])] = d.get('nombre') or d.get('usuario', 'Desconocida')

        # ── 3. Contadores globales + por doctora + por establecimiento ──
        total_evaluados  = 0
        total_pendientes = 0
        neuro_count      = 0
        familiar_count   = 0
        doctor_stats     = defaultdict(lambda: {'completed': 0, 'total': 0, 'nombre': ''})
        est_stats        = defaultdict(lambda: {'completed': 0, 'total': 0, 'nombre': ''})
        # Tendencia: acumular fechas de evaluacion
        daily_counts     = defaultdict(int)
        weekly_counts    = defaultdict(int)
        hoy              = date.today()

        for nom in nominas:
            nom_id    = str(nom.get("id"))
            tipo      = (nom.get("tipo_nomina") or "").lower().strip()
            doc_id    = str(nom.get("doctora_id") or "")
            est_nombre = nom.get("nombre_colegio") or nom.get("nombre_nomina") or "Sin nombre"

            evaluados  = get_supabase_count(f"nomina_id=eq.{nom_id}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")
            pendientes = get_supabase_count(f"nomina_id=eq.{nom_id}&evaluado_flag=eq.false&estado_asistencia=in.(activo,extra)")
            subtotal   = evaluados + pendientes

            total_evaluados  += evaluados
            total_pendientes += pendientes

            # Especialidad
            if "neuro" in tipo:
                neuro_count += evaluados
            elif "familiar" in tipo or "medicina" in tipo:
                familiar_count += evaluados

            # Por doctora
            if doc_id:
                doctor_stats[doc_id]['completed'] += evaluados
                doctor_stats[doc_id]['total']     += subtotal
                doctor_stats[doc_id]['nombre']     = doctoras_map.get(doc_id, 'Desconocida')

            # Por establecimiento
            est_stats[est_nombre]['completed'] += evaluados
            est_stats[est_nombre]['total']     += subtotal
            est_stats[est_nombre]['nombre']     = est_nombre

            # Tendencia: obtener fechas de evaluaciones de esta nómina
            try:
                url_fechas = (
                    f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                    f"?nomina_id=eq.{nom_id}&evaluado_flag=eq.true"
                    f"&select=fecha_evaluacion"
                )
                res_f = requests.get(url_fechas, headers=SUPABASE_SERVICE_HEADERS)
                if res_f.ok:
                    for row in res_f.json():
                        fe = row.get('fecha_evaluacion')
                        if fe:
                            try:
                                fe_date = datetime.strptime(str(fe)[:10], '%Y-%m-%d').date()
                                delta = (hoy - fe_date).days
                                if 0 <= delta <= 30:
                                    daily_counts[str(fe_date)] += 1
                                delta_w = (hoy - fe_date).days // 7
                                if 0 <= delta_w < 12:
                                    iso_w = fe_date.isocalendar()
                                    wk_key = f"{iso_w[0]}-W{iso_w[1]:02d}"
                                    weekly_counts[wk_key] += 1
                            except Exception:
                                pass
            except Exception:
                pass

        # ── 4. Construir ranking de doctoras (top 10) ──────────────────
        ranking = []
        for did, stats in doctor_stats.items():
            t = stats['total'] or 1
            pct = round(stats['completed'] / t * 100, 1)
            ranking.append({
                'id':        did,
                'nombre':    stats['nombre'],
                'completed': stats['completed'],
                'total':     stats['total'],
                'percent':   pct
            })
        ranking.sort(key=lambda x: x['percent'], reverse=True)
        ranking = ranking[:10]

        # ── 5. Top establecimientos (top 10) ───────────────────────────
        establecimientos = []
        for ename, stats in est_stats.items():
            t = stats['total'] or 1
            pct = round(stats['completed'] / t * 100, 1)
            establecimientos.append({
                'nombre':    ename,
                'completed': stats['completed'],
                'total':     stats['total'],
                'percent':   pct
            })
        establecimientos.sort(key=lambda x: x['completed'], reverse=True)
        establecimientos = establecimientos[:10]

        # ── 6. Totales globales ────────────────────────────────────────
        total_alumnos = total_evaluados + total_pendientes
        percent = round((total_evaluados / total_alumnos * 100), 1) if total_alumnos > 0 else 0

        return jsonify({
            "success":         True,
            "total":           total_alumnos,
            "completed":       total_evaluados,
            "pending":         total_pendientes,
            "percent":         f"{percent}%",
            "neuro":           neuro_count,
            "familiar":        familiar_count,
            "ranking":         ranking,
            "establecimientos": establecimientos,
            "trend": {
                "daily":   dict(sorted(daily_counts.items())),
                "weekly":  dict(sorted(weekly_counts.items()))
            },
            # backward compat
            "chart_data": {r['id']: r['completed'] for r in ranking}
        })

    except Exception as e:
        import traceback
        print(f"❌ Error en get_admin_stats: {e}")
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)})
        
# app-30.py (Reemplaza la función dashboard completa)
@app.route('/dashboard')
def dashboard():
    if 'usuario' not in session:
        return redirect(url_for('index'))

    user_role = session.get('usuario')
    user_id = session.get('usuario_id')
    
    print(f"DEBUG: Accediendo a dashboard para usuario: {user_role}, ID: {user_id}")

    # --- Variables de inicialización ---
    doctoras_all = [] # Lista completa de doctoras (para select en admin)
    admin_nominas_cargadas = [] # Lista de nóminas para admin
    assigned_nominations = [] # Lista de nóminas para doctora
    proyectos = []
    
    # --- 1. Lógica de carga de PROYECTOS (ACTUALIZADO) ---
    try:
        # Añadimos descripcion_proyecto y created_at al select
        url_p = f"{SUPABASE_URL}/rest/v1/proyectos?select=id,nombre_proyecto,descripcion_proyecto,created_at&order=nombre_proyecto.asc"
        res_p = requests.get(url_p, headers=SUPABASE_SERVICE_HEADERS)
        if res_p.ok:
            raw_proyectos = res_p.json()
            # Mapeamos 'created_at' a 'fecha_creacion' para que el HTML lo entienda
            proyectos = []
            for p in raw_proyectos:
                proyectos.append({
                    "id": p['id'],
                    "nombre_proyecto": p['nombre_proyecto'],
                    "descripcion_proyecto": p.get('descripcion_proyecto'),
                    "fecha_creacion": p.get('created_at'), # <--- MAPEO CLAVE
                    "nominas": [] # Se llena más abajo
                })
    except Exception as e:
        print(f"❌ ERROR AL OBTENER PROYECTOS: {e}")
        proyectos = []
    
    # --- 2. Lógica de carga de USUARIOS ---
    try:
        url_doctoras = f"{SUPABASE_URL}/rest/v1/doctoras?select=id,usuario,rol,nombre" 
        res_doctoras = requests.get(url_doctoras, headers=SUPABASE_SERVICE_HEADERS) 
        res_doctoras.raise_for_status()
        doctoras_raw = res_doctoras.json()
        
        # Filtros de roles
        doctoras_all = [{'id': doc['id'], 'usuario': doc['usuario'], 'rol': doc.get('rol'), 'nombre': doc.get('nombre')} for doc in doctoras_raw]
        all_users_for_lookup = doctoras_all # Para las búsquedas en el HTML
        doctoras_relleno = [user for user in doctoras_all if user['rol'] == 'doctora'] 
        coordinadoras_generales = [user for user in doctoras_all if user['rol'] == 'coordinadora'] 
        coordinadores_escuela = [user for user in doctoras_all if user['rol'] == 'coordinador_escuela']

    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR AL OBTENER DATOS DE DOCTORAS: {e}")
        flash('Error al cargar datos de doctoras o usuarios.', 'error')
        all_users_for_lookup = []
        doctoras_relleno = []
        coordinadoras_generales = []
        coordinadores_escuela = []

    # --- 3. Lógica de carga de NÓMINAS (Admin y Doctora) ---
    if user_role == 'admin':
        # Admin ve TODAS las nóminas
        # NUEVO: Se agregó 'proyecto_id' al select
        url_nominas = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?select=id,nombre_nomina,tipo_nomina,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,url_excel_original,nombre_excel_original,form_type,doctora_id_para_formulario,nombre_colegio,coord_general_id,coord_escuela_id,proyecto_id"
            f"&order=nombre_nomina.asc"
        )
        
        try:
            res_nominas = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS) 
            res_nominas.raise_for_status()
            nominas_raw = res_nominas.json()

            for nom in nominas_raw:
                doctora_obj = next((doc for doc in all_users_for_lookup if doc['id'] == nom.get('doctora_id')), None)
                
                # Preparamos el objeto de la nómina
                datos_nomina = {
                    'id': nom['id'],
                    'nombre_nomina': nom['nombre_nomina'],
                    'tipo_nomina': nom['tipo_nomina'],
                    'doctora_id': nom['doctora_id'],
                    'doctora_id_2': nom.get('doctora_id_2'),
                    'url_excel_original': nom['url_excel_original'],
                    'nombre_excel_original': nom['nombre_excel_original'],
                    'form_type': nom['form_type'],
                    'doctora_id_para_formulario': nom.get('doctora_id_para_formulario'),
                    'coord_general_id': nom.get('coord_general_id'),
                    'coord_escuela_id': nom.get('coord_escuela_id'),
                    'proyecto_id': nom.get('proyecto_id'), # NUEVO
                    'nombre_colegio': nom.get('nombre_colegio') or nom['nombre_nomina'] 
                }
                
                # 1. La agregamos a la lista general que ya tenías
                admin_nominas_cargadas.append(datos_nomina)

                # 2. NUEVO: Lógica de Carpetas
                # Buscamos el proyecto correspondiente en la lista 'proyectos' y le metemos esta nómina
                for p in proyectos:
                    if str(p['id']) == str(nom.get('proyecto_id')):
                        if 'nominas' not in p:
                            p['nominas'] = []
                        p['nominas'].append(datos_nomina)

        except requests.exceptions.RequestException as e:
            print(f"❌ ERROR AL OBTENER NÓMINAS (ADMIN): {e}")
            flash('Error al cargar nóminas del administrador.', 'error')
            admin_nominas_cargadas = []

    elif user_role == 'doctora':
        # Doctora ve sus nóminas: como doctora principal Y como doctora 2, 3 o 4 (nómina compartida)
        url_nominas_principal = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id=eq.{user_id}"
            f"&select=id,nombre_nomina,tipo_nomina,form_type,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,doctora_id_para_formulario,nombre_colegio,proyecto_id"
            f"&order=nombre_nomina.asc"
        )
        url_nominas_compartida_2 = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id_2=eq.{user_id}"
            f"&select=id,nombre_nomina,tipo_nomina,form_type,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,doctora_id_para_formulario,nombre_colegio,proyecto_id"
            f"&order=nombre_nomina.asc"
        )
        url_nominas_compartida_3 = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id_3=eq.{user_id}"
            f"&select=id,nombre_nomina,tipo_nomina,form_type,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,doctora_id_para_formulario,nombre_colegio,proyecto_id"
            f"&order=nombre_nomina.asc"
        )
        url_nominas_compartida_4 = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id_4=eq.{user_id}"
            f"&select=id,nombre_nomina,tipo_nomina,form_type,doctora_id,doctora_id_2,doctora_id_3,doctora_id_4,doctora_id_para_formulario,nombre_colegio,proyecto_id"
            f"&order=nombre_nomina.asc"
        )
        
        try:
            res_n1 = requests.get(url_nominas_principal,    headers=SUPABASE_SERVICE_HEADERS)
            res_n2 = requests.get(url_nominas_compartida_2, headers=SUPABASE_SERVICE_HEADERS)
            res_n3 = requests.get(url_nominas_compartida_3, headers=SUPABASE_SERVICE_HEADERS)
            res_n4 = requests.get(url_nominas_compartida_4, headers=SUPABASE_SERVICE_HEADERS)
            nominas_lista_1 = res_n1.json() if res_n1.ok else []
            nominas_lista_2 = res_n2.json() if res_n2.ok else []
            nominas_lista_3 = res_n3.json() if res_n3.ok else []
            nominas_lista_4 = res_n4.json() if res_n4.ok else []

            # Combinar sin duplicados
            ids_vistos = set()
            nominas_raw = []
            for nom in nominas_lista_1 + nominas_lista_2 + nominas_lista_3 + nominas_lista_4:
                if nom['id'] not in ids_vistos:
                    ids_vistos.add(nom['id'])
                    nominas_raw.append(nom)
            
            for nom in nominas_raw:
                # Buscar el nombre del proyecto si tiene proyecto_id
                proyecto_nombre = None
                if nom.get('proyecto_id'):
                    proyecto_encontrado = next((p for p in proyectos if str(p['id']) == str(nom.get('proyecto_id'))), None)
                    if proyecto_encontrado:
                        proyecto_nombre = proyecto_encontrado['nombre_proyecto']
                
                assigned_nominations.append({
                    'id': nom['id'],
                    'nombre_establecimiento': nom['nombre_nomina'],
                    'tipo_nomina_display': nom['tipo_nomina'].replace('_', ' ').title(),
                    'form_type': nom.get('form_type'),
                    'doctora_id': nom.get('doctora_id'),
                    'doctora_id_2': nom.get('doctora_id_2'),
                    'doctora_id_para_formulario': nom.get('doctora_id_para_formulario'),
                    'nombre_colegio': nom.get('nombre_colegio') or nom['nombre_nomina'],
                    'proyecto_id': nom.get('proyecto_id'),
                    'proyecto_nombre': proyecto_nombre
                })
        
        except requests.exceptions.RequestException as e:
            print(f"❌ ERROR AL OBTENER NÓMINAS ASIGNADAS (DOCTORA): {e}")
            flash('Error al cargar nóminas asignadas.', 'error')
            assigned_nominations = []

    # --- Lógica de carga para Coordinador de Escuela (Usa datos de Session) ---
    colegios_asignados_escuela = session.get('colegios_asignados', [])

    # ------------------ Renderizado del Dashboard ------------------
    return render_template(
        'dashboard-23.html',
        usuario=user_role,
        rol=user_role, # Mantener rol para compatibilidad
        
        # Datos para ADMIN/COORDINADORA
        admin_nominas_cargadas=admin_nominas_cargadas,
        doctoras=doctoras_relleno, # Doctoras para SELECTS
        coordinadoras_generales=coordinadoras_generales, 
        proyectos=proyectos,
        coordinadores_escuela=coordinadores_escuela,
        all_users_for_lookup=all_users_for_lookup, # CRUCIAL: Lista de todos los usuarios
        
        # Datos para DOCTORA
        assigned_nominations=assigned_nominations,
        
        # Datos para COORDINADOR DE ESCUELA
        colegios_asignados=colegios_asignados_escuela,
        
        # Relleno de variables vacías (para evitar TemplateErrors)
        eventos=[], 
        formularios=[], 
        conteo={}, 
        establecimientos=[],
        doctor_performance_data={}, 
        doctor_performance_data_single_doctor={'completed': 0, 'pending': 0, 'total': 0},
        nombre_establecimiento_coordinador=None, 
        nominas_completadas_escuela=None
    )
    
@app.route('/logout')
def logout():
    # Marcar doctora como desconectada en presencia antes de limpiar sesión
    usuario_id = session.get('usuario_id')
    rol        = session.get('usuario')
    if usuario_id and rol == 'doctora':
        try:
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/presencia_doctoras?doctora_id=eq.{usuario_id}",
                headers=SUPABASE_SERVICE_HEADERS,
                json={"estado": "desconectada", "ultima_actividad": str(datetime.utcnow().isoformat()) + "Z"}
            )
        except Exception:
            pass
    session.clear()
    flash('Has cerrado sesión correctamente.', 'info')
    return redirect(url_for('index'))

# Coloque esto en su archivo principal de Flask (app.py)

from flask import jsonify, session

# ... (otras importaciones y configuración de Supabase)

# app-38.py (Alrededor de la línea 1307)

# app-38.py (Añadir esta ruta en cualquier parte, por ejemplo, cerca de /api/correcciones_pendientes)

@app.route('/api/correcciones/pendientes_detalle', methods=['GET'])
def get_correcciones_detalle():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    
    try:
        # CONSULTA: Obtener todas las solicitudes pendientes, incluyendo datos del alumno y solicitante
        url_requests = (
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion"
            f"?select=*,estudiantes_nomina(nombre,rut,nomina_id),doctoras(usuario)"
            f"&estado=eq.Pendiente"
            f"&order=fecha_solicitud.desc"
        )
        
        res = requests.get(url_requests, headers=SUPABASE_SERVICE_HEADERS)
        res.raise_for_status()
        solicitudes = res.json()
        
        # Procesamiento de datos para el frontend
        processed_requests = []
        for req in solicitudes:
            solicitante_usuario = req.get('doctoras', {}).get('usuario') if req.get('doctoras') else 'N/A'
            est_data  = req.get('estudiantes_nomina') or {}
            alumno_nombre = est_data.get('nombre', 'Alumno Desconocido')
            alumno_rut    = format_rut_python(est_data.get('rut', ''))
            alumno_id     = req.get('alumno_id')

            # Obtener nomina_id y form_type del alumno para poder abrir el formulario
            nomina_id  = est_data.get('nomina_id')
            form_type  = None
            nom_nombre = None
            nom_colegio = None
            proyecto_nombre = None
            if nomina_id:
                url_nom = (f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                           f"?id=eq.{nomina_id}&select=form_type,nombre_nomina,nombre_colegio,proyecto_id")
                r_nom = requests.get(url_nom, headers=SUPABASE_SERVICE_HEADERS)
                if r_nom.ok and r_nom.json():
                    nd = r_nom.json()[0]
                    form_type   = nd.get('form_type')
                    nom_nombre  = nd.get('nombre_nomina')
                    nom_colegio = nd.get('nombre_colegio') or nd.get('nombre_nomina')
                    pid = nd.get('proyecto_id')
                    if pid:
                        r_p = requests.get(f"{SUPABASE_URL}/rest/v1/proyectos?id=eq.{pid}&select=nombre_proyecto",
                                           headers=SUPABASE_SERVICE_HEADERS)
                        if r_p.ok and r_p.json():
                            proyecto_nombre = r_p.json()[0].get('nombre_proyecto')

            processed_requests.append({
                'id':               req['id'],
                'alumno_id':        str(alumno_id) if alumno_id else None,
                'alumno_nombre':    alumno_nombre,
                'alumno_rut':       alumno_rut,
                'detalles':         req.get('detalles', ''),
                'solicitante':      solicitante_usuario,
                'fecha':            req['fecha_solicitud'].split('T')[0] if req.get('fecha_solicitud') else 'N/A',
                'estado':           req['estado'],
                'nomina_id':        str(nomina_id) if nomina_id else None,
                'form_type':        form_type,
                'nombre_establecimiento': nom_nombre,
                'nombre_colegio':   nom_colegio,
                'proyecto_nombre':  proyecto_nombre,
                'url_documento_corregido': req.get('url_documento_corregido'),
                'respuesta_admin':  req.get('respuesta_admin', ''),
            })
            
        return jsonify({"success": True, "data": processed_requests})
        
    except requests.exceptions.RequestException as e:
        print(f"ERROR al obtener detalle de solicitudes: {e}")
        return jsonify({"success": False, "message": f"Error de conexión con BD: {str(e)}"}), 500
        
@app.route('/api/correcciones_pendientes', methods=['GET'])
def get_correcciones_pendientes():
    # 1. Ajustar la verificación de rol
    # Supabase guarda el rol como 'admin', no 'administrador'
    if 'usuario' not in session or session.get('usuario') != 'admin': 
        return jsonify({"count": 0, "success": True}), 200
    
    try:
        # 2. La consulta debe usar requests directamente si no tienes el objeto supabase-py
        # Usaremos get_supabase_count o una consulta directa con Prefer: count=exact

        url_count = (
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion"
            f"?select=id"
            f"&estado=eq.Pendiente" # Filtro crucial por el estado
        )
        
        # Usamos requests para obtener el conteo exacto (requiere el header Prefer: count=exact)
        res = requests.get(url_count, headers=SUPABASE_SERVICE_HEADERS)
        res.raise_for_status()
        
        # Leemos el conteo del Content-Range header
        content_range = res.headers.get("Content-Range")
        count_data = 0
        if content_range and '/' in content_range:
            count_data = int(content_range.split('/')[-1])
            
        print(f"DEBUG: Solicitudes Pendientes encontradas: {count_data}")

        return jsonify({
            "success": True, 
            "count": count_data 
        })

    except Exception as e:
        print(f"ERROR al consultar correcciones pendientes: {e}")
        return jsonify({"success": False, "count": 0, "message": "Error interno al consultar la base de datos."}), 500
        
# ==================================================================================
# RUTA /api/correccion/solicitar (ACTUALIZADA PARA USAR requests)
# ==================================================================================

# ==================================================================================
# RUTA /api/correccion/solicitar (SOLUCIÓN DEFINITIVA Y FINAL)
# ==================================================================================
# ==================================================================================
# RUTA /api/correccion/solicitar (SOLUCIÓN FINAL CON COLUMNAS SIMPLES)
# ==================================================================================
# ==================================================================================
# RUTA /api/correccion/solicitar (SOLUCIÓN FINAL CON COLUMNA SOLICITANTE_ID)
# ==================================================================================
@app.route('/api/correccion/solicitar', methods=['POST'])
def solicitar_correccion():
    # 1. Verificar sesión con la clave 'usuario_id'
    usuario_solicitante_id = session.get('usuario_id')
    
    if not usuario_solicitante_id:
        return jsonify({"success": False, "message": "Acceso no autorizado. Debe iniciar sesión para solicitar una corrección."}), 403

    try:
        # 2. Obtener los datos enviados (JSON)
        data = request.get_json()
        
        alumno_id = data.get('alumno_id')
        detalles = data.get('detalles')
        
        if not alumno_id or not detalles:
            return jsonify({"success": False, "message": "Faltan campos requeridos (alumno_id o detalles)."}), 400

        # 3. Preparar el cuerpo de la petición para Supabase
        payload = {
            "alumno_id": alumno_id,
            "detalles": detalles,        
            "estado": "Pendiente", 
            
            # 🟢 CORRECCIÓN CLAVE: Usamos la nueva columna solicitante_id (tipo UUID)
            "solicitante_id": usuario_solicitante_id,  
            
            # Nota: La columna "id" (clave primaria) se dejará para que Supabase la autogenere.
        }
        
        # 4. Enviar la petición POST al endpoint de Supabase
        url_supabase_insert = f"{SUPABASE_URL}/rest/v1/solicitudes_correccion"
        res = requests.post(url_supabase_insert, headers=SUPABASE_SERVICE_HEADERS, json=payload)
        
        # 5. Manejar la respuesta
        res.raise_for_status() 
        
        if res.status_code == 201:
            return jsonify({"success": True, "message": "Solicitud enviada correctamente."}), 201
        else:
            return jsonify({"success": False, "message": f"Error desconocido al insertar. Código: {res.status_code}"}), 500

    except requests.exceptions.HTTPError as http_err:
        print(f"❌ Error HTTP de Supabase: {http_err}. Respuesta: {http_err.response.text}")
        return jsonify({"success": False, "message": f"Error de Base de Datos (Supabase). Detalles: {http_err.response.text}"}), 500
        
    except Exception as e:
        print(f"❌ ERROR al procesar solicitud de corrección: {e}")
        return jsonify({"success": False, "message": f"Error interno del servidor. Detalle: {str(e)}"}), 500

# app-38.py (Añadir esta ruta)

@app.route('/api/correcciones/actualizar_estado', methods=['POST'])
def actualizar_estado_correccion():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    
    try:
        data = request.get_json()
        request_id = data.get('request_id')
        nuevo_estado = data.get('estado') # 'Aprobada' o 'Rechazada'

        if not request_id or nuevo_estado not in ['Aprobada', 'Rechazada']:
            return jsonify({"success": False, "message": "Datos de entrada inválidos"}), 400

        # 1. Preparar el payload de actualización
        respuesta = (data.get('respuesta_admin') or '').strip() or None
        update_data = {
            "estado":             nuevo_estado,
            "fecha_resolucion":   str(date.today()),
            "respuesta_admin":    respuesta,
            "notificacion_vista": False,  # coordinadora aún no vio la respuesta
        }
        response = requests.patch(
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?id=eq.{request_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json=update_data
        )
        response.raise_for_status()
        return jsonify({"success": True, "message": "Estado actualizado"})

    except requests.exceptions.RequestException as e:
        print(f"ERROR al actualizar estado de corrección: {e}")
        return jsonify({"success": False, "message": f"Error de conexión con BD: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR inesperado al actualizar estado: {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500
        
@app.route('/admin/agregar', methods=['POST'])
def admin_agregar():
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))

    # ... (Lógica existente para agregar establecimiento) ...
    nombre = request.form.get('nombre')
    fecha = request.form.get('fecha')
    horario = request.form.get('horario')
    obs = request.form.get('obs')
    doctora_id_from_form = request.form.get('doctora', '').strip()
    cantidad_alumnos = request.form.get('alumnos')
    
    if not all([nombre, fecha, horario, doctora_id_from_form]):
        flash("❌ Faltan campos obligatorios para el establecimiento.", 'error')
        return redirect(url_for('dashboard'))

    nuevo_id = str(uuid.uuid4())
    
    data_establecimiento = {
        "id": nuevo_id,
        "nombre": nombre,
        "fecha": fecha,
        "horario": horario,
        "observaciones": obs,
        "doctora_id": doctora_id_from_form,
        "cantidad_alumnos": int(cantidad_alumnos) if cantidad_alumnos else None,
        "url_archivo": None,
        "nombre_archivo": None
    }
    
    try:
        response_db = requests.post(
            f"{SUPABASE_URL}/rest/v1/establecimientos",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=data_establecimiento
        )
        response_db.raise_for_status()
        flash("✅ Establecimiento agregado correctamente.", 'success')
    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR AL GUARDAR ESTABLECIMIENTO EN DB: {e} - {response_db.text if 'response_db' in locals() else ''}")
        flash("❌ Error al guardar el establecimiento en la base de datos.", 'error')
    except Exception as e:
        print(f"❌ Error inesperado al guardar establecimiento: {e}")
        flash("❌ Error inesperado al guardar el establecimiento.", 'error')

    return redirect(url_for('dashboard'))


# - Modificar la función admin_cargar_nomina

# - Modificar la función admin_cargar_nomina

# Asegúrate de que esta importación esté en la parte superior de app-30.py:
# import secrets 

@app.route('/admin/cargar_nomina', methods=['POST'])
def admin_cargar_nomina():
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))
    
    # 1. Obtener datos del formulario
    tipo_nomina_raw = request.form.get('tipo_nomina', '').strip()
    nombre_colegio_o_establecimiento = request.form.get('nombre_especifico', '').strip() # ¡Usamos este campo como nombre del colegio!
    doctora_id_from_form = request.form.get('doctora', '').strip()
    excel_file = request.files.get('excel')
    doctora_id_para_formulario = request.form.get('doctora_id_para_formulario', '').strip()
    proyecto_id_from_form = request.form.get('proyecto_id', '').strip()
    
    # Doctoras adicionales (nómina compartida — opcionales)
    doctora_id_2_from_form = request.form.get('doctora_id_2', '').strip() or None
    doctora_id_3_from_form = request.form.get('doctora_id_3', '').strip() or None
    doctora_id_4_from_form = request.form.get('doctora_id_4', '').strip() or None

    # Fecha de evaluación fija (opcional)
    fecha_evaluacion_fija_from_form = request.form.get('fecha_evaluacion_fija', '').strip() or None
    
    # Obtener IDs de coordinación
    coord_general_id_from_form = request.form.get('coord_general_id', '').strip()
    coord_escuela_id_from_form = request.form.get('coord_escuela_id', '').strip()

    proyecto_id_db = proyecto_id_from_form if proyecto_id_from_form else None    
    
    tipo_nomina_normalized = tipo_nomina_raw.strip().lower() if tipo_nomina_raw else ''
    
    form_type = None
    if 'neurologia' in tipo_nomina_normalized: 
        form_type = 'neurologia'
    elif 'familiar' in tipo_nomina_normalized or 'medicina familiar' in tipo_nomina_normalized: 
        form_type = 'medicina_familiar'
    # 🟢 CORRECCIÓN DE INDENTACIÓN (Asegurado que esté alineado con los 'elif' anteriores)
    elif 'informe' in tipo_nomina_normalized and 'neuro' in tipo_nomina_normalized:
        form_type = 'informe_neurologico'
    # -----------------------------------------------------------------------------------
        
    # Validaciones básicas
    if not all([tipo_nomina_raw, nombre_colegio_o_establecimiento, doctora_id_from_form, excel_file]):
        flash('❌ Falta uno o más campos obligatorios.', 'error')
        return redirect(url_for('dashboard'))

    if form_type is None: 
        flash(f'❌ El tipo de nómina "{tipo_nomina_raw}" no se pudo mapear a un tipo de formulario conocido.', 'error')
        return redirect(url_for('dashboard'))

    if form_type == 'neurologia' and not doctora_id_para_formulario:
        flash('❌ Para nóminas de tipo "Neurología", debe seleccionar la Doctora para el formulario.', 'error')
        return redirect(url_for('dashboard'))


    if not permitido(excel_file.filename):
        flash('❌ Archivo Excel o CSV no válido. Extensiones permitidas: .xls, .xlsx, .csv', 'error')
        return redirect(url_for('dashboard'))

    nomina_id = str(uuid.uuid4())
    excel_filename = secure_filename(excel_file.filename)
    excel_file_data = excel_file.read()

    try:
        upload_path = f"nominas-medicas/{nomina_id}/{excel_filename}" 
        upload_url = f"{SUPABASE_URL}/storage/v1/object/{upload_path}"
        
        res_upload = requests.put(upload_url, headers=SUPABASE_SERVICE_HEADERS, data=excel_file_data)
        res_upload.raise_for_status()
        
        url_excel_publica = f"{SUPABASE_URL}/storage/v1/object/public/{upload_path}" 
    except requests.exceptions.RequestException as e:
        error_detail = res_upload.text if 'res_upload' in locals() else 'No response from Supabase Storage.'
        flash(f"❌ Error al subir el archivo de la nómina a Supabase Storage: {error_detail}", 'error')
        return redirect(url_for('dashboard'))

    # Mapear cadenas vacías a None para la DB (Crucial para NULL en UUID)
    coord_general_id_db = coord_general_id_from_form if coord_general_id_from_form else None
    coord_escuela_id_db = coord_escuela_id_from_form if coord_escuela_id_from_form else None

    # Limpiar todos los campos UUID — Supabase rechaza string vacío en columna UUID
    def uuid_or_none(v): return v if v else None
    doctora_id_from_form         = uuid_or_none(doctora_id_from_form)
    doctora_id_2_from_form       = uuid_or_none(doctora_id_2_from_form)
    doctora_id_3_from_form       = uuid_or_none(doctora_id_3_from_form)
    doctora_id_4_from_form       = uuid_or_none(doctora_id_4_from_form)
    doctora_id_para_formulario   = uuid_or_none(doctora_id_para_formulario)
    proyecto_id_db               = uuid_or_none(proyecto_id_db)

    # GENERACIÓN DEL TOKEN DE ACCESO (solo si hay coordinador de escuela asignado)
    token_generado = None
    if coord_escuela_id_db: 
        token_generado = secrets.token_hex(2) 

    # 3. Payload de Inserción (NOMBRES DE COLUMNA EXACTOS)
    data_nomina = {
        "id": nomina_id,
        "nombre_nomina": nombre_colegio_o_establecimiento, 
        "tipo_nomina": tipo_nomina_raw, 
        "doctora_id": doctora_id_from_form,
        "doctora_id_2": doctora_id_2_from_form,
        "doctora_id_3": doctora_id_3_from_form,
        "doctora_id_4": doctora_id_4_from_form,
        "url_excel_original": url_excel_publica,
        "nombre_excel_original": excel_filename,
        "form_type": form_type, 
        # doctora_id_para_formulario aplica a neurología Y medicina_familiar
        "doctora_id_para_formulario": doctora_id_para_formulario if form_type in ('neurologia', 'medicina_familiar', 'informe_neurologico') else None,
        "fecha_evaluacion_fija": fecha_evaluacion_fija_from_form,
        
        # --- CAMPOS CLAVE 100% INTEGRADOS ---
        "nombre_colegio": nombre_colegio_o_establecimiento,
        "coord_general_id": coord_general_id_db,
        "coord_escuela_id": coord_escuela_id_db,
        "token_acceso": token_generado,
        "establecimiento_id": None,
        "proyecto_id": proyecto_id_db,
    }
    
    try:
        # Intento de inserción en nominas_medicas
        res_insert_nomina = requests.post(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=data_nomina
        )
        res_insert_nomina.raise_for_status()

    except requests.exceptions.RequestException as e:
        error_detail = res_insert_nomina.text if 'res_insert_nomina' in locals() else 'No response from Supabase.'
        print(f"❌ ERROR AL GUARDAR NÓMINA EN DB: {error_detail}")
        flash(f"❌ Error al guardar los datos de la nómina en la base de datos. {error_detail}", 'error')
        # Rollback
        try:
            requests.delete(upload_url, headers=SUPABASE_SERVICE_HEADERS)
        except Exception: pass
        return redirect(url_for('dashboard'))

    excel_data_stream = io.BytesIO(excel_file_data)
    
    
    if excel_filename.endswith(('.xls', '.xlsx')):
        df = pd.read_excel(excel_data_stream)
    elif excel_filename.endswith('.csv'):
        df = pd.read_csv(excel_data_stream, encoding='utf-8')
    else:
        flash('❌ Formato de archivo no soportado para la nómina.', 'error')
        # Rollback
        try:
            requests.delete(upload_url, headers=SUPABASE_SERVICE_HEADERS)
            requests.delete(f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}", headers=SUPABASE_SERVICE_HEADERS)
        except Exception: pass
        return redirect(url_for('dashboard'))

    estudiantes_a_insertar = []
    omitidos               = []   # filas sin nombre — no se pueden cargar
    con_datos_incompletos  = []   # cargados pero con datos faltantes
    df.columns = [normalizar(col) for col in df.columns]

    column_mapping = {
        'nombre_completo': ['nombre_completo', 'nombre_del_estudiante', 'nombre'], 
        'rut': ['rut'],
        'fecha_nacimiento': ['fecha_nacimiento', 'fecha_de_nacimiento'],
        'nacionalidad': ['nacionalidad'],
        'diagnostico_previo': ['diagnostico previo', 'diagnostico_previo', 'diagnóstico previo'],
    }
    
    col_map = {}
    for key, possible_names in column_mapping.items():
        for name in possible_names:
            if name in df.columns:
                col_map[key] = name
                break
    
    # ✅ Solo el nombre es obligatorio — rut y fecha son opcionales
    if 'nombre_completo' not in col_map:
        flash("❌ El archivo no contiene la columna 'nombre' o 'nombre_completo'. Sin nombre no es posible cargar alumnos.", 'error')
        try:
            requests.delete(upload_url, headers=SUPABASE_SERVICE_HEADERS)
            requests.delete(f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}", headers=SUPABASE_SERVICE_HEADERS)
        except Exception: pass
        return redirect(url_for('dashboard'))
        
    establecimiento_id_db_para_estudiantes = None

    def _es_vacio(val):
        """Devuelve True si el valor es None, NaN o string vacío."""
        if val is None:
            return True
        try:
            if pd.isna(val):
                return True
        except Exception:
            pass
        return str(val).strip() == ''

    for index, row in df.iterrows():
        try:
            nombre_completo_raw   = row.get(col_map.get('nombre_completo'))
            rut_raw               = row.get(col_map.get('rut'))               if col_map.get('rut')               else None
            fecha_nacimiento_raw  = row.get(col_map.get('fecha_nacimiento'))  if col_map.get('fecha_nacimiento')  else None
            nacionalidad_raw      = row.get(col_map.get('nacionalidad'))      if col_map.get('nacionalidad')      else None
            diagnostico_previo_raw= row.get(col_map.get('diagnostico_previo'))if col_map.get('diagnostico_previo')else None

            # Nombre vacío → omitir fila (no hay nada que hacer)
            if _es_vacio(nombre_completo_raw):
                omitidos.append(index + 2)
                continue

            nombre_str = str(nombre_completo_raw).strip()

            # ── RUT: opcional ─────────────────────────────────────────────
            rut_limpio = None
            rut_faltante = _es_vacio(rut_raw)
            if not rut_faltante:
                rut_limpio = str(rut_raw).replace('.', '').replace('-', '').strip()

            # ── FECHA NACIMIENTO: opcional ────────────────────────────────
            fecha_nac_str  = None
            edad_calculada = None
            fecha_faltante = _es_vacio(fecha_nacimiento_raw)

            if not fecha_faltante:
                if isinstance(fecha_nacimiento_raw, (datetime, date)):
                    fecha_nac_str = fecha_nacimiento_raw.strftime('%Y-%m-%d')
                else:
                    s = str(fecha_nacimiento_raw).strip()
                    parsed_date = None
                    for fmt in ('%d/%m/%Y', '%d.%m.%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d.%m.%y', '%d-%m-%y'):
                        try:
                            parsed_date = datetime.strptime(s, fmt)
                            break
                        except ValueError:
                            continue
                    if parsed_date is None:
                        parsed_pd = pd.to_datetime(s, dayfirst=True, errors='coerce')
                        if pd.notna(parsed_pd):
                            parsed_date = parsed_pd.to_pydatetime()
                    if parsed_date:
                        fecha_nac_str = parsed_date.strftime('%Y-%m-%d')
                    else:
                        fecha_faltante = True  # valor inválido → tratar como faltante

                if fecha_nac_str:
                    try:
                        edad_calculada = calculate_age(datetime.strptime(fecha_nac_str, '%Y-%m-%d').date())
                    except Exception:
                        pass

            # ── Registrar campos faltantes ────────────────────────────────
            campos_faltantes = []
            if rut_faltante:   campos_faltantes.append('RUT')
            if fecha_faltante: campos_faltantes.append('Fecha de nacimiento')

            sexo_adivinado = guess_gender(nombre_str)
            nacionalidad_valor = str(nacionalidad_raw).strip() if not _es_vacio(nacionalidad_raw) else 'Chilena'

            estudiante = {
                "nomina_id":           nomina_id,
                "nombre":              nombre_str,
                "rut":                 rut_limpio,
                "fecha_nacimiento":    fecha_nac_str,
                "nacionalidad":        nacionalidad_valor,
                "sexo":                sexo_adivinado,
                "edad":                edad_calculada,
                "fecha_relleno":       None,
                "evaluado_flag":       False,
                "datos_incompletos":   len(campos_faltantes) > 0,
                "campos_faltantes":    ', '.join(campos_faltantes) if campos_faltantes else None,
                "diagnostico_sospecha": str(diagnostico_previo_raw).strip() if not _es_vacio(diagnostico_previo_raw) else None,
            }
            if form_type == 'informe_neurologico':
                estudiante["tipo_registro_individual"] = "INFORME_NEURO"

            if campos_faltantes:
                con_datos_incompletos.append(nombre_str)

            estudiantes_a_insertar.append(estudiante)
            
        except Exception as e:
            print(f"❌ Error fila {index+2}: {e}. Datos: {row.to_dict()}")
            continue  # ✅ No abortar todo por una fila con error

    if not estudiantes_a_insertar:
        flash("⚠️ El archivo Excel/CSV no contiene datos válidos. Verifica que exista al menos una columna 'nombre'.", 'warning')
        return redirect(url_for('dashboard'))

    try:
        res_insert_estudiantes = requests.post(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=estudiantes_a_insertar
        )
        res_insert_estudiantes.raise_for_status()

        msg = f"✅ Nómina '{nombre_colegio_o_establecimiento}' cargada. {len(estudiantes_a_insertar)} estudiantes agregados."
        if token_generado:
            msg += f" Token: {token_generado}."
        if omitidos:
            msg += f" ⚠️ {len(omitidos)} fila(s) sin nombre omitidas."
        if con_datos_incompletos:
            msg += f" ⚠️ {len(con_datos_incompletos)} alumno(s) cargados con datos incompletos (RUT o fecha faltante) — la doctora verá una alerta al evaluarlos."
        flash(msg, 'success')
        return redirect(url_for('dashboard'))

    except requests.exceptions.RequestException as e:
        error_detail = res_insert_estudiantes.text if 'res_insert_estudiantes' in locals() else 'No response from Supabase.'
        flash(f"❌ Error al guardar los estudiantes en la base de datos. La nómina fue creada, pero no se agregaron los estudiantes. ({e}). Detalles: {error_detail}", 'error')
        return redirect(url_for('dashboard'))


# La ruta '/enviar_formulario_a_drive' ha sido eliminada por completo.

@app.route('/subir/<establecimiento>', methods=['POST'])
def subir(establecimiento):
    if 'usuario' not in session:
        return redirect(url_for('index'))

    archivos = request.files.getlist('archivo')
    print(f"DEBUG: subir - Establecimiento ID: {establecimiento}, Cantidad de archivos: {len(archivos)}")
    print(f"DEBUG: ID de usuario en sesión (doctora) para /subir: {session.get('usuario_id')}")


    if not archivos or archivos[0].filename == '':
        flash('No se seleccionó ningún archivo para subir.', 'error')
        return redirect(url_for('dashboard'))

    usuario_id = session['usuario_id']
    mensajes = []

    for archivo in archivos:
        if permitido(archivo.filename):
            filename = secure_filename(archivo.filename)
            file_data = archivo.read()

            unique_file_id = str(uuid.uuid4())

            upload_path = f"formularios_completados/{establecimiento}/{unique_file_id}/{filename}"
            upload_url = f"{SUPABASE_URL}/storage/v1/object/{upload_path}"
            print(f"DEBUG: Subiendo archivo completado a Storage: {upload_url}")
            
            try:
                res_upload = requests.put(upload_url, headers=SUPABASE_SERVICE_HEADERS, data=file_data)
                res_upload.raise_for_status()
                
                url_publica = f"{SUPABASE_URL}/storage/v1/object/public/{upload_path}" 
                print(f"DEBUG: Archivo completado subido, URL pública: {url_publica}")

                data = {
                    "doctoras_id": usuario_id,
                    "establecimientos_id": establecimiento,
                    "nombre_archivo": filename,
                    "url_archivo": url_publica
                }
                print(f"DEBUG: Payload para insertar formulario subido en DB: {data}")

                res_insert = requests.post(
                    f"{SUPABASE_URL}/rest/v1/formularios_subidos",
                    headers=SUPABASE_SERVICE_HEADERS, 
                    json=data
                )
                res_insert.raise_for_status()
                mensajes.append(f"✅ Archivo '{filename}' subido y registrado correctamente.")
            
            except requests.exceptions.RequestException as e:
                error_msg = f"❌ Error al subir o registrar '{filename}': {e} - {res_upload.text if 'res_upload' in locals() else res_insert.text if 'res_insert' in locals() else 'No response'}"
                print(error_msg)
                mensajes.append(error_msg)
            except Exception as e:
                error_msg = f"❌ Error inesperado al procesar '{filename}': {e}"
                print(error_msg)
                mensajes.append(error_msg)
        else:
            mensajes.append(f"⚠️ Archivo '{archivo.filename}' no permitido.")
    
    for msg in mensajes:
        flash(msg, 'success' if '✅' in msg else 'error' if '❌' in msg else 'warning')

    return redirect(url_for('dashboard'))

@app.route('/colegios')
def colegios():
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))
    
    return render_template('colegios.html')

@app.route('/mis_nominas')
def mis_nominas():
    if 'usuario' not in session:
        return redirect(url_for('index'))
    
    usuario_id = session.get('usuario_id')
    assigned_nominations = []

    if not usuario_id:
        flash("No se pudo obtener el ID de usuario.", "error")
        return redirect(url_for('dashboard'))

    try:
        url_nominas_asignadas = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id=eq.{usuario_id}"
            f"&select=id,nombre_nomina,tipo_nomina,form_type,doctora_id_para_formulario" 
        )
        res_nominas_asignadas = requests.get(url_nominas_asignadas, headers=SUPABASE_HEADERS)
        res_nominas_asignadas.raise_for_status()
        raw_nominas = res_nominas_asignadas.json()

        for nom in raw_nominas:
            display_name = nom['tipo_nomina'].replace('_', ' ').title()
            assigned_nominations.append({
                'id': nom['id'],
                'nombre_establecimiento': nom['nombre_nomina'],
                'tipo_nomina_display': display_name,
                'form_type': nom.get('form_type'),
                'doctora_id_para_formulario': nom.get('doctora_id_para_formulario')
            })

    except requests.exceptions.RequestException as e:
        print(f"❌ Error al obtener mis nóminas: {e}")
        flash('Error al cargar sus nóminas asignadas.', 'error')
    except Exception as e:
        print(f"❌ Error inesperado al procesar mis nóminas: {e}")
        flash('Error inesperado al cargar sus nóminas asignadas.', 'error')


    return render_template('mis_nominas.html', assigned_nominations=assigned_nominations)

@app.route('/evaluados/<establecimiento>', methods=['POST'])
def evaluados(establecimiento):
    if 'usuario' not in session:
        return redirect(url_for('index'))

    alumnos_evaluados = request.form.get('alumnos')
    
    data_update = {
        "cantidad_alumnos_evaluados": int(alumnos_evaluados) if alumnos_evaluados else 0
    }

    try:
        response_db = requests.patch(
            f"{SUPABASE_URL}/rest/v1/establecimientos?id=eq.{establecimiento}",
            headers=SUPABASE_SERVICE_HEADERS, 
            json=data_update
        )
        response_db.raise_for_status()
        flash("✅ Cantidad de alumnos evaluados registrada correctamente.", 'success')
    except requests.exceptions.RequestException as e:
        print(f"❌ Error al registrar alumnos evaluados: {e} - {response_db.text if 'response_db' in locals() else ''}")
        flash("❌ Error al registrar la cantidad de alumnos evaluados.", 'error')
    except Exception as e:
        print(f"❌ Error inesperado al registrar alumnos evaluados: {e}")
        flash("❌ Error inesperado al registrar la cantidad de alumnos evaluados.", 'error')

    return redirect(url_for('dashboard'))

# --- NUEVA RUTA: RUTA DE DESBLOQUEO DEL COORDINADOR DE ESCUELA ---
# app-30.py (Reemplaza la función /api/nomina/desbloquear completa)
# app-30.py (Reemplaza la función /api/nomina/desbloquear completa)
# app-30.py (Reemplaza la función /api/nomina/desbloquear completa)
# app-30.py (Reemplaza la función /api/nomina/desbloquear completa)
@app.route('/api/nomina/desbloquear', methods=['POST'])
def desbloquear_nomina():
    # 1. Verificación Inicial de Seguridad
    if session.get('usuario') != 'coordinador_escuela':
        return jsonify({"success": False, "message": "Acceso no autorizado al recurso"}), 403
    
    data = request.get_json()
    password_ingresada = data.get('password')
    school_name = data.get('school_id') 
    
    # 2. Verificación de Asignación
    colegios_permitidos = session.get('colegios_asignados_ids', [])
    if school_name not in colegios_permitidos:
        return jsonify({"success": False, "message": "No tiene permiso asignado para este establecimiento."}), 403
        
    # 3. Validar el token de acceso para ese colegio (Envuelto en TRY/EXCEPT para errores de conexión)
    try:
        url_token = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?nombre_colegio=eq.{school_name}"
            f"&coord_escuela_id=eq.{session.get('usuario_id')}"
            f"&select=token_acceso,id"
        )
        
        res_token = requests.get(url_token, headers=SUPABASE_SERVICE_HEADERS)
        res_token.raise_for_status()
        nominas_con_token = res_token.json()
        
        token_esperado = nominas_con_token[0].get('token_acceso') if nominas_con_token and nominas_con_token[0] else None
        
        # 4. Comparamos la contraseña
        if token_esperado and token_esperado == password_ingresada:
            
            nomina_ids = [nom.get('id') for nom in nominas_con_token if nom.get('id')]
            nomina_ids_str = ",".join(nomina_ids)

            if not nomina_ids:
                return jsonify({"success": True, "nominas": []}) 
                
            # 5. CONSULTA FINAL DE ESTUDIANTES (Sin filtro de fecha)
            url_students = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=in.({nomina_ids_str})"
                f"&select=id,nombre,rut,fecha_evaluacion,fecha_relleno"
                f"&order=nombre.asc"
            )

            nominas_raw = requests.get(url_students, headers=SUPABASE_SERVICE_HEADERS).json()
            
            # 6. Procesamiento de datos (Recuperación del nombre de nómina en el bucle)
            nominas_procesadas = []
            for alumno in nominas_raw:
                
                if not isinstance(alumno, dict):
                    print(f"ADVERTENCIA: Elemento inesperado saltado: {alumno}")
                    continue 
                
                # Consulta de apoyo para el nombre de la nómina (usando el nomina_id)
                nombre_nomina = 'N/A'
                if alumno.get('nomina_id'):
                    url_nomina_name = f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{alumno['nomina_id']}&select=nombre_nomina"
                    res_nomina_name = requests.get(url_nomina_name, headers=SUPABASE_SERVICE_HEADERS)
                    if res_nomina_name.ok and res_nomina_name.json():
                        nombre_nomina = res_nomina_name.json()[0]['nombre_nomina']

                fecha_relleno = alumno.get('fecha_relleno')
                estado_evaluacion = "Evaluado" if fecha_relleno else "PENDIENTE"
                puede_descargar = estado_evaluacion == "Evaluado"
                
                nominas_procesadas.append({
                    'id': alumno.get('id'),
                    'nombre_alumno': alumno.get('nombre'), 
                    'rut_alumno': format_rut_python(alumno.get('rut')), 
                    'fecha_evaluacion': alumno.get('fecha_evaluacion') or 'N/A',
                    'nombre_nomina': nombre_nomina, 
                    'estado': estado_evaluacion, 
                    'descarga_habilitada': puede_descargar 
                })
            
            return jsonify({"success": True, "nominas": nominas_procesadas})
        else:
            # Token incorrecto
            return jsonify({"success": False, "message": "Token de acceso incorrecto"}), 401
            
    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR DE CONEXIÓN CON SUPABASE: {e}")
        return jsonify({"success": False, "message": f"Error de conexión con Supabase al validar token: {str(e)}"}), 500
    except Exception as e:
        print(f"❌ ERROR INESPERADO AL DESBLOQUEAR NÓMINA: {e}")
        # Retornar 500 para el cliente
        return jsonify({"success": False, "message": f"Error inesperado del servidor: {str(e)}"}), 500
        
# --- NUEVA RUTA: DESCARGA DE PDF POR ALUMNO ID ---@app.route('/descargar_pdf_alumno/<alumno_id>', methods=['GET'])
@app.route('/descargar_pdf_alumno/<alumno_id>', methods=['GET'])
def descargar_pdf_alumno(alumno_id):
    if session.get('usuario') not in ('coordinador_escuela', 'coordinadora', 'admin'):
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))

    try:
        # 1. CONSULTA ÚNICA: DATOS COMPLETOS DEL ESTUDIANTE (Select ALL FIELDS)
        url_student_data = (
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?id=eq.{alumno_id}"
            f"&select=id,nombre,rut,fecha_nacimiento,nacionalidad,sexo,estado_general,diagnostico,derivaciones,fecha_evaluacion,fecha_reevaluacion,fecha_relleno,diagnostico_1,diagnostico_2,diagnostico_complementario,clasificacion,observacion_1,observacion_2,observacion_3,observacion_4,observacion_5,observacion_6,observacion_7,check_cesarea,check_atermino,check_vaginal,check_prematuro,check_acorde,check_retrasogeneralizado,check_esquemac,check_esquemai,check_alergiano,check_alergiasi,check_cirugiano,check_cirugiasi,check_retraso,check_visionsinalteracion,check_visionrefraccion,check_audicionnormal,check_hipoacusia,check_tapondecerumen,check_sinhallazgos,check_caries,check_apinamientodental,check_retenciondental,check_frenillolingual,check_hipertrofia,altura,peso,imc,indicaciones,doctora_evaluadora_id,nomina_id,clasificacion_imc,motivo_consulta,observaciones,observacion_neurologia,estado_asistencia,motivo_ausencia,deficit" 
        )
        res_student = requests.get(url_student_data, headers=SUPABASE_SERVICE_HEADERS)
        res_student.raise_for_status() 
        student_data = res_student.json()

        if not student_data or not student_data[0].get('fecha_relleno'):
            flash(f"❌ Alumno ID {alumno_id} no encontrado o no evaluado.", 'error')
            return redirect(url_for('dashboard'))

        est = student_data[0]
        nomina_id_fk = est.get('nomina_id') 
        
        # 2. CONSULTA 2: METADATA DE LA NÓMINA
        url_nomina_meta = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?id=eq.{nomina_id_fk}"
            f"&select=form_type,nombre_nomina,doctora_id_para_formulario"
        )
        res_nomina = requests.get(url_nomina_meta, headers=SUPABASE_SERVICE_HEADERS)
        res_nomina.raise_for_status()
        nomina_meta = res_nomina.json()[0] if res_nomina.json() else {}
        
        form_type = nomina_meta.get('form_type', 'neurologia')
        nombre_nomina = nomina_meta.get('nombre_nomina', 'Valoracion')
        doctora_evaluadora_id = est.get('doctora_evaluadora_id')
        doctora_id_para_formulario = nomina_meta.get('doctora_id_para_formulario')
        
        # 4. LÓGICA DE PLANTILLA (Selección del PDF)
        pdf_base_path = ''
        base_dir = os.path.dirname(os.path.abspath(__file__))
        full_pdf_bases_dir_path = os.path.join(base_dir, PDF_BASES_NEUROLOGIA_DIR)

        if form_type == 'neurologia':
            doc_id_for_pdf = doctora_id_para_formulario or doctora_evaluadora_id
            specific_pdf_filename = f"FORMULARIO TIPO NEUROLOGIA_{doc_id_for_pdf}.pdf"
            specific_pdf_path = os.path.join(full_pdf_bases_dir_path, specific_pdf_filename)

            if doc_id_for_pdf and os.path.exists(specific_pdf_path):
                pdf_base_path = specific_pdf_path
            else:
                pdf_base_path = os.path.join(base_dir, PDF_BASE_NEUROLOGIA)

        elif form_type == 'informe_neurologico':
            doc_id_for_pdf = doctora_id_para_formulario or doctora_evaluadora_id
            specific_pdf_filename = f"INFORME_NEUROLOGICO_BASE_{doc_id_for_pdf}.pdf"
            specific_pdf_path = os.path.join(full_pdf_bases_dir_path, specific_pdf_filename)

            if doc_id_for_pdf and os.path.exists(specific_pdf_path):
                pdf_base_path = specific_pdf_path
            else:
                pdf_base_path = os.path.join(base_dir, PDF_BASE_INFORME_NEURO)
                
        elif form_type == 'medicina_familiar':
            # Soporte para PDF personalizado por doctora — fallback al base
            doc_id_for_pdf = doctora_id_para_formulario or doctora_evaluadora_id
            specific_pdf_filename_fam = f"FORMULARIO_FAMILIAR_{doc_id_for_pdf}.pdf"
            full_pdf_bases_familiar_dir = os.path.join(base_dir, PDF_BASES_FAMILIAR_DIR)
            specific_pdf_path_fam = os.path.join(full_pdf_bases_familiar_dir, specific_pdf_filename_fam)
            pdf_base_path = specific_pdf_path_fam if (doc_id_for_pdf and os.path.exists(specific_pdf_path_fam)) else os.path.join(base_dir, PDF_BASE_FAMILIAR)
        
        else:
            raise FileNotFoundError(f"Tipo de formulario no reconocido: {form_type}")
        
        if not os.path.exists(pdf_base_path):
            raise FileNotFoundError(f"Archivo base del formulario no encontrado: {pdf_base_path}")
            
        print(f"DEBUG: Usando PDF Base Path: {pdf_base_path}") 


        # 5. INICIALIZAR EL RELLENADOR DE PDF
        reader = PdfReader(pdf_base_path) 
        writer = PdfWriter()
        
        # Iterar sobre las páginas
        for page in reader.pages:
            writer.add_page(page)

        # 6. Preparar y Mapear Campos del PDF 
        nombre = est.get('nombre', '')
        rut = format_rut_python(est.get('rut', ''))
        
        # --- FUNCIÓN AUXILIAR PARA MAPEO TEXTO A '/Yes' ---
        def map_db_value_to_yes_pdf(db_value):
            if db_value is True or (isinstance(db_value, str) and db_value.strip()):
                return "/Yes" 
            return ""

        # --- Cálculo de campos y formato de fechas ---
        _fecha_eval_ref = datetime.strptime(est['fecha_evaluacion'], '%Y-%m-%d').date() if est.get('fecha_evaluacion') else None
        edad = calculate_age(datetime.strptime(est['fecha_nacimiento'], '%Y-%m-%d').date(), fecha_ref=_fecha_eval_ref) if est.get('fecha_nacimiento') else 'N/A'
        fecha_nac_formato = datetime.strptime(est['fecha_nacimiento'], '%Y-%m-%d').strftime('%d/%m/%Y') if est.get('fecha_nacimiento') else ''
        fecha_evaluacion_formatted = datetime.strptime(est['fecha_evaluacion'], '%Y-%m-%d').strftime('%d/%m/%Y') if est.get('fecha_evaluacion') else ''
        fecha_reeval_pdf = datetime.strptime(est['fecha_reevaluacion'], '%Y-%m-%d').strftime('%d/%m/%Y') if est.get('fecha_reevaluacion') else ''
        
        # --- Mapeo General ---
        campos = {}
        
        if form_type == 'neurologia':
            def _sg(f): v=est.get(f); return '' if v is None or str(v).strip() in ('None','null','') else str(v).strip()
            _sexo_raw = est.get('sexo', '')
            campos = {
                "nombre": nombre, "rut": rut, "fecha_nacimiento": fecha_nac_formato,
                "nacionalidad": _sg('nacionalidad'), "edad": edad,
                "diagnostico_1": _sg('diagnostico'), "diagnostico_2": _sg('diagnostico'),
                "estado_general": wrap_texto_pdf(_sg('estado_general')),
                "derivaciones":   wrap_texto_pdf(_sg('derivaciones')),
                "fecha_evaluacion": fecha_evaluacion_formatted, "fecha_reevaluacion": fecha_reeval_pdf,
                "sexo_f": "X" if _es_femenino(_sexo_raw) else "",
                "sexo_m": "X" if _es_masculino(_sexo_raw) else "",
            }

        elif form_type == 'informe_neurologico':
            campos = {
                "nombre": nombre, "rut": rut, "fecha_nacimiento": fecha_nac_formato,
                "edad": edad, "nacionalidad": est.get('nacionalidad', ''),
                "genero_m": "X" if est.get('sexo') == "M" else "",
                "genero_f": "X" if est.get('sexo') == "F" else "",
                "motivo_consulta":        wrap_texto_pdf(est.get('motivo_consulta', '')),
                "observaciones":          wrap_texto_pdf(est.get('observaciones', '')),
                "observacion_neurologia": wrap_texto_pdf(est.get('observacion_neurologia', '')),
                "diagnostico":            wrap_texto_pdf(est.get('diagnostico', '')),
                "indicaciones":           wrap_texto_pdf(est.get('indicaciones', '')),
                "derivaciones":           wrap_texto_pdf(est.get('derivaciones', '')),
                "fecha_evaluacion": fecha_evaluacion_formatted,
                "fecha_reevaluacion": fecha_reeval_pdf,
            }
        
        elif form_type == 'medicina_familiar':
            # Los campos check_ son campos de TEXTO en el PDF base — usan "X", no "/Yes"
            def map_check_db(val):
                if val is True or (isinstance(val, str) and val.strip() and val.strip().lower() not in ('false','0','no','')):
                    return "X"
                return ""

            # Safe getter: nunca retorna None ni el string "None"
            def sg(field):
                v = est.get(field)
                if v is None or str(v).strip() in ('None','null',''):
                    return ''
                return str(v).strip()

            campos = {
                "nombre": nombre, "rut": rut, "fecha_nacimiento": fecha_nac_formato, "edad": edad, "nacionalidad": sg('nacionalidad'),
                "sexo_f": "X" if est.get('sexo') == "F" else "", "sexo_m": "X" if est.get('sexo') == "M" else "",

                "diagnostico_1": sg('diagnostico_1'), "diagnostico_2": sg('diagnostico_2'),
                "diagnostico_complementario": sg('diagnostico_complementario'), "clasificacion": sg('clasificacion_imc'),
                "indicaciones": sg('indicaciones'), "derivaciones": sg('derivaciones'),
                "fecha_evaluacion": fecha_evaluacion_formatted, "fecha_reevaluacion": fecha_reeval_pdf,

                "altura": sg('altura'), "peso": sg('peso'), "imc": sg('imc'),
                "observacion_1": sg('observacion_1'), "observacion_2": sg('observacion_2'),
                "observacion_3": sg('observacion_3'), "observacion_4": sg('observacion_4'),
                "observacion_5": sg('observacion_5'), "observacion_6": sg('observacion_6'),
                "observacion_7": sg('observacion_7'),

                # Campos de texto que muestran "X" — misma lógica que generar_pdf (doctora)
                "check_cesarea":             map_check_db(est.get('check_cesarea')),
                "check_atermino":            map_check_db(est.get('check_atermino')),
                "check_vaginal":             map_check_db(est.get('check_vaginal')),
                "check_prematuro":           map_check_db(est.get('check_prematuro')),
                "check_acorde":              map_check_db(est.get('check_acorde')),
                "check_retraso":             map_check_db(est.get('check_retraso')),
                "check_retrasogeneralizado": map_check_db(est.get('check_retrasogeneralizado')),
                "check_esquemac":            map_check_db(est.get('check_esquemac')),
                "check_esquemai":            map_check_db(est.get('check_esquemai')),
                "check_alergiano":           map_check_db(est.get('check_alergiano')),
                "check_alergiasi":           map_check_db(est.get('check_alergiasi')),
                "check_cirugiano":           map_check_db(est.get('check_cirugiano')),
                "check_cirugiasi":           map_check_db(est.get('check_cirugiasi')),
                "check_visionsinalteracion": map_check_db(est.get('check_visionsinalteracion')),
                "check_visionrefraccion":    map_check_db(est.get('check_visionrefraccion')),
                "check_audicionnormal":      map_check_db(est.get('check_audicionnormal')),
                "check_hipoacusia":          map_check_db(est.get('check_hipoacusia')),
                "check_tapondecerumen":      map_check_db(est.get('check_tapondecerumen')),
                "check_sinhallazgos":        map_check_db(est.get('check_sinhallazgos')),
                "check_caries":              map_check_db(est.get('check_caries')),
                "check_apinamientodental":   map_check_db(est.get('check_apinamientodental')),
                "check_retenciondental":     map_check_db(est.get('check_retenciondental')),
                "check_frenillolingual":     map_check_db(est.get('check_frenillolingual')),
                "check_hipertrofia":         map_check_db(est.get('check_hipertrofia')),
                "deficit":                   "X" if (est.get('deficit') or '').strip() == 'X' else "",
            }


        # 7. Generar PDF final
        if form_type in ('neurologia', 'informe_neurologico'):
            # NEUROLOGÍA: ReportLab con coordenadas hardcodeadas
            # (PDF base tiene todos los campos ReadOnly)
            pdf_final = generar_pdf_neurologia_overlay(pdf_base_path, campos)
            if not pdf_final:
                flash("❌ Error al generar el PDF de neurología.", 'error')
                return redirect(url_for('dashboard'))
        else:
            # FAMILIAR y otros: flujo original PyPDF2 (funciona bien)
            if "/AcroForm" not in writer._root_object:
                writer._root_object.update({
                    NameObject("/AcroForm"): DictionaryObject()
                })
            for page in writer.pages:
                writer.update_page_form_field_values(page, campos)
            writer._root_object["/AcroForm"].update({
                NameObject("/NeedAppearances"): BooleanObject(True)
            })
            aplicar_autosize_campos(writer)
            output = io.BytesIO()
            writer.write(output)
            output.seek(0)
            pdf_final = aplicar_overlay_texto_largo(output.read(), campos)

        # Nombre del archivo para la descarga
        nombre_archivo_descarga = f"Valoracion_{nombre.replace(' ', '_')}_{rut}_{nombre_nomina.replace(' ', '_')}.pdf"
        
        return send_file(io.BytesIO(pdf_final), as_attachment=True, download_name=nombre_archivo_descarga, mimetype='application/pdf')

    except requests.exceptions.RequestException as e:
        print(f"❌ Error al obtener datos de Supabase para PDF: {e}")
        flash(f"❌ Error al generar PDF: Fallo de conexión/consulta. Detalles: {e}. Revise el log para las columnas.", 'error')
        return redirect(url_for('dashboard'))
    except FileNotFoundError as e:
        # Esto atrapará si el PDF_BASE_INFORME_NEURO no existe
        print(f"❌ Error File Not Found: {e}")
        flash(f"❌ Error al generar el PDF: Archivo base no encontrado. Detalle: {e}", 'error')
        return redirect(url_for('dashboard'))
    except Exception as e:
        print(f"❌ Error inesperado al generar PDF de alumno: {e}")
        flash(f"❌ Error inesperado al generar el PDF. Detalle: {e}", 'error')
        return redirect(url_for('dashboard'))
        
# app-30.py (Define esta ruta después de las funciones auxiliares)

# app-30.py (Define esta ruta después de las funciones auxiliares)

# app-30.py (Define esta ruta después de las funciones auxiliares)

# app-30.py (Reemplaza la función /api/dashboard_counts completa)
# app-30.py (Reemplaza la función /api/dashboard_counts completa y final)
# app-30.py (Reemplaza la función /api/dashboard_counts FINAL)
# app-30.py (Reemplaza la función /api/dashboard_counts FINAL)
# app-30.py (Reemplaza la función /api/dashboard_counts con la versión final basada en estado_general)
# app-30.py (Reemplaza la función /api/dashboard_counts con la versión de CONTEO INVERSO)
# app-30.py (Reemplaza la función /api/dashboard_counts con la versión de RESTA POR ESPECIALIDAD)
# app-30.py (Reemplaza la función /api/dashboard_counts con la versión HÍBRIDA FINAL)
# app-30.py (Reemplaza la función /api/dashboard_counts con la versión FINAL BASADA EN CONTEO DIRECTO)
# app-30.py (Versión de CONTINGENCIA FINAL)
# Asegúrate de que las importaciones de requests, uuid, y el resto estén al inicio de app.py
# from datetime import datetime, date  <-- CRÍTICO
# import requests
# import uuid
# ...

# Asegúrate de que las importaciones de requests, uuid, y datetime estén al inicio de app.py

# Reemplaza la función dashboard_counts en app.py con esta versión:

# Asegúrate de que las importaciones de requests, uuid, y datetime estén al inicio de app.py
# from datetime import datetime, date  <-- CRÍTICO
# import requests 
# import uuid
# ...

# Asegúrate de que las importaciones de requests, uuid, y datetime estén al inicio de app.py
# y que la función get_supabase_count esté definida.

# app.py (Reemplazo de la función dashboard_counts)

# app.py (Reemplazo de la función dashboard_counts)

# app.py (Reemplazo de la función dashboard_counts)

# app.py (Reemplazo de la función dashboard_counts)

# app.py (Reemplazo de la función dashboard_counts completa)

# app.py (Reemplazo de la función dashboard_counts completa)

# app.py (Reemplazo de la función dashboard_counts completa)

# app.py (Reemplazo de la función dashboard_counts completa)

# app.py (Reemplazo de la función dashboard_counts completa)

# app.py (Reemplazo de la función dashboard_counts completa)

@app.route('/api/dashboard_counts', methods=['GET'])
def dashboard_counts():
    print("🔍 Iniciando cálculo de dashboard_counts...")

    user_role = session.get('usuario')
    user_id = session.get('usuario_id')

    if not user_role or not user_id:
        print("❌ Usuario no autenticado")
        return jsonify({"error": "Usuario no autenticado", "success": False}), 401

    if user_role not in ['coordinadora', 'coordinadora_general', 'coordinador_general', 'admin']:
        print("❌ Usuario sin permisos")
        return jsonify({"error": "Permisos insuficientes", "success": False}), 403

    # --- CONSULTA DE NOMINAS MEDICAS ---
    try:
        if user_role == 'admin':
            url_nominas = f"{SUPABASE_URL}/rest/v1/nominas_medicas?select=id,tipo_nomina"
        else:
            url_nominas = (
                f"{SUPABASE_URL}/rest/v1/nominas_medicas?"
                f"coord_general_id=eq.{user_id}&select=id,tipo_nomina"
            )

        print("DEBUG URL NOMINAS:", url_nominas)

        res_n = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS)
        res_n.raise_for_status()
        nominas = res_n.json()

        print("NOMINAS RECIBIDAS:", nominas)

    except Exception as e:
        print("❌ Error al obtener nóminas:", e)
        return jsonify({"success": False, "error": str(e)}), 500

    # Desduplicar por id (por si acaso)
    nominas_unicas = { n["id"]: n for n in nominas if n.get("id") }
    nominas = list(nominas_unicas.values())

    # --- CONTADORES ---
    total_evaluados = 0
    total_pendientes = 0
    neuro_count = 0
    familiar_count = 0

    for nom in nominas:

        nom_id = nom.get("id")
        tipo = (nom.get("tipo_nomina") or "").lower().strip()

        if not nom_id:
            continue

        evaluados = get_supabase_count(
            f"nomina_id=eq.{nom_id}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)"
        )

        pendientes = get_supabase_count(
            f"nomina_id=eq.{nom_id}&evaluado_flag=eq.false&estado_asistencia=in.(activo,extra)"
        )

        print(f"DEBUG NOMINA {nom_id} → {tipo} | Evaluados={evaluados}, Pendientes={pendientes}")

        total_evaluados += evaluados
        total_pendientes += pendientes

        if "neuro" in tipo:
            neuro_count += evaluados
        elif "familiar" in tipo or "medicina" in tipo:
            familiar_count += evaluados

    return jsonify({
        "success": True,
        "total_evaluados": total_evaluados,
        "evaluaciones_pendientes": total_pendientes,
        "neurologia_count": neuro_count,
        "familiar_count": familiar_count
    })

# --- FIN MODIFICACIONES CLAVE PARA COORDINADOR DE ESCUELA ---

@app.route('/doctor_performance/<doctor_id>')
def doctor_performance_detail(doctor_id):
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))

    doctor_name = "Doctora Desconocida"
    evaluated_students = []

    try:
        url_doctora = f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{doctor_id}&select=usuario"
        res_doctora = requests.get(url_doctora, headers=SUPABASE_SERVICE_HEADERS)
        res_doctora.raise_for_status()
        doctor_data = res_doctora.json()
        if doctor_data:
            doctor_name = doctor_data[0]['usuario']
        print(f"DEBUG: Obteniendo rendimiento para doctora: {doctor_name} (ID: {doctor_id})")

        url_students = (
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?doctora_evaluadora_id=eq.{doctor_id}" 
            f"&fecha_relleno.not.is.null" 
            f"&select=nombre,rut,fecha_nacimiento,fecha_relleno,nomina_id,nominas_medicas(nombre_nomina)" 
            f"&order=nombre.asc" 
        )
        print(f"DEBUG: URL para obtener estudiantes evaluados: {url_students}")
        res_students = requests.get(url_students, headers=SUPABASE_SERVICE_HEADERS)
        res_students.raise_for_status()
        students_raw = res_students.json()
        print(f"DEBUG: Estudiantes evaluados recibidos: {students_raw}")

        for student in students_raw:
            formatted_date = student.get('fecha_relleno')
            if formatted_date and isinstance(formatted_date, str):
                try:
                    formatted_date = datetime.strptime(formatted_date, '%Y-%m-%d').strftime('%d-%m-%Y')
                except ValueError:
                    pass 
            
            nomina_nombre = "Nómina Desconocida"
            if student.get('nominas_medicas') and student['nominas_medicas']:
                if isinstance(student['nominas_medicas'], list) and student['nominas_medicas']:
                    nomina_nombre = student['nominas_medicas'][0].get('nombre_nomina', nomina_nombre)
                elif isinstance(student['nominas_medicas'], dict):
                    nomina_nombre = student['nominas_medicas'].get('nombre_nomina', nomina_nombre)


            evaluated_students.append({
                'nombre': student.get('nombre'),
                'rut': format_rut_python(student.get('rut')), # APLICA FORMATO AQUÍ TAMBIÉN SI SE MUESTRA EN ESTA VISTA
                'fecha_relleno': formatted_date,
                'nomina_nombre': nomina_nombre 
            })

    except requests.exceptions.RequestException as e:
        print(f"ERROR: Error de solicitud al obtener el rendimiento de la doctora: {e} - {res_students.text if 'res_students' in locals() else 'No response'}")
        flash('Error al cargar el detalle de rendimiento de la doctora.', 'error')
    except Exception as e:
        print(f"❌ Error inesperado al cargar rendimiento de doctora: {e}")
        flash('Error inesperado al cargar el detalle de rendimiento de la doctora.', 'error')

    return render_template('doctor_performance.html', 
                           doctor_name=doctor_name, 
                           evaluated_students=evaluated_students)

@app.route('/admin/crear_proyecto', methods=['POST'])
def crear_proyecto():
    if session.get('usuario') != 'admin':
        flash('Acceso denegado.', 'error')
        return redirect(url_for('dashboard'))

    # 1. Obtener datos del formulario
    nombre = request.form.get('nombre_proyecto', '').strip()
    descripcion = request.form.get('descripcion_proyecto', '').strip()
    
    # Obtenemos el ID del admin de la sesión para cumplir con la Foreign Key
    admin_id = session.get('usuario_id') 
    
    if not nombre:
        flash('❌ El nombre del proyecto es obligatorio.', 'error')
        return redirect(url_for('dashboard'))

    # 2. Payload con nombres largos (como confirmaste)
    payload = {
        "nombre_proyecto": nombre,
        "descripcion_proyecto": descripcion,
        "doctora_id": admin_id  # <--- Esto evita el error de la Foreign Key
    }

    try:
        proyectos_url = f"{SUPABASE_URL}/rest/v1/proyectos"
        response = requests.post(proyectos_url, json=payload, headers=SUPABASE_SERVICE_HEADERS)
        
        if response.status_code not in [200, 201]:
            print(f"❌ ERROR SUPABASE ({response.status_code}): {response.text}")
            flash(f"Error al guardar: {response.text}", 'error')
        else:
            flash(f"✅ Proyecto '{nombre}' creado con éxito.", 'success')
        
    except Exception as e:
        print(f"❌ ERROR CRÍTICO: {e}")
        flash(f"Error inesperado: {str(e)}", 'error')

    return redirect(url_for('dashboard'))
    
# app.py (Nueva ruta)

# app.py

# Asegúrate de que las siguientes líneas estén importadas al inicio de app.py:
# from datetime import datetime, date
# import uuid
# import requests
# ...

@app.route('/agregar_alumno_manual', methods=['POST'])
def agregar_alumno_manual():
    """
    Ruta para que una doctora agregue manualmente un nuevo estudiante a su nómina.
    El estudiante se marca automáticamente como 'evaluado' y 'agregado_manual'.
    """
    # 1. Verificar sesión y tipo de usuario
    if 'usuario' not in session or session.get('tipo_usuario') != 'doctora':
        return jsonify({"success": False, "message": "Acceso denegado. Debe ser una doctora."}), 401

    try:
        data = request.json
        # Datos requeridos del formulario HTML
        nombre = data.get('nombre')
        rut = data.get('rut')
        fecha_nac_str = data.get('fecha_nacimiento')
        nomina_id = data.get('nomina_id')
        doctora_id = session.get('usuario_id')
        
        if not all([nombre, rut, fecha_nac_str, nomina_id, doctora_id]):
            return jsonify({"success": False, "message": "Faltan campos requeridos. Por favor, complete todos los campos."}), 400

        # 2. Procesar datos (reutilizando funciones existentes)
        rut_formateado = format_rut_python(rut) # Función existente
        
        try:
            fecha_nac = datetime.strptime(fecha_nac_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({"success": False, "message": "Formato de fecha de nacimiento inválido."}), 400
        
        edad_calculada = calculate_age(fecha_nac) # Función existente
        sexo_adivinado = guess_gender(nombre) # Función existente

        # 3. Preparar datos para Supabase
        new_estudiante_id = str(uuid.uuid4()) # Generar un ID único
        
        payload = {
            "id": new_estudiante_id,
            "nomina_id": nomina_id,
            "nombre": nombre,
            "rut": rut_formateado,
            "fecha_nacimiento_formato": fecha_nac.strftime('%d/%m/%Y'),
            "fecha_nacimiento": fecha_nac_str,
            "edad": edad_calculada,
            "nacionalidad": data.get('nacionalidad', 'Chilena'), 
            "sexo": sexo_adivinado,
            "evaluado_flag": True, # CRÍTICO: Marcado como evaluado
            "agregado_manual": True, # CRÍTICO: Flag manual
            "fecha_evaluacion": date.today().isoformat(), # Fecha de registro/evaluación
            "doctora_evaluadora_id": doctora_id,
            "estado_general": "Pendiente", # Estado inicial
            # Dejar otros campos (diagnostico, etc.) en NULL/vacío para que la doctora los complete
        }

        # 4. Insertar en Supabase
        url_insert = f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
        # Usamos SUPABASE_SERVICE_HEADERS para inserciones (asegura permisos)
        res = requests.post(url_insert, headers=SUPABASE_SERVICE_HEADERS, json=payload)
        res.raise_for_status()

        # 5. Respuesta exitosa (devolvemos los datos para actualizar la UI sin recargar)
        return jsonify({
            "success": True, 
            "message": f"Alumno '{nombre}' agregado y marcado como evaluado (Manual).",
            "estudiante_data": payload 
        })

    except requests.exceptions.HTTPError as e:
        print(f"❌ ERROR AL INSERTAR EN SUPABASE: {e} - {e.response.text}")
        return jsonify({"success": False, "message": f"Error al guardar el alumno en la base de datos. Verifique los datos o contacte a soporte."}), 500
    except Exception as e:
        print(f"❌ ERROR INESPERADO en /agregar_alumno_manual: {e}")
        return jsonify({"success": False, "message": f"Error interno del servidor: {str(e)}"}), 500

# -------------------- FIN DE LA NUEVA RUTA --------------------        
@app.route('/descargar_excel_evaluados/<nomina_id>', methods=['GET'])
def descargar_excel_evaluados(nomina_id):
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    
    try:
        url_students = (
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&fecha_relleno.not.is.null" 
            f"&select=nombre,rut,fecha_nacimiento,fecha_relleno" 
            f"&order=nombre.asc" 
        )
        print(f"DEBUG: URL para descargar Excel de evaluados (simplificado): {url_students}")
        res_students = requests.get(url_students, headers=SUPABASE_SERVICE_HEADERS)
        res_students.raise_for_status()
        evaluated_students_data = res_students.json()
        print(f"DEBUG: Datos de estudiantes evaluados para Excel: {evaluated_students_data}")

        if not evaluated_students_data:
            return jsonify({"success": False, "message": "No hay formularios evaluados para esta nómina."}), 404

        df = pd.DataFrame(evaluated_students_data)

        df.rename(columns={
            'nombre': 'Nombre Completo',
            'rut': 'RUT',
            'fecha_nacimiento': 'Fecha de Nacimiento',
            'fecha_relleno': 'Fecha de Evaluación'
        }, inplace=True)

        # Formatear el RUT en el DataFrame antes de exportar a Excel
        df['RUT'] = df['RUT'].apply(format_rut_python) # APLICA EL FORMATO AQUÍ PARA EL EXCEL

        for col in ['Fecha de Nacimiento', 'Fecha de Evaluación']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce').dt.strftime('%d/%m/%Y').fillna('')
        
        df['Estado de Evaluación'] = df['Fecha de Evaluación'].apply(lambda x: 'Evaluado' if pd.notnull(x) and x != '' else 'Pendiente')

        df = df[['Nombre Completo', 'RUT', 'Fecha de Nacimiento', 'Estado de Evaluación']]

        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='xlsxwriter')
        df.to_excel(writer, index=False, sheet_name='Formularios Evaluados')
        writer.close() 
        output.seek(0)

        establecimiento_nombre = session.get('establecimiento_nombre', 'Nomina_Desconocida').replace(' ', '_')
        excel_filename = f"Formularios_Evaluados_{establecimiento_nombre}_{date.today().strftime('%Y%m%d')}.xlsx"

        return send_file(output, as_attachment=True, download_name=excel_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except requests.exceptions.RequestException as e:
        print(f"ERROR: Error de solicitud al descargar Excel de evaluados: {e}")
        return jsonify({"success": False, "message": f"Error de conexión con Supabase: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR: Error inesperado al generar Excel: {e}")
        return jsonify({"success": False, "message": f"Error interno del servidor al generar el Excel: {str(e)}"}), 500

@app.route('/generar_zip_pdfs_evaluados', methods=['POST'])
def generar_zip_pdfs_evaluados():
    if 'user_id' not in session:
        return jsonify({"success": False, "message": "Acceso denegado."}), 403

    data = request.get_json()
    nomina_id = data.get('nomina_id')
    student_ids = data.get('student_ids', [])

    if not nomina_id or not student_ids:
        return jsonify({"success": False, "message": "IDs de nómina o estudiantes faltantes."}), 400

    merged_pdf_writer = PdfWriter()
    # Obtener el tipo de formulario y el ID de la doctora para el PDF base de la nómina
    res_nomina = requests.get(
        f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}&select=nombre_nomina,form_type,doctora_id_para_formulario",
        headers=SUPABASE_SERVICE_HEADERS
    )
    res_nomina.raise_for_status()
    nomina_info = res_nomina.json()
    if not nomina_info:
        return jsonify({"success": False, "message": "Nómina no encontrada."}), 404
    
    nomina_nombre = nomina_info[0]['nombre_nomina']
    form_type = nomina_info[0]['form_type']
    doctora_id_para_formulario = nomina_info[0].get('doctora_id_para_formulario')

    pdf_base_path = None
    if form_type == 'neurologia':
        pdf_base_path = PDF_BASE_NEUROLOGIA 
        if not doctora_id_para_formulario:
            return jsonify({"success": False, "message": "No se especificó la doctora para el formulario de neurología."}), 400
    elif form_type == 'medicina_familiar':
        pdf_base_path = PDF_BASE_FAMILIAR
    else:
        return jsonify({"success": False, "message": "Tipo de formulario no soportado."}), 400

    if not os.path.exists(pdf_base_path):
        return jsonify({"success": False, "message": f"Archivo PDF base no encontrado: {pdf_base_path}"}), 500

    # ... (Lógica de generación de ZIP omitida por brevedad, usa la función generar_pdfs_visibles) ...

    return jsonify({"success": False, "message": "Ruta de ZIP no implementada completamente."}), 501

@app.route('/generar_pdfs_visibles', methods=['POST'])
def generar_pdfs_visibles():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    data = request.get_json()
    nomina_id = data.get('nomina_id')
    student_ids = data.get('student_ids')

    if not nomina_id or not student_ids or not isinstance(student_ids, list):
        return jsonify({"success": False, "message": "Datos de entrada inválidos."}), 400

    # --- FUNCIONES INTERNAS PARA EVITAR ERROS DE "NOT DEFINED" ---
    def mapear_check_interno(campo_nombre, diccionario_est):
        """Busca el valor en el diccionario del estudiante y devuelve 'X' si es verdadero."""
        val = diccionario_est.get(campo_nombre)
        if val is True or val in ['on', 'X', 'true', 'True', 1, '1']:
            return "X"
        return ""

    merged_pdf_writer = PdfWriter()
    form_type = session.get('current_form_type', 'neurologia') 
    doctora_id_para_formulario = session.get('doctora_id_para_formulario')

    base_dir = os.path.dirname(os.path.abspath(__file__))
    
    if form_type == 'neurologia':
        if doctora_id_para_formulario:
            pdf_base_path = get_doctor_specific_neurologia_pdf(doctora_id_para_formulario)
        else:
            pdf_base_path = os.path.join(base_dir, PDF_BASE_NEUROLOGIA)
    elif form_type == 'medicina_familiar':
        specific_pdf_filename_fam = f"FORMULARIO_FAMILIAR_{doctora_id_para_formulario}.pdf"
        full_pdf_bases_familiar_dir = os.path.join(base_dir, PDF_BASES_FAMILIAR_DIR)
        specific_pdf_path_fam = os.path.join(full_pdf_bases_familiar_dir, specific_pdf_filename_fam)
        pdf_base_path = specific_pdf_path_fam if (doctora_id_para_formulario and os.path.exists(specific_pdf_path_fam)) else os.path.join(base_dir, PDF_BASE_FAMILIAR)
    else:
        return jsonify({"success": False, "message": "Tipo de formulario no reconocido."}), 400

    if not os.path.exists(pdf_base_path):
        return jsonify({"success": False, "message": "Archivo base no encontrado."}), 500

    try:
        for student_id in student_ids:
            url_student_data = f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{student_id}&select=*"
            res_student = requests.get(url_student_data, headers=SUPABASE_SERVICE_HEADERS)
            res_student.raise_for_status()
            student_data = res_student.json()

            if not student_data:
                continue

            est = student_data[0] 

            # Preparación de datos básicos
            nombre_est = est.get('nombre', '')
            rut_est = format_rut_python(est.get('rut', ''))
            nacionalidad_est = est.get('nacionalidad', '')
            edad_est = est.get('edad', '')
            
            # Formateo de fechas (Uso de variables consistentes)
            fecha_nac_pdf = ''
            if est.get('fecha_nacimiento'):
                try:
                    fecha_nac_pdf = datetime.strptime(est['fecha_nacimiento'], '%Y-%m-%d').strftime('%d/%m/%Y')
                except: pass

            fecha_eval_pdf = ''
            if est.get('fecha_evaluacion'):
                try:
                    fecha_eval_pdf = datetime.strptime(est['fecha_evaluacion'], '%Y-%m-%d').strftime('%d/%m/%Y')
                except: pass

            fecha_reeval_pdf = ''
            if est.get('fecha_reevaluacion'):
                try:
                    fecha_reeval_pdf = datetime.strptime(est['fecha_reevaluacion'], '%Y-%m-%d').strftime('%d/%m/%Y')
                except: pass

            # Procesar PDF
            reader = PdfReader(pdf_base_path)
            writer_single_pdf = PdfWriter()
            writer_single_pdf.add_page(reader.pages[0])

            campos = {}
            if form_type == 'neurologia':
                campos = {
                    "nombre": nombre_est, "rut": rut_est, "fecha_nacimiento": fecha_nac_pdf, 
                    "nacionalidad": nacionalidad_est, "edad": edad_est,
                    "diagnostico_1": est.get('diagnostico', ''), "diagnostico_2": est.get('diagnostico', ''), 
                    "estado_general": est.get('estado_general', ''), "fecha_evaluacion": fecha_eval_pdf, 
                    "fecha_reevaluacion": fecha_reeval_pdf, "derivaciones": est.get('derivaciones', ''),
                    "sexo_f": "X" if est.get('sexo') == "F" else "", "sexo_m": "X" if est.get('sexo') == "M" else "",
                }
            elif form_type == 'medicina_familiar':
                # Nota: En esta ruta masiva usamos los datos de la DB (est.get)
                diag_unif = est.get('diagnostico_1') or est.get('diagnostico_unificado', '')
                
                campos = {
                    "nombre": nombre_est, "rut": rut_est, "fecha_nacimiento": fecha_nac_pdf, 
                    "edad": edad_est, "nacionalidad": nacionalidad_est,
                    "sexo_f": "X" if est.get('sexo') == "F" else "", "sexo_m": "X" if est.get('sexo') == "M" else "",
                    "diagnostico_1": diag_unif, "diagnostico_2": diag_unif, 
                    "diagnostico_complementario": est.get('diagnostico_complementario', ''),
                    "clasificacion": est.get('clasificacion_imc', '') or est.get('clasificacion', ''),
                    "indicaciones": wrap_texto_pdf(est.get('indicaciones', '')), 
                    "derivaciones": wrap_texto_pdf(est.get('derivaciones', '')), 
                    "fecha_evaluacion": fecha_eval_pdf, 
                    "fecha_reevaluacion": fecha_reeval_pdf,
                    "altura": est.get('altura', ''), "peso": est.get('peso', ''), "imc": est.get('imc', ''),
                    "observacion_1": est.get('observacion_1', ''), "observacion_2": est.get('observacion_2', ''),
                    "observacion_3": est.get('observacion_3', ''), "observacion_4": est.get('observacion_4', ''),
                    "observacion_5": est.get('observacion_5', ''), "observacion_6": est.get('observacion_6', ''),
                    "observacion_7": est.get('observacion_7', ''),
                    
                    # Uso de la función interna corregida
                    "check_cesarea": mapear_check_interno('check_cesarea', est),
                    "check_atermino": mapear_check_interno('check_atermino', est),
                    "check_vaginal": mapear_check_interno('check_vaginal', est),
                    "check_prematuro": mapear_check_interno('check_prematuro', est),
                    "check_acorde": mapear_check_interno('check_acorde', est),
                    "check_retraso": mapear_check_interno('check_retraso', est),
                    "check_retrasogeneralizado": mapear_check_interno('check_retrasogeneralizado', est),
                    "check_esquemac": mapear_check_interno('check_esquemac', est),
                    "check_esquemai": mapear_check_interno('check_esquemai', est),
                    "check_alergiano": mapear_check_interno('check_alergiano', est),
                    "check_alergiasi": mapear_check_interno('check_alergiasi', est),
                    "check_cirugiano": mapear_check_interno('check_cirugiano', est),
                    "check_cirugiasi": mapear_check_interno('check_cirugiasi', est),
                    "check_visionsinalteracion": mapear_check_interno('check_visionsinalteracion', est),
                    "check_visionrefraccion": mapear_check_interno('check_visionrefraccion', est),
                    "check_audicionnormal": mapear_check_interno('check_audicionnormal', est),
                    "check_hipoacusia": mapear_check_interno('check_hipoacusia', est),
                    "check_tapondecerumen": mapear_check_interno('check_tapondecerumen', est),
                    "check_sinhallazgos": mapear_check_interno('check_sinhallazgos', est),
                    "check_caries": mapear_check_interno('check_caries', est),
                    "check_apinamientodental": mapear_check_interno('check_apinamientodental', est),
                    "check_retenciondental": mapear_check_interno('check_retenciondental', est),
                    "check_frenillolingual": mapear_check_interno('check_frenillolingual', est),
                    "check_hipertrofia": mapear_check_interno('check_hipertrofia', est),
                }

            if form_type == 'neurologia':
                pdf_bytes = generar_pdf_neurologia_overlay(pdf_base_path, campos)
                if not pdf_bytes:
                    continue
                temp_reader = PdfReader(io.BytesIO(pdf_bytes))
                merged_pdf_writer.add_page(temp_reader.pages[0])
            else:
                if "/AcroForm" not in writer_single_pdf._root_object:
                    writer_single_pdf._root_object.update({NameObject("/AcroForm"): DictionaryObject()})
                writer_single_pdf.update_page_form_field_values(writer_single_pdf.pages[0], campos)
                writer_single_pdf._root_object["/AcroForm"].update({NameObject("/NeedAppearances"): BooleanObject(True)})
                aplicar_autosize_campos(writer_single_pdf)
                temp_output = io.BytesIO()
                writer_single_pdf.write(temp_output)
                temp_output.seek(0)
                temp_bytes = aplicar_overlay_texto_largo(temp_output.read(), campos)
                temp_reader = PdfReader(io.BytesIO(temp_bytes))
                merged_pdf_writer.add_page(temp_reader.pages[0])

        final_output_pdf = io.BytesIO()
        merged_pdf_writer.write(final_output_pdf)
        final_output_pdf.seek(0)

        pdf_filename = f"PDFs_Visibles_{date.today().strftime('%Y%m%d')}.pdf"
        return send_file(final_output_pdf, as_attachment=False, download_name=pdf_filename, mimetype='application/pdf')

    except Exception as e:
        print(f"ERROR: {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500

# --- Rutas de Eliminación (Solo para Admin) ---

@app.route('/admin/eliminar_establecimiento/<establecimiento_id>', methods=['DELETE'])
def eliminar_establecimiento(establecimiento_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado. Solo administradores pueden eliminar."}), 403
    
    print(f"DEBUG: Intentando eliminar establecimiento con ID: {establecimiento_id}")

    try:
        # Eliminar el establecimiento
        res_delete_est = requests.delete(
            f"{SUPABASE_URL}/rest/v1/establecimientos?id=eq.{establecimiento_id}",
            headers=SUPABASE_SERVICE_HEADERS
        )
        res_delete_est.raise_for_status()

        if res_delete_est.status_code == 204: # 204 No Content typically means successful deletion
            print(f"DEBUG: Establecimiento {establecimiento_id} eliminado de la DB.")
            return jsonify({"success": True, "message": "Colegio eliminado correctamente."})
        else:
            print(f"ERROR: Error inesperado al eliminar establecimiento. Status: {res_delete_est.status_code}, Response: {res_delete_est.text}")
            return jsonify({"success": False, "message": f"Error al eliminar el colegio: {res_delete_est.text}"}), 500

    except requests.exceptions.RequestException as e:
        print(f"ERROR: Error de solicitud al eliminar establecimiento: {e}")
        return jsonify({"success": False, "message": f"Error de conexión al eliminar colegio: {str(e)}"}), 500
    except Exception as e:
        print(f"ERROR: Error inesperado al eliminar establecimiento: {e}")
        return jsonify({"success": False, "message": f"Error interno del servidor al eliminar colegio: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  Eliminar NÓMINA completa (alumnos + solicitudes_correccion + nómina)
#  DELETE /admin/eliminar_nomina/<nomina_id>   ← ya existía, mejorada
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/admin/eliminar_proyecto/<proyecto_id>', methods=['DELETE'])
def eliminar_proyecto(proyecto_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        # 1. Obtener todas las nóminas del proyecto
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?proyecto_id=eq.{proyecto_id}&select=id",
            headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_nom.json() if res_nom.ok else []
        nomina_ids = [n['id'] for n in nominas]

        for nid in nomina_ids:
            # 2a. Obtener alumnos de la nómina
            res_alum = requests.get(
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?nomina_id=eq.{nid}&select=id",
                headers=SUPABASE_SERVICE_HEADERS)
            alumnos = res_alum.json() if res_alum.ok else []
            alumno_ids = [a['id'] for a in alumnos]

            # 2b. Eliminar solicitudes_correccion de esos alumnos
            if alumno_ids:
                aids = ','.join(str(i) for i in alumno_ids)
                requests.delete(
                    f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?alumno_id=in.({aids})",
                    headers=SUPABASE_SERVICE_HEADERS)

            # 2c. Eliminar estudiantes de la nómina
            requests.delete(
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?nomina_id=eq.{nid}",
                headers=SUPABASE_SERVICE_HEADERS)

            # 2d. Eliminar la nómina
            requests.delete(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nid}",
                headers=SUPABASE_SERVICE_HEADERS)

        # 3. Eliminar el proyecto
        res_del = requests.delete(
            f"{SUPABASE_URL}/rest/v1/proyectos?id=eq.{proyecto_id}",
            headers=SUPABASE_SERVICE_HEADERS)
        res_del.raise_for_status()

        return jsonify({"success": True, "message": f"Proyecto y {len(nomina_ids)} nómina(s) eliminados correctamente."})

    except Exception as e:
        print(f"ERROR eliminar_proyecto: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/admin/eliminar_nomina/<nomina_id>', methods=['DELETE'])
def eliminar_nomina(nomina_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        # 1. Obtener alumnos de la nómina
        res_alum = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?nomina_id=eq.{nomina_id}&select=id",
            headers=SUPABASE_SERVICE_HEADERS)
        alumnos = res_alum.json() if res_alum.ok else []
        alumno_ids = [a['id'] for a in alumnos]

        # 2. Eliminar solicitudes_correccion de esos alumnos (evita FK error)
        if alumno_ids:
            aids = ','.join(str(i) for i in alumno_ids)
            requests.delete(
                f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?alumno_id=in.({aids})",
                headers=SUPABASE_SERVICE_HEADERS)

        # 3. Eliminar estudiantes
        requests.delete(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?nomina_id=eq.{nomina_id}",
            headers=SUPABASE_SERVICE_HEADERS)

        # 4. Eliminar la nómina
        res_del = requests.delete(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}",
            headers=SUPABASE_SERVICE_HEADERS)
        res_del.raise_for_status()

        if res_del.status_code == 204:
            return jsonify({"success": True, "message": "Nómina y sus alumnos eliminados correctamente."})
        else:
            return jsonify({"success": False, "message": f"Error al eliminar: {res_del.text}"}), 500

    except Exception as e:
        print(f"ERROR eliminar_nomina: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ============================================================
# NUEVAS RUTAS API — Doctora (rendimiento + visitas) y
#                    Coordinadora General (stats completas)
# ============================================================

# ─────────────────────────────────────────────────────────────
# RUTA 1: Rendimiento real de la Doctora logueada
#  GET /api/doctor/performance
#  Retorna: total, completed, pending, percent, nomina_labels,
#           nomina_completed, nomina_totals, by_type (neuro/familiar)
# ─────────────────────────────────────────────────────────────
@app.route('/api/doctor/performance', methods=['GET'])
def api_doctor_performance():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_id = session.get('usuario_id')
    if not user_id:
        return jsonify({"success": False, "message": "ID de usuario no encontrado en sesión"}), 400

    # Filtro de proyecto opcional (query param ?proyecto=NombreProyecto)
    proyecto_filtro = request.args.get('proyecto', '').strip()

    try:
        # 1. Obtener nóminas donde esta doctora es principal O segunda doctora
        url_nominas_principal = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id=eq.{user_id}"
            f"&select=id,nombre_nomina,nombre_colegio,form_type,tipo_nomina,proyecto_id"
        )
        url_nominas_compartida = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id_2=eq.{user_id}"
            f"&select=id,nombre_nomina,nombre_colegio,form_type,tipo_nomina,proyecto_id"
        )
        res_n1 = requests.get(url_nominas_principal,  headers=SUPABASE_SERVICE_HEADERS)
        res_n2 = requests.get(url_nominas_compartida, headers=SUPABASE_SERVICE_HEADERS)
        lista_n1 = res_n1.json() if res_n1.ok else []
        lista_n2 = res_n2.json() if res_n2.ok else []

        # Combinar sin duplicados
        ids_vistos = set()
        nominas_raw = []
        for n in lista_n1 + lista_n2:
            if n['id'] not in ids_vistos:
                ids_vistos.add(n['id'])
                nominas_raw.append(n)

        # 2. Resolver nombres de proyectos en batch
        proyecto_ids = list({n['proyecto_id'] for n in nominas_raw if n.get('proyecto_id')})
        proyecto_map = {}  # id -> nombre_proyecto
        if proyecto_ids:
            ids_str = ','.join(str(pid) for pid in proyecto_ids)
            url_proy = (
                f"{SUPABASE_URL}/rest/v1/proyectos"
                f"?id=in.({ids_str})"
                f"&select=id,nombre_proyecto"
            )
            res_proy = requests.get(url_proy, headers=SUPABASE_SERVICE_HEADERS)
            if res_proy.ok:
                for p in res_proy.json():
                    proyecto_map[str(p['id'])] = p['nombre_proyecto']

        # 3. Enriquecer cada nómina con proyecto_nombre
        for n in nominas_raw:
            pid = n.get('proyecto_id')
            n['proyecto_nombre'] = proyecto_map.get(str(pid), 'Sin Proyecto') if pid else 'Sin Proyecto'

        nominas_todas = nominas_raw

        # 4. Aplicar filtro de proyecto si se indicó
        if proyecto_filtro:
            nominas = [n for n in nominas_todas if n['proyecto_nombre'] == proyecto_filtro]
        else:
            nominas = nominas_todas

        if not nominas:
            return jsonify({
                "success": True,
                "total": 0, "completed": 0, "pending": 0, "percent": "0%",
                "nominas_count": 0,
                "nomina_labels": [], "nomina_completed": [], "nomina_totals": [],
                "by_type": {"neurologia": 0, "medicina_familiar": 0}
            })

        nomina_labels = []
        nomina_completed = []
        nomina_totals = []
        nomina_proyectos = []
        total_global = 0
        completed_global = 0
        neuro_completed = 0
        familiar_completed = 0

        for nomina in nominas:
            nid   = nomina['id']
            label = (nomina.get('nombre_colegio') or nomina.get('nombre_nomina') or 'Nómina')[:30]
            ftype = nomina.get('form_type', '')

            total_n = get_supabase_count(f"nomina_id=eq.{nid}&estado_asistencia=in.(activo,extra)")
            comp_n  = get_supabase_count(f"nomina_id=eq.{nid}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")

            nomina_labels.append(label)
            nomina_totals.append(total_n)
            nomina_completed.append(comp_n)
            nomina_proyectos.append(nomina.get('proyecto_nombre', 'Sin Proyecto'))
            total_global     += total_n
            completed_global += comp_n

            if ftype == 'neurologia':
                neuro_completed += comp_n
            elif ftype == 'medicina_familiar':
                familiar_completed += comp_n

        pending_global = total_global - completed_global
        percent = round((completed_global / total_global * 100), 1) if total_global > 0 else 0

        # Lista de proyectos únicos asignados a esta doctora (para poblar el select en el front)
        proyectos_unicos = sorted(list({n['proyecto_nombre'] for n in nominas_todas}))

        print(f"DEBUG api_doctor_performance: total={total_global}, comp={completed_global}, pct={percent}%, proyectos={proyectos_unicos}")

        return jsonify({
            "success": True,
            "total": total_global,
            "completed": completed_global,
            "pending": pending_global,
            "percent": f"{percent}%",
            "nominas_count": len(nominas),
            "nomina_labels": nomina_labels,
            "nomina_completed": nomina_completed,
            "nomina_totals": nomina_totals,
            "nomina_proyectos": nomina_proyectos,
            "proyectos_list": proyectos_unicos,
            "by_type": {
                "neurologia": neuro_completed,
                "medicina_familiar": familiar_completed
            }
        })

    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR api_doctor_performance (requests): {e}")
        return jsonify({"success": False, "message": f"Error de conexión con BD: {str(e)}"}), 500
    except Exception as e:
        print(f"❌ ERROR api_doctor_performance: {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# RUTA 2: Visitas / eventos de la Doctora para el calendario
#  GET /api/doctor/visitas
#  Retorna: lista de eventos con fecha en formato YYYY-MM-DD
#  (el admin las crea desde /admin/agregar → tabla establecimientos
#   con columna doctora_id)
# ─────────────────────────────────────────────────────────────
@app.route('/api/doctor/visitas', methods=['GET'])
def api_doctor_visitas():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_id = session.get('usuario_id')
    if not user_id:
        return jsonify({"success": False, "message": "ID de usuario no encontrado en sesión"}), 400

    try:
        url_eventos = (
            f"{SUPABASE_URL}/rest/v1/establecimientos"
            f"?doctora_id=eq.{user_id}"
            f"&select=id,nombre,fecha,horario,cantidad_alumnos,observaciones,url_archivo"
            f"&order=fecha.asc"
        )
        res = requests.get(url_eventos, headers=SUPABASE_SERVICE_HEADERS)
        res.raise_for_status()
        eventos_raw = res.json()

        eventos_formateados = []
        for ev in eventos_raw:
            fecha_raw  = ev.get('fecha', '') or ''
            fecha_norm = fecha_raw

            # Normalizar a YYYY-MM-DD si viene en otro formato
            if fecha_raw and '/' in fecha_raw:
                try:
                    fecha_norm = datetime.strptime(fecha_raw, '%d/%m/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    fecha_norm = fecha_raw

            eventos_formateados.append({
                "id":               ev.get('id', ''),
                "nombre":           ev.get('nombre', ''),
                "fecha":            fecha_norm,
                "horario":          ev.get('horario', ''),
                "cantidad_alumnos": ev.get('cantidad_alumnos'),
                "observaciones":    ev.get('observaciones', ''),
                "url_archivo":      ev.get('url_archivo', '')
            })

        print(f"DEBUG api_doctor_visitas: {len(eventos_formateados)} eventos para doctora {user_id}")
        return jsonify({"success": True, "eventos": eventos_formateados})

    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR api_doctor_visitas (requests): {e}")
        return jsonify({"success": False, "message": f"Error de conexión con BD: {str(e)}"}), 500
    except Exception as e:
        print(f"❌ ERROR api_doctor_visitas: {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# RUTA 3: Estadísticas completas para la Coordinadora General
#  GET /api/coordinadora/stats
#  Retorna: totales, desglose neuro/familiar, datos por día y
#           semana (últimos 30 días / 12 semanas), ranking de
#           doctoras y top de establecimientos
# ─────────────────────────────────────────────────────────────
@app.route('/api/coordinadora/stats', methods=['GET'])
def api_coordinadora_stats():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_role = session.get('usuario')
    user_id   = session.get('usuario_id')

    if user_role not in ('coordinadora', 'admin'):
        return jsonify({"success": False, "message": "Acceso denegado"}), 403

    try:
        # ── 1. Nóminas de la coordinadora (o todas si es admin) ──────────
        if user_role == 'coordinadora':
            url_nominas = (
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?coord_general_id=eq.{user_id}"
                f"&select=id,form_type,nombre_colegio,nombre_nomina,doctora_id,tipo_nomina"
            )
        else:
            url_nominas = (
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?select=id,form_type,nombre_colegio,nombre_nomina,doctora_id,tipo_nomina"
            )

        res_nominas = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS)
        res_nominas.raise_for_status()
        nominas = res_nominas.json()

        # Respuesta vacía si no hay nóminas
        if not nominas:
            return jsonify({
                "success": True,
                "total": 0, "completed": 0, "pending": 0, "percent": 0,
                "neuro_total": 0, "neuro_completed": 0,
                "familiar_total": 0, "familiar_completed": 0,
                "por_dia": {}, "por_semana": {},
                "doctoras_ranking": [],
                "establecimientos": []
            })

        nomina_ids   = [n['id'] for n in nominas]
        nominas_set  = set(nomina_ids)

        # ── 2. Conteos globales usando get_supabase_count ──
        total_global     = 0
        completed_global = 0
        neuro_total      = 0
        neuro_completed  = 0
        familiar_total   = 0
        familiar_completed = 0

        for nomina in nominas:
            nid   = nomina['id']
            ftype = nomina.get('form_type', '') or ''
            tipo  = (nomina.get('tipo_nomina') or '').lower()

            t = get_supabase_count(f"nomina_id=eq.{nid}&estado_asistencia=in.(activo,extra)")
            c = get_supabase_count(f"nomina_id=eq.{nid}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")

            total_global     += t
            completed_global += c

            # Detectar tipo por form_type primero, luego por tipo_nomina
            if ftype == 'neurologia' or 'neuro' in tipo:
                neuro_total     += t
                neuro_completed += c
            elif ftype == 'medicina_familiar' or 'familiar' in tipo or 'medicina' in tipo:
                familiar_total     += t
                familiar_completed += c

        pending_global = total_global - completed_global
        percent = round((completed_global / total_global * 100), 1) if total_global > 0 else 0

        # ── 3. Datos diarios y semanales (misma lógica que admin) ──────────
        from collections import defaultdict
        hoy      = date.today()
        hace_30  = hoy - timedelta(days=30)
        hace_12s = hoy - timedelta(weeks=12)

        por_dia    = defaultdict(lambda: {"neurologia": 0, "medicina_familiar": 0, "total": 0})
        por_semana = defaultdict(lambda: {"neurologia": 0, "medicina_familiar": 0, "total": 0})

        for nom in nominas:
            nid    = nom['id']
            ftype  = (nom.get('form_type') or '').lower().strip()
            tipo_n = (nom.get('tipo_nomina') or '').lower().strip()
            if 'neuro' in ftype or 'neuro' in tipo_n:
                tipo_key = 'neurologia'
            elif 'familiar' in ftype or 'medicina' in ftype or 'familiar' in tipo_n:
                tipo_key = 'medicina_familiar'
            else:
                tipo_key = 'otro'
            try:
                url_f = (
                    f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                    f"?nomina_id=eq.{nid}"
                    f"&evaluado_flag=eq.true"
                    f"&estado_asistencia=in.(activo,extra)"
                    f"&select=fecha_evaluacion"
                )
                res_f = requests.get(url_f, headers=SUPABASE_SERVICE_HEADERS)
                if not res_f.ok:
                    continue
                for row in res_f.json():
                    fe = (row.get('fecha_evaluacion') or '').strip()
                    if not fe:
                        continue
                    fe_date = None
                    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
                        try:
                            fe_date = datetime.strptime(fe[:10], fmt).date()
                            break
                        except ValueError:
                            continue
                    if not fe_date:
                        continue
                    delta = (hoy - fe_date).days
                    if 0 <= delta <= 30:
                        k = str(fe_date)
                        por_dia[k]["total"] += 1
                        if tipo_key != 'otro':
                            por_dia[k][tipo_key] += 1
                    delta_w = delta // 7
                    if 0 <= delta_w < 12:
                        iso_w = fe_date.isocalendar()
                        ks = f"S{iso_w[1]:02d}/{iso_w[0]}"
                        por_semana[ks]["total"] += 1
                        if tipo_key != 'otro':
                            por_semana[ks][tipo_key] += 1
            except Exception:
                pass

        # Ordenar por clave cronológicamente
        por_dia_sorted    = dict(sorted(por_dia.items()))
        por_semana_sorted = dict(sorted(por_semana.items()))

        # ── 4. Ranking de doctoras ────────────────────────────────────────
        # Agrupar nóminas por doctora_id
        doctoras_nominas = defaultdict(list)
        for n in nominas:
            did = n.get('doctora_id')
            if did:
                doctoras_nominas[did].append(n['id'])

        # Obtener nombres de doctoras de la tabla doctoras
        doctoras_ranking = []
        if doctoras_nominas:
            url_docs = f"{SUPABASE_URL}/rest/v1/doctoras?select=id,usuario,nombre"
            res_docs = requests.get(url_docs, headers=SUPABASE_SERVICE_HEADERS)
            doctoras_info = {d['id']: d for d in (res_docs.json() if res_docs.ok else [])}

            for did, nids in doctoras_nominas.items():
                doc_total = sum(get_supabase_count(f"nomina_id=eq.{nid}&estado_asistencia=in.(activo,extra)") for nid in nids)
                doc_comp  = sum(get_supabase_count(f"nomina_id=eq.{nid}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)") for nid in nids)
                doc_pct   = round((doc_comp / doc_total * 100), 1) if doc_total > 0 else 0
                # Usar 'usuario' o 'nombre' para mostrar
                doc_info  = doctoras_info.get(did, {})
                doc_name  = doc_info.get('nombre') or doc_info.get('usuario') or f'Doctora {str(did)[:6]}'
                doctoras_ranking.append({
                    "id":        did,
                    "nombre":    doc_name,
                    "total":     doc_total,
                    "completed": doc_comp,
                    "percent":   doc_pct
                })

        doctoras_ranking.sort(key=lambda x: x['percent'], reverse=True)

        # ── 5. Top establecimientos (agrupados por nombre_colegio) ────────
        est_data = defaultdict(lambda: {"total": 0, "completed": 0})
        for nomina in nominas:
            nid   = nomina['id']
            ename = nomina.get('nombre_colegio') or nomina.get('nombre_nomina') or 'Sin nombre'
            t = get_supabase_count(f"nomina_id=eq.{nid}&estado_asistencia=in.(activo,extra)")
            c = get_supabase_count(f"nomina_id=eq.{nid}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")
            est_data[ename]["total"]     += t
            est_data[ename]["completed"] += c

        establecimientos = [
            {
                "nombre":    k,
                "total":     v["total"],
                "completed": v["completed"],
                "percent":   round((v["completed"] / v["total"] * 100), 1) if v["total"] > 0 else 0
            }
            for k, v in est_data.items()
        ]
        establecimientos.sort(key=lambda x: x["completed"], reverse=True)

        print(f"DEBUG api_coordinadora_stats: total={total_global}, comp={completed_global}, doctoras={len(doctoras_ranking)}")

        return jsonify({
            "success": True,
            "total":              total_global,
            "completed":          completed_global,
            "pending":            pending_global,
            "percent":            percent,
            "neuro_total":        neuro_total,
            "neuro_completed":    neuro_completed,
            "familiar_total":     familiar_total,
            "familiar_completed": familiar_completed,
            "por_dia":            por_dia_sorted,
            "por_semana":         por_semana_sorted,
            "doctoras_ranking":   doctoras_ranking[:10],
            "establecimientos":   establecimientos[:10]
        })

    except requests.exceptions.RequestException as e:
        print(f"❌ ERROR api_coordinadora_stats (requests): {e}")
        return jsonify({"success": False, "message": f"Error de conexión con BD: {str(e)}"}), 500
    except Exception as e:
        print(f"❌ ERROR api_coordinadora_stats: {e}")
        return jsonify({"success": False, "message": f"Error interno: {str(e)}"}), 500


# ─────────────────────────────────────────────────────────────
# RUTA: Stats de todas las nóminas de un proyecto
#  GET /api/admin/proyecto_stats/<project_id>
#  Retorna: [{nomina_id, total, evaluados, pct}]
# ─────────────────────────────────────────────────────────────
@app.route('/api/admin/proyecto_stats/<project_id>', methods=['GET'])
def api_admin_proyecto_stats(project_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        # 1. Obtener nóminas del proyecto
        url_n = (f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                 f"?proyecto_id=eq.{project_id}&select=id")
        res_n = requests.get(url_n, headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_n.json() if res_n.ok else []

        result = []
        proj_total = 0
        proj_eval  = 0

        for nom in nominas:
            nid   = nom['id']
            total = get_supabase_count(f"nomina_id=eq.{nid}&estado_asistencia=in.(activo,extra)")
            eval_ = get_supabase_count(f"nomina_id=eq.{nid}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")
            pct   = round(eval_ / total * 100, 1) if total > 0 else 0
            proj_total += total
            proj_eval  += eval_
            result.append({"nomina_id": nid, "total": total, "evaluados": eval_, "pct": pct})

        proj_pct = round(proj_eval / proj_total * 100, 1) if proj_total > 0 else 0
        return jsonify({
            "success": True,
            "nominas": result,
            "proyecto": {"total": proj_total, "evaluados": proj_eval, "pct": proj_pct}
        })
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ─────────────────────────────────────────────────────────────
# RUTA: Visitas de una doctora específica (para calendario admin)
#  GET /api/admin/visitas_doctora/<doctor_id>
# ─────────────────────────────────────────────────────────────
@app.route('/api/admin/visitas_doctora/<doctor_id>', methods=['GET'])
def api_admin_visitas_doctora(doctor_id):
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        url = (
            f"{SUPABASE_URL}/rest/v1/establecimientos"
            f"?doctora_id=eq.{doctor_id}"
            f"&select=id,nombre,fecha,horario,cantidad_alumnos,observaciones"
            f"&order=fecha.asc"
        )
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        eventos_raw = res.json() if res.ok else []

        eventos = []
        for ev in eventos_raw:
            fecha_raw = ev.get('fecha', '') or ''
            if fecha_raw and '/' in fecha_raw:
                try:
                    fecha_raw = datetime.strptime(fecha_raw, '%d/%m/%Y').strftime('%Y-%m-%d')
                except ValueError:
                    pass
            eventos.append({
                "id":               ev.get('id', ''),
                "nombre":           ev.get('nombre', ''),
                "fecha":            fecha_raw,
                "horario":          ev.get('horario', ''),
                "cantidad_alumnos": ev.get('cantidad_alumnos'),
                "observaciones":    ev.get('observaciones', '')
            })

        return jsonify({"success": True, "eventos": eventos})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


# ─────────────────────────────────────────────────────────────
# RUTA: Reporte detallado doctora (alumnos por nomina)
#  GET /api/doctor/reporte_detalle
# ─────────────────────────────────────────────────────────────
@app.route('/api/doctor/reporte_detalle', methods=['GET'])
def api_doctor_reporte_detalle():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_id = session.get('usuario_id')
    if not user_id:
        return jsonify({"success": False, "message": "ID de usuario no encontrado"}), 400

    try:
        # 1. Nóminas de la doctora
        url_nominas = (
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?doctora_id=eq.{user_id}"
            f"&select=id,nombre_nomina,nombre_colegio,form_type"
        )
        res_n = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_n.json() if res_n.ok else []

        detalles = []
        global_total = 0
        global_evaluados = 0

        for nom in nominas:
            url_e = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(activo,extra)"
                f"&select=id,nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia,fecha_relleno"
                f"&order=nombre.asc"
            )
            res_e = requests.get(url_e, headers=SUPABASE_SERVICE_HEADERS)
            alumnos = res_e.json() if res_e.ok else []

            url_aus = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(no_asiste_reemplazado,no_asiste_sin_reemplazo)"
                f"&select=nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia"
                f"&order=nombre.asc"
            )
            res_aus = requests.get(url_aus, headers=SUPABASE_SERVICE_HEADERS)
            ausentes = res_aus.json() if res_aus.ok else []

            ev_count    = len([a for a in alumnos if a.get('evaluado_flag') is True])
            total_count = len(alumnos)
            global_total     += total_count
            global_evaluados += ev_count

            colegio = nom.get('nombre_colegio') or nom.get('nombre_nomina') or 'Sin nombre'
            detalles.append({
                "colegio":   colegio,
                "alumnos":   alumnos,
                "ausentes":  ausentes,
                "total":     total_count,
                "evaluados": ev_count
            })

        return jsonify({
            "success": True,
            "resumen": {
                "total":      global_total,
                "evaluados":  global_evaluados,
                "pendientes": global_total - global_evaluados,
                "porcentaje": f"{round(global_evaluados/global_total*100,1) if global_total>0 else 0}%"
            },
            "data": detalles
        })

    except Exception as e:
        print(f"ERROR api_doctor_reporte_detalle: {e}")
        return jsonify({"success": False, "error": str(e)})


# ─────────────────────────────────────────────────────────────
# RUTA: Reporte detallado coordinadora (alumnos por nomina)
#  GET /api/coordinadora/reporte_detalle
# ─────────────────────────────────────────────────────────────
@app.route('/api/coordinadora/reporte_detalle', methods=['GET'])
def api_coordinadora_reporte_detalle():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_role = session.get('usuario')
    user_id   = session.get('usuario_id')

    if user_role not in ('coordinadora', 'admin'):
        return jsonify({"success": False, "message": "Acceso denegado"}), 403

    try:
        # 1. Nóminas según rol
        if user_role == 'coordinadora':
            url_nominas = (
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?coord_general_id=eq.{user_id}"
                f"&select=id,nombre_nomina,nombre_colegio,form_type,doctora_id"
            )
        else:
            url_nominas = (
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?select=id,nombre_nomina,nombre_colegio,form_type,doctora_id"
            )

        res_n = requests.get(url_nominas, headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_n.json() if res_n.ok else []

        # 2. Lookup nombres doctoras
        url_docs = f"{SUPABASE_URL}/rest/v1/doctoras?select=id,nombre,usuario"
        res_docs = requests.get(url_docs, headers=SUPABASE_SERVICE_HEADERS)
        doctoras_map = {}
        if res_docs.ok:
            for d in res_docs.json():
                doctoras_map[str(d['id'])] = d.get('nombre') or d.get('usuario', 'Doctora')

        detalles = []
        global_total = 0
        global_evaluados = 0

        for nom in nominas:
            url_e = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(activo,extra)"
                f"&select=id,nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia,fecha_relleno"
                f"&order=nombre.asc"
            )
            res_e = requests.get(url_e, headers=SUPABASE_SERVICE_HEADERS)
            alumnos = res_e.json() if res_e.ok else []

            url_aus = (
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nom['id']}"
                f"&estado_asistencia=in.(no_asiste_reemplazado,no_asiste_sin_reemplazo)"
                f"&select=nombre,rut,evaluado_flag,estado_asistencia,motivo_ausencia"
                f"&order=nombre.asc"
            )
            res_aus = requests.get(url_aus, headers=SUPABASE_SERVICE_HEADERS)
            ausentes = res_aus.json() if res_aus.ok else []

            ev_count    = len([a for a in alumnos if a.get('evaluado_flag') is True])
            total_count = len(alumnos)
            global_total     += total_count
            global_evaluados += ev_count

            colegio    = nom.get('nombre_colegio') or nom.get('nombre_nomina') or 'Sin nombre'
            doc_id     = str(nom.get('doctora_id') or '')
            doc_nombre = doctoras_map.get(doc_id, '')
            label = colegio + ('  (' + doc_nombre + ')' if doc_nombre else '')

            detalles.append({
                "colegio":   label,
                "alumnos":   alumnos,
                "ausentes":  ausentes,
                "total":     total_count,
                "evaluados": ev_count
            })

        return jsonify({
            "success": True,
            "resumen": {
                "total":      global_total,
                "evaluados":  global_evaluados,
                "pendientes": global_total - global_evaluados,
                "porcentaje": f"{round(global_evaluados/global_total*100,1) if global_total>0 else 0}%"
            },
            "data": detalles
        })

    except Exception as e:
        print(f"ERROR api_coordinadora_reporte_detalle: {e}")
        return jsonify({"success": False, "error": str(e)})



# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Stats de una nómina individual (para barras de progreso doctora)
#  GET /api/doctor/nomina_stats/<nomina_id>
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/doctor/nomina_stats/<nomina_id>', methods=['GET'])
def api_doctor_nomina_stats(nomina_id):
    """Retorna total y evaluados de una nómina específica. Accesible por doctora y admin."""
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    try:
        total    = get_supabase_count(f"nomina_id=eq.{nomina_id}&estado_asistencia=in.(activo,extra)")
        evaluados = get_supabase_count(f"nomina_id=eq.{nomina_id}&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")
        pct = round(evaluados / total * 100, 1) if total > 0 else 0
        return jsonify({"success": True, "total": total, "evaluados": evaluados, "pct": pct})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Admin descarga listado de alumnos en EXCEL por nómina
#  POST /api/admin/listado_alumnos_excel
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/listado_alumnos_excel', methods=['POST'])
def api_admin_listado_excel():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        data = request.get_json()
        nomina_id = data.get('nomina_id')
        filtro    = data.get('filtro', 'todos')
        if not nomina_id:
            return jsonify({"success": False, "message": "nomina_id requerido"}), 400

        # Obtener datos de la nómina
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}&select=nombre_nomina,nombre_colegio",
            headers=SUPABASE_SERVICE_HEADERS)
        nomina_info = res_nom.json()[0] if res_nom.ok and res_nom.json() else {}
        nombre_nomina = nomina_info.get('nombre_colegio') or nomina_info.get('nombre_nomina') or 'Nómina'

        # Obtener alumnos
        res_alumnos = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&estado_asistencia=in.(activo,extra)"
            f"&select=nombre,rut,fecha_nacimiento,sexo,fecha_evaluacion,evaluado_flag,estado_asistencia"
            f"&order=nombre.asc",
            headers=SUPABASE_SERVICE_HEADERS)
        todos_alumnos = res_alumnos.json() if res_alumnos.ok else []

        # Aplicar filtro
        if filtro == 'evaluados':
            alumnos = [a for a in todos_alumnos if a.get('evaluado_flag')]
            filtro_label = 'Solo evaluados'
        elif filtro == 'pendientes':
            alumnos = [a for a in todos_alumnos if not a.get('evaluado_flag')]
            filtro_label = 'Solo pendientes'
        else:
            alumnos = todos_alumnos
            filtro_label = 'Todos los alumnos'

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado Alumnos"

        # Estilos
        hdr_fill  = PatternFill("solid", fgColor="0F3460")
        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        body_font = Font(name="Calibri", size=10)
        alt_fill  = PatternFill("solid", fgColor="EEF4FF")
        border    = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Título
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Listado de Alumnos — {nombre_nomina}"
        ws['A1'].font = Font(name="Calibri", bold=True, size=14, color="0F3460")
        ws['A1'].alignment = center_align
        ws.row_dimensions[1].height = 26

        ws.merge_cells('A2:H2')
        from datetime import datetime
        ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} · Filtro: {filtro_label} · CardioHome"
        ws['A2'].font = Font(name="Calibri", size=9, color="888888", italic=True)
        ws['A2'].alignment = center_align

        # Encabezados — agregada columna Tipo
        headers    = ['#', 'Nombre', 'RUT', 'Fecha Nacimiento', 'Sexo', 'Tipo', 'Fecha Evaluación', 'Estado']
        col_widths = [5,   35,       14,    16,                 8,     14,     16,                  14    ]
        for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center_align
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = w
        ws.row_dimensions[4].height = 20

        # Datos activos + extra
        for i, alumno in enumerate(alumnos):
            row  = i + 5
            fill = alt_fill if i % 2 == 0 else PatternFill()
            tipo = 'Reemplazo' if alumno.get('estado_asistencia') == 'extra' else 'Titular'
            values = [
                i + 1,
                alumno.get('nombre', ''),
                alumno.get('rut', ''),
                alumno.get('fecha_nacimiento', '') or '',
                'M' if alumno.get('sexo') == 'M' else 'F' if alumno.get('sexo') == 'F' else '',
                tipo,
                alumno.get('fecha_evaluacion', '') or '',
                'Evaluado' if alumno.get('evaluado_flag') else 'Pendiente'
            ]
            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = body_font
                cell.fill = fill
                cell.alignment = center_align if col_idx in [1,3,4,5,6,7,8] else Alignment(vertical='center')
                cell.border = border
                # Color especial para reemplazo
                if col_idx == 6 and tipo == 'Reemplazo':
                    cell.font = Font(name="Calibri", size=10, color="7C3AED", bold=True)
                # Color verde/naranja para estado
                if col_idx == 8:
                    cell.font = Font(name="Calibri", size=10, bold=True,
                                     color="059669" if alumno.get('evaluado_flag') else "D97706")
            ws.row_dimensions[row].height = 16

        # Fila totales
        total_row = len(alumnos) + 5
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name="Calibri", size=10)
        ws.cell(row=total_row, column=2, value=len(alumnos)).font = Font(bold=True, name="Calibri", size=10)
        evaluated = sum(1 for a in alumnos if a.get('evaluado_flag'))
        ws.cell(row=total_row, column=8, value=f"{evaluated}/{len(alumnos)} evaluados").font = Font(bold=True, name="Calibri", size=10, color="059669")

        # ── Hoja 2: Alumnos Anulados ──────────────────────────────────────
        res_ausentes = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&estado_asistencia=in.(no_asiste_reemplazado,no_asiste_sin_reemplazo)"
            f"&select=nombre,rut,fecha_nacimiento,sexo,estado_asistencia,motivo_ausencia"
            f"&order=nombre.asc",
            headers=SUPABASE_SERVICE_HEADERS)
        ausentes = res_ausentes.json() if res_ausentes.ok else []

        if ausentes:
            ws2 = wb.create_sheet(title="Anulados")
            ws2.merge_cells('A1:F1')
            ws2['A1'] = f"Alumnos Anulados — {nombre_nomina}"
            ws2['A1'].font = Font(name="Calibri", bold=True, size=13, color="B91C1C")
            ws2['A1'].alignment = center_align

            aus_hdr_fill = PatternFill("solid", fgColor="B91C1C")
            aus_headers  = ['#', 'Nombre', 'RUT', 'Fecha Nacimiento', 'Tipo Anulación', 'Motivo']
            aus_widths   = [5,   35,       14,    16,                  20,               40     ]
            for ci, (h, w) in enumerate(zip(aus_headers, aus_widths), 1):
                c = ws2.cell(row=3, column=ci, value=h)
                c.font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
                c.fill = aus_hdr_fill
                c.alignment = center_align
                c.border = border
                ws2.column_dimensions[c.column_letter].width = w

            for i, a in enumerate(ausentes):
                r = i + 4
                tipo_an = 'Reemplazado' if a.get('estado_asistencia') == 'no_asiste_reemplazado' else 'Sin reemplazo'
                vals = [i+1, a.get('nombre',''), a.get('rut',''),
                        a.get('fecha_nacimiento','') or '', tipo_an, a.get('motivo_ausencia','') or '']
                for ci, v in enumerate(vals, 1):
                    c = ws2.cell(row=r, column=ci, value=v)
                    c.font = body_font
                    c.fill = alt_fill if i % 2 == 0 else PatternFill()
                    c.alignment = center_align if ci in [1,3,4,5] else Alignment(vertical='center')
                    c.border = border
                ws2.row_dimensions[r].height = 16

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = nombre_nomina.replace(' ', '_').replace('/', '-')
        from flask import send_file
        return send_file(
            output,
            as_attachment=True,
            download_name=f"Listado_{safe_name}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        print(f"ERROR listado_alumnos_excel: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Admin descarga listado de alumnos en PDF por nómina
#  POST /api/admin/listado_alumnos_pdf
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/listado_alumnos_pdf', methods=['POST'])
def api_admin_listado_pdf():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        data = request.get_json()
        nomina_id = data.get('nomina_id')
        filtro    = data.get('filtro', 'todos')   # 'todos' | 'evaluados' | 'pendientes'
        if not nomina_id:
            return jsonify({"success": False, "message": "nomina_id requerido"}), 400

        # Obtener datos de la nómina
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas?id=eq.{nomina_id}&select=nombre_nomina,nombre_colegio",
            headers=SUPABASE_SERVICE_HEADERS)
        nomina_info = res_nom.json()[0] if res_nom.ok and res_nom.json() else {}
        nombre_nomina = nomina_info.get('nombre_colegio') or nomina_info.get('nombre_nomina') or 'Nómina'

        # Obtener alumnos
        res_alumnos = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&estado_asistencia=in.(activo,extra)"
            f"&select=nombre,rut,fecha_nacimiento,sexo,fecha_evaluacion,evaluado_flag"
            f"&order=nombre.asc",
            headers=SUPABASE_SERVICE_HEADERS)
        todos_alumnos = res_alumnos.json() if res_alumnos.ok else []

        # Aplicar filtro
        if filtro == 'evaluados':
            alumnos = [a for a in todos_alumnos if a.get('evaluado_flag')]
            filtro_label = 'Solo evaluados'
        elif filtro == 'pendientes':
            alumnos = [a for a in todos_alumnos if not a.get('evaluado_flag')]
            filtro_label = 'Solo pendientes'
        else:
            alumnos   = todos_alumnos
            filtro_label = 'Todos los alumnos'

        from datetime import datetime
        fecha_gen = datetime.now().strftime('%d/%m/%Y %H:%M')
        evaluados_total = sum(1 for a in todos_alumnos if a.get('evaluado_flag'))

        rows_html = ''
        for i, alumno in enumerate(alumnos):
            bg = '#f0f6ff' if i % 2 == 0 else 'white'
            estado_color = '#059669' if alumno.get('evaluado_flag') else '#dc2626'
            estado_text  = 'Evaluado' if alumno.get('evaluado_flag') else 'Pendiente'
            rows_html += f"""
            <tr style="background:{bg};">
                <td style="text-align:center;">{i+1}</td>
                <td>{alumno.get('nombre','')}</td>
                <td style="text-align:center;">{alumno.get('rut','')}</td>
                <td style="text-align:center;">{alumno.get('fecha_nacimiento','') or ''}</td>
                <td style="text-align:center;">{'M' if alumno.get('sexo')=='M' else 'F' if alumno.get('sexo')=='F' else ''}</td>
                <td style="text-align:center;">{alumno.get('fecha_evaluacion','') or ''}</td>
                <td style="text-align:center;color:{estado_color};font-weight:700;">{estado_text}</td>
            </tr>"""

        pct = round(evaluados_total / len(todos_alumnos) * 100) if todos_alumnos else 0

        html = f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Listado — {nombre_nomina}</title>
<style>
  /* Reset print conflicts — previene duplicación en Chrome */
  *, *::before, *::after {{ box-sizing: border-box; margin: 0; padding: 0; }}
  html, body {{ width: 100%; height: auto; overflow: visible; }}
  body {{ font-family: Arial, sans-serif; font-size: 11px; color: #1a2332; padding: 20px; }}
  .header {{ background: #0f3460; color: white; padding: 16px 20px; border-radius: 8px; margin-bottom: 16px; }}
  .header h1 {{ font-size: 15px; margin-bottom: 4px; }}
  .header p {{ font-size: 10px; opacity: .75; margin: 0; }}
  .kpis {{ display: flex; gap: 12px; margin-bottom: 14px; }}
  .kpi {{ background: #f0f6ff; border: 1px solid #d4e5f5; border-radius: 8px; padding: 10px 16px; flex: 1; text-align: center; }}
  .kpi .n {{ font-size: 20px; font-weight: 900; color: #0f3460; }}
  .kpi .l {{ font-size: 9px; color: #94a3b8; text-transform: uppercase; letter-spacing: .5px; }}
  .filtro-badge {{ display: inline-block; background: #e0f2fe; color: #0369a1; border: 1px solid #bae6fd;
      border-radius: 20px; padding: 3px 10px; font-size: 9px; font-weight: 700; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
  th {{ background: #0f3460; color: white; padding: 8px 10px; font-size: 10px; text-align: center; }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #e8f0f6; font-size: 10px; word-break: break-word; }}
  .footer {{ margin-top: 16px; font-size: 9px; color: #94a3b8; text-align: center;
      border-top: 1px solid #e8f0f6; padding-top: 10px; }}
  @media print {{
    html, body {{ height: auto !important; overflow: visible !important; }}
    /* Evita que Chrome repita el body al imprimir */
    body {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
    .no-print {{ display: none !important; }}
    table {{ page-break-inside: auto; }}
    tr {{ page-break-inside: avoid; page-break-after: auto; }}
    thead {{ display: table-header-group; }}
    tfoot {{ display: table-footer-group; }}
  }}
</style></head><body>
<div class="header">
  <h1>Listado de Alumnos — {nombre_nomina}</h1>
  <p>Generado el {fecha_gen} &nbsp;·&nbsp; CardioHome</p>
</div>
<div class="filtro-badge">Filtro: {filtro_label}</div>
<div class="kpis">
  <div class="kpi"><div class="n">{len(todos_alumnos)}</div><div class="l">Total nómina</div></div>
  <div class="kpi"><div class="n" style="color:#059669">{evaluados_total}</div><div class="l">Evaluados</div></div>
  <div class="kpi"><div class="n" style="color:#dc2626">{len(todos_alumnos)-evaluados_total}</div><div class="l">Pendientes</div></div>
  <div class="kpi"><div class="n" style="color:#7c3aed">{pct}%</div><div class="l">Avance</div></div>
  <div class="kpi"><div class="n" style="color:#0369a1">{len(alumnos)}</div><div class="l">En este listado</div></div>
</div>
<table>
  <thead><tr>
    <th style="width:30px">#</th>
    <th style="text-align:left">Nombre</th>
    <th style="width:80px">RUT</th>
    <th style="width:75px">Fecha Nac.</th>
    <th style="width:35px">Sexo</th>
    <th style="width:75px">Fecha Eval.</th>
    <th style="width:70px">Estado</th>
  </tr></thead>
  <tbody>{rows_html}</tbody>
</table>
<div class="footer">CardioHome SpA &nbsp;·&nbsp; Jorge Enrique Miranda Kirk &nbsp;·&nbsp; RUT 77.028.328-0</div>
</body></html>"""

        from flask import Response
        safe_name = nombre_nomina.replace(' ', '_').replace('/', '-')
        return Response(
            html,
            content_type='text/html; charset=utf-8',
            headers={
                'Content-Disposition': f'inline; filename="Listado_{safe_name}.html"',
                'Cache-Control': 'no-store'
            })

    except Exception as e:
        print(f"ERROR listado_alumnos_pdf: {e}")
        return jsonify({"success": False, "message": str(e)}), 500

# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Admin sube el formulario corregido de un alumno específico
#  POST /api/admin/subir_documento_corregido
#  Body: multipart/form-data — alumno_id, solicitud_id, file
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/admin/subir_documento_corregido', methods=['POST'])
def api_admin_subir_documento_corregido():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        solicitud_id = request.form.get('solicitud_id')
        alumno_id    = request.form.get('alumno_id')
        file         = request.files.get('file')

        if not solicitud_id or not file:
            return jsonify({"success": False, "message": "Faltan datos requeridos"}), 400

        # Subir archivo al storage de Supabase
        bucket_name   = "documentos-corregidos"
        ext           = os.path.splitext(file.filename)[1].lower() or '.pdf'
        storage_path  = f"correcciones/{solicitud_id}{ext}"
        file_content  = file.read()

        upload_url = f"{SUPABASE_URL}/storage/v1/object/{bucket_name}/{storage_path}"
        upload_res = requests.post(
            upload_url,
            headers={
                "apikey":        SUPABASE_SERVICE_KEY,
                "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
                "Content-Type":  file.mimetype or 'application/octet-stream',
                "x-upsert":      "true"
            },
            data=file_content)

        # URL pública del documento
        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{bucket_name}/{storage_path}"

        # Actualizar la solicitud: estado Aprobada + url_documento_corregido
        from datetime import date
        patch_res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?id=eq.{solicitud_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json={
                "estado":                  "Aprobada",
                "fecha_resolucion":        str(date.today()),
                "url_documento_corregido": public_url
            })
        patch_res.raise_for_status()

        # También actualizar el estudiante si se proporcionó alumno_id
        if alumno_id:
            requests.patch(
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{alumno_id}",
                headers=SUPABASE_SERVICE_HEADERS,
                json={"url_formulario_corregido": public_url})

        return jsonify({"success": True, "url": public_url,
                        "message": "Documento subido y solicitud marcada como Aprobada"})

    except Exception as e:
        print(f"ERROR subir_documento_corregido: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Coordinador escuela consulta correcciones resueltas de su colegio
#  GET /api/coordinador_escuela/correcciones_resueltas/<school_id>
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/coordinador_escuela/correcciones_resueltas/<school_id>', methods=['GET'])
def api_correcciones_resueltas_escuela(school_id):
    if session.get('usuario') != 'coordinador_escuela':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        # Obtener nominas del colegio
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?nombre_colegio=eq.{school_id}"
            f"&coord_escuela_id=eq.{session.get('usuario_id')}"
            f"&select=id",
            headers=SUPABASE_SERVICE_HEADERS)
        nominas = res_nom.json() if res_nom.ok else []
        if not nominas:
            return jsonify({"success": True, "data": []})

        nomina_ids = [n['id'] for n in nominas]
        ids_str    = ','.join(str(i) for i in nomina_ids)

        # Obtener TODOS los alumnos de estas nóminas para poder filtrar por alumno_id
        res_alumnos = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=in.({ids_str})"
            f"&select=id,nombre,rut",
            headers=SUPABASE_SERVICE_HEADERS)
        alumnos_list = res_alumnos.json() if res_alumnos.ok else []
        alumno_ids_set = {str(a['id']) for a in alumnos_list}
        alumno_info    = {str(a['id']): a for a in alumnos_list}

        # Obtener solicitudes resueltas de estos alumnos
        result = []
        if alumno_ids_set:
            aid_str = ','.join(alumno_ids_set)
            url_sol = (
                f"{SUPABASE_URL}/rest/v1/solicitudes_correccion"
                f"?estado=in.(Aprobada,Rechazada)"
                f"&alumno_id=in.({aid_str})"
                f"&select=id,alumno_id,detalles,fecha_resolucion,"
                f"url_documento_corregido,respuesta_admin,notificacion_vista,estado"
                f"&order=fecha_resolucion.desc"
            )
            res_sol = requests.get(url_sol, headers=SUPABASE_SERVICE_HEADERS)
            solicitudes = res_sol.json() if res_sol.ok else []
            for s in solicitudes:
                aid  = str(s.get('alumno_id', ''))
                info = alumno_info.get(aid, {})
                result.append({
                    'id':                 s['id'],
                    'alumno_id':          aid,
                    'alumno_nombre':      info.get('nombre', 'N/A'),
                    'alumno_rut':         info.get('rut', 'N/A'),
                    'detalles':           s.get('detalles', ''),
                    'fecha_resolucion':   s.get('fecha_resolucion', ''),
                    'url_documento':      s.get('url_documento_corregido', ''),
                    'respuesta_admin':    s.get('respuesta_admin') or '',
                    'notificacion_vista': s.get('notificacion_vista', True),
                    'estado':             s.get('estado', 'Aprobada'),
                })

        return jsonify({"success": True, "data": result})

    except Exception as e:
        print(f"ERROR correcciones_resueltas_escuela: {e}")
        return jsonify({"success": False, "error": str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Marcar ausencia / reactivar alumno
#  POST /api/estudiante/estado_ausencia
#  Body JSON: { estudiante_id, estado_asistencia, motivo_ausencia }
#  estados válidos: 'activo' | 'no_asiste_reemplazado' | 'no_asiste_sin_reemplazo'
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/estudiante/estado_ausencia', methods=['POST'])
def api_estudiante_estado_ausencia():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "Body JSON requerido"}), 400

        estudiante_id    = data.get('estudiante_id', '').strip()
        estado           = data.get('estado_asistencia', '').strip()
        motivo           = data.get('motivo_ausencia')

        estados_validos = {'activo', 'no_asiste_reemplazado', 'no_asiste_sin_reemplazo', 'extra'}
        if not estudiante_id or estado not in estados_validos:
            return jsonify({"success": False,
                            "message": f"Parámetros inválidos. estado='{estado}', id='{estudiante_id}'"}), 400

        # Validar formato UUID
        try:
            uuid.UUID(estudiante_id)
        except ValueError:
            return jsonify({"success": False,
                            "message": f"ID de estudiante no es UUID válido: '{estudiante_id}'"}), 400

        payload = {"estado_asistencia": estado}
        if motivo is not None:
            payload["motivo_ausencia"] = motivo if motivo else None

        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{estudiante_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json=payload)

        if res.status_code in (200, 204):
            return jsonify({"success": True, "message": "Estado actualizado correctamente"})
        else:
            return jsonify({"success": False,
                            "message": f"Error Supabase: {res.status_code} — {res.text}"}), 500

    except Exception as e:
        print(f"ERROR api_estudiante_estado_ausencia: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Agregar alumno extra / de reemplazo
#  POST /api/estudiante/agregar
#  Body JSON: { nomina_id, nombre, rut, fecha_nacimiento, sexo, motivo_ingreso }
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/estudiante/agregar', methods=['POST'])
def api_estudiante_agregar():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "Body JSON requerido"}), 400

        nomina_id      = data.get('nomina_id', '').strip()
        nombre         = data.get('nombre', '').strip()
        rut            = data.get('rut', '').strip()
        fecha_nac      = data.get('fecha_nacimiento') or None
        sexo           = data.get('sexo') or None
        nacionalidad        = data.get('nacionalidad') or None
        diagnostico_previo  = data.get('diagnostico_previo') or None
        motivo_ingreso      = data.get('motivo_ingreso') or 'extra'
        estado_asist        = data.get('estado_asistencia') or 'extra'

        if not nomina_id or not nombre:
            return jsonify({"success": False, "message": "nomina_id y nombre son requeridos"}), 400

        new_id = str(uuid.uuid4())
        payload = {
            "id":                  new_id,
            "nomina_id":           nomina_id,
            "nombre":              nombre,
            "rut":                 rut or None,
            "fecha_nacimiento":    fecha_nac,
            "sexo":                sexo,
            "nacionalidad":        nacionalidad,
            "diagnostico_sospecha": diagnostico_previo,
            "estado_asistencia":   estado_asist,
            "motivo_ausencia":     motivo_ingreso,
            "evaluado_flag":       False
        }

        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json=payload)

        if res.status_code == 201:
            nuevo = res.json()[0] if res.json() else {}
            return jsonify({"success": True, "id": nuevo.get('id', new_id),
                            "message": "Alumno agregado correctamente"})
        else:
            return jsonify({"success": False,
                            "message": f"Error Supabase: {res.status_code} — {res.text}"}), 500

    except Exception as e:
        print(f"ERROR api_estudiante_agregar: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  PRESENCIA EN TIEMPO REAL
#  POST /api/presencia   — heartbeat de la doctora mientras evalúa
#  GET  /api/presencia   — lista para admin / coordinadora
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/presencia', methods=['POST'])
def api_presencia_post():
    """Recibe heartbeat de una doctora y upsert en presencia_doctoras."""
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    # Solo las doctoras registran presencia — admin y coordinadoras no aparecen en el panel
    if session.get('usuario') != 'doctora':
        return jsonify({"success": True, "skipped": True}), 200
    try:
        data          = request.get_json() or {}
        doctora_id    = session.get('usuario_id')
        nomina_id     = data.get('nomina_id') or session.get('current_nomina_id') or ''
        establec      = data.get('establecimiento') or ''
        accion        = data.get('accion', 'heartbeat')   # heartbeat | logout

        if accion == 'logout':
            estado = 'desconectada'
        else:
            estado = 'evaluando'

        payload = {
            "doctora_id":            doctora_id,
            "nomina_id":             nomina_id or None,
            "establecimiento":       establec,
            "ultima_actividad":      "now()",
            "estado":                estado,
        }

        # Upsert: si ya existe la fila para esta doctora la actualiza
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/presencia_doctoras",
            headers={**SUPABASE_SERVICE_HEADERS,
                     "Prefer": "resolution=merge-duplicates,return=minimal",
                     "on_conflict": "doctora_id"},
            json=payload
        )
        # Supabase devuelve 200/201/204 en upsert
        return jsonify({"success": res.status_code in (200, 201, 204)})

    except Exception as e:
        print(f"ERROR api_presencia_post: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/presencia', methods=['GET'])
def api_presencia_get():
    """Devuelve la lista de presencia de todas las doctoras para admin / coordinadora."""
    if 'usuario' not in session:
        return jsonify({"success": False}), 401
    if session.get('usuario') not in ('admin', 'coordinadora'):
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        # 1. Traer todas las filas de presencia
        res_p = requests.get(
            f"{SUPABASE_URL}/rest/v1/presencia_doctoras"
            f"?select=doctora_id,establecimiento,ultima_actividad,estado",
            headers=SUPABASE_SERVICE_HEADERS
        )
        filas = res_p.json() if res_p.ok else []

        # 2. Traer nombres de doctoras
        res_d = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?rol=eq.doctora&select=id,nombre,usuario",
            headers=SUPABASE_SERVICE_HEADERS
        )
        doctoras_map = {}
        if res_d.ok:
            for d in res_d.json():
                doctoras_map[str(d['id'])] = d.get('nombre') or d.get('usuario', '—')

        # 3. Calcular estado dinámico según última actividad
        from datetime import timezone
        ahora = datetime.now(timezone.utc)
        resultado = []
        for fila in filas:
            did   = str(fila.get('doctora_id', ''))
            ua    = fila.get('ultima_actividad', '')
            estado = fila.get('estado', 'desconectada')

            # Si tiene timestamp, recalcular
            minutos = None
            if ua and estado != 'desconectada':
                try:
                    # Supabase devuelve ISO con timezone
                    ts = datetime.fromisoformat(ua.replace('Z', '+00:00'))
                    minutos = int((ahora - ts).total_seconds() / 60)
                    if minutos > 20:
                        estado = 'desconectada'
                    elif minutos > 8:
                        estado = 'en_pausa'
                    else:
                        estado = 'evaluando'
                except Exception:
                    pass

            resultado.append({
                "doctora_id":      did,
                "nombre":          doctoras_map.get(did, '—'),
                "establecimiento": fila.get('establecimiento', ''),
                "ultima_actividad": ua,
                "minutos":         minutos,
                "estado":          estado,
            })

        # Solo incluir en resultado filas cuyo doctora_id esté en doctoras_map (rol=doctora)
        resultado = [r for r in resultado if r['doctora_id'] in doctoras_map]

        # Agregar doctoras sin fila (nunca se conectaron)
        ids_con_fila = {r['doctora_id'] for r in resultado}
        for did, nombre in doctoras_map.items():
            if did not in ids_con_fila:
                resultado.append({
                    "doctora_id": did, "nombre": nombre,
                    "establecimiento": "", "ultima_actividad": None,
                    "minutos": None, "estado": "desconectada"
                })

        # 4. Enriquecer con datos de nómina actual y velocidad
        for r in resultado:
            r['evaluados_nomina'] = 0
            r['total_nomina']     = 0
            r['pct_nomina']       = 0
            r['nomina_nombre']    = ''
            r['mins_por_eval']    = None  # promedio minutos por evaluación

        # Traer nomina_id de presencia
        res_p2 = requests.get(
            f"{SUPABASE_URL}/rest/v1/presencia_doctoras"
            f"?select=doctora_id,nomina_id",
            headers=SUPABASE_SERVICE_HEADERS)
        nomina_por_doc = {}
        if res_p2.ok:
            for row in res_p2.json():
                if row.get('nomina_id'):
                    nomina_por_doc[str(row['doctora_id'])] = str(row['nomina_id'])

        # Para cada doctora activa, traer datos de su nómina
        for r in resultado:
            if r['estado'] == 'desconectada':
                continue
            nid = nomina_por_doc.get(r['doctora_id'])
            if not nid:
                continue
            # Nombre nómina
            res_nom = requests.get(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?id=eq.{nid}&select=nombre_nomina,nombre_colegio",
                headers=SUPABASE_SERVICE_HEADERS)
            if res_nom.ok and res_nom.json():
                nd = res_nom.json()[0]
                r['nomina_nombre'] = nd.get('nombre_colegio') or nd.get('nombre_nomina') or ''

            # Alumnos activos y evaluados
            res_alum = requests.get(
                f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                f"?nomina_id=eq.{nid}"
                f"&estado_asistencia=in.(activo,extra)"
                f"&select=evaluado_flag,fecha_relleno",
                headers=SUPABASE_SERVICE_HEADERS)
            if res_alum.ok:
                alumnos = res_alum.json()
                total   = len(alumnos)
                evaluados_list = [a for a in alumnos if a.get('evaluado_flag')]
                evaluados = len(evaluados_list)
                r['total_nomina']     = total
                r['evaluados_nomina'] = evaluados
                r['pct_nomina']       = round(evaluados / total * 100, 1) if total > 0 else 0

                # Calcular promedio de minutos por evaluación
                # Usamos fecha_relleno de hoy como proxy de evaluaciones de esta jornada
                from datetime import timezone, timedelta
                hoy = datetime.now(timezone.utc).date()
                tiempos = []
                for a in evaluados_list:
                    fr = a.get('fecha_relleno')
                    if fr:
                        try:
                            fd = datetime.fromisoformat(fr.replace('Z','+00:00')).date() if 'T' in fr else datetime.strptime(fr,'%Y-%m-%d').date()
                            if fd == hoy:
                                tiempos.append(fr)
                        except:
                            pass

                if len(tiempos) >= 2 and r.get('ultima_actividad'):
                    # Estimación: tiempo total activo / evaluaciones de hoy
                    # Usamos minutos de la sesión actual dividido por evaluados hoy
                    mins_sesion = r.get('minutos') or 0
                    if mins_sesion > 0 and len(tiempos) > 0:
                        r['mins_por_eval'] = round(mins_sesion / len(tiempos), 1)

        # Ordenar: más rápidas primero (menor mins_por_eval), luego pausa, luego desconectadas
        def sort_key(x):
            if x['estado'] == 'desconectada': return (2, 9999)
            if x['estado'] == 'en_pausa':     return (1, 9999)
            # evaluando: ordenar por mins_por_eval (None = sin datos, va al final)
            mpe = x.get('mins_por_eval')
            return (0, mpe if mpe is not None else 9999)

        resultado.sort(key=sort_key)
        return jsonify({"success": True, "data": resultado})

    except Exception as e:
        print(f"ERROR api_presencia_get: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  ELIMINAR ALUMNO INDIVIDUAL
#  DELETE /api/estudiante/eliminar/<estudiante_id>
#  Accesible por admin y doctora (solo sus propias nóminas)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/estudiante/eliminar/<estudiante_id>', methods=['DELETE'])
def api_estudiante_eliminar(estudiante_id):
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    user_role = session.get('usuario')
    user_id   = session.get('usuario_id')

    try:
        # 1. Verificar que el alumno existe y obtener su nomina_id
        res_est = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?id=eq.{estudiante_id}&select=id,nombre,nomina_id",
            headers=SUPABASE_SERVICE_HEADERS
        )
        if not res_est.ok or not res_est.json():
            return jsonify({"success": False, "message": "Alumno no encontrado"}), 404

        alumno    = res_est.json()[0]
        nomina_id = alumno.get('nomina_id')

        # 2. Si es doctora, verificar que la nómina le pertenece
        if user_role == 'doctora':
            res_nom = requests.get(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?id=eq.{nomina_id}&doctora_id=eq.{user_id}&select=id",
                headers=SUPABASE_SERVICE_HEADERS
            )
            if not res_nom.ok or not res_nom.json():
                return jsonify({"success": False,
                                "message": "No tienes permiso para eliminar este alumno"}), 403

        # 3. Eliminar
        res_del = requests.delete(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{estudiante_id}",
            headers=SUPABASE_SERVICE_HEADERS
        )
        res_del.raise_for_status()

        print(f"INFO: Alumno {estudiante_id} ({alumno.get('nombre')}) eliminado por {user_role} {user_id}")
        return jsonify({
            "success": True,
            "message": f"Alumno '{alumno.get('nombre')}' eliminado correctamente.",
            "alumno_id": estudiante_id,
            "nomina_id": nomina_id
        })

    except requests.exceptions.RequestException as e:
        detail = e.response.text if hasattr(e, 'response') and e.response else str(e)
        return jsonify({"success": False, "message": f"Error de conexión: {detail}"}), 500
    except Exception as e:
        print(f"ERROR api_estudiante_eliminar: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  GUARDAR SIN PDF — permite guardar correcciones aunque ya esté evaluado
#  POST /guardar_evaluacion
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/guardar_evaluacion', methods=['POST'])
def guardar_evaluacion():
    """Guarda los datos del formulario SIN generar PDF y SIN cambiar evaluado_flag.
       Permite corregir datos de alumnos ya evaluados."""
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401

    estudiante_id = request.form.get('estudiante_id')
    nomina_id     = request.form.get('nomina_id') or session.get('current_nomina_id')
    doctora_id    = session.get('usuario_id')
    form_type     = session.get('current_form_type', 'medicina_familiar')

    if not all([estudiante_id, nomina_id, doctora_id]):
        return jsonify({"success": False, "message": "Faltan datos obligatorios"}), 400

    # Reutiliza las mismas funciones auxiliares internas de marcar_evaluado
    def map_to_boolean_local(field_name):
        raw = request.form.get(field_name)
        if raw is None:
            return None   # campo ausente → no modificar
        if raw.strip():
            return True   # valor presente → marcado
        return False      # string vacío → desmarcado

    # Datos comunes
    update_data = {
        'nombre':           get_form_field_value('nombre', request.form),
        'rut':              get_form_field_value('rut', request.form),
        'fecha_nacimiento': get_form_field_value('fecha_nacimiento', request.form, return_none_if_empty=True),
        'fecha_evaluacion': get_form_field_value('fecha_evaluacion', request.form, return_none_if_empty=True),
        'fecha_reevaluacion': get_form_field_value('fecha_reevaluacion', request.form, return_none_if_empty=True),
        'edad':             get_form_field_value('edad', request.form),
        'nacionalidad':     get_form_field_value('nacionalidad', request.form),
        'sexo':             get_form_field_value('sexo', request.form),
        # NO tocamos evaluado_flag ni fecha_relleno → permite correcciones
    }

    if form_type == 'medicina_familiar':
        diagnostico_unificado_valor = get_form_field_value('diagnostico_unificado', request.form)
        genero_f = get_form_field_value('genero_f', request.form)
        genero_m = get_form_field_value('genero_m', request.form)
        if genero_f:   update_data['sexo'] = 'F'
        elif genero_m: update_data['sexo'] = 'M'

        update_data.update({
            'diagnostico_1': diagnostico_unificado_valor,
            # Limpiar dx_previo solo si se guardó un valor válido del desplegable
            'diagnostico_sospecha': ('' if diagnostico_unificado_valor and diagnostico_unificado_valor.strip() else None),
            'diagnostico_2': diagnostico_unificado_valor,
            'diagnostico_complementario': get_form_field_value('diagnostico_complementario', request.form),
            'clasificacion':  get_form_field_value('clasificacion_imc', request.form) or None,
            'derivaciones':   get_form_field_value('derivaciones', request.form),
            'indicaciones':   get_form_field_value('indicaciones', request.form),
            'fecha_reevaluacion_select': get_form_field_value('fecha_reevaluacion_select', request.form, return_none_if_empty=True),
            'observacion_1':  get_form_field_value('observacion_1', request.form),
            'observacion_2':  get_form_field_value('observacion_2', request.form),
            'observacion_3':  get_form_field_value('observacion_3', request.form),
            'observacion_4':  get_form_field_value('observacion_4', request.form),
            'observacion_5':  get_form_field_value('observacion_5', request.form),
            'observacion_6':  get_form_field_value('observacion_6', request.form),
            'observacion_7':  get_form_field_value('observacion_7', request.form),
            'check_cesarea':              map_to_boolean_local('check_cesarea'),
            'check_atermino':             map_to_boolean_local('check_atermino'),
            'check_vaginal':              map_to_boolean_local('check_vaginal'),
            'check_prematuro':            map_to_boolean_local('check_prematuro'),
            'check_acorde':               map_to_boolean_local('check_acorde'),
            'check_retraso':               map_to_boolean_local('check_retraso'),
            'check_retrasogeneralizado':  map_to_boolean_local('check_retrasogeneralizado'),
            'check_esquemac':             map_to_boolean_local('check_esquemac'),
            'check_esquemai':             map_to_boolean_local('check_esquemai'),
            'check_alergiano':            map_to_boolean_local('check_alergiano'),
            'check_alergiasi':            map_to_boolean_local('check_alergiasi'),
            'check_cirugiano':            map_to_boolean_local('check_cirugiano'),
            'check_cirugiasi':            map_to_boolean_local('check_cirugiasi'),
            'check_visionsinalteracion':  map_to_boolean_local('check_visionsinalteracion'),
            'check_visionrefraccion':     map_to_boolean_local('check_visionrefraccion'),
            'check_audicionnormal':       map_to_boolean_local('check_audicionnormal'),
            'check_hipoacusia':           map_to_boolean_local('check_hipoacusia'),
            'check_tapondecerumen':       map_to_boolean_local('check_tapondecerumen'),
            'check_sinhallazgos':         map_to_boolean_local('check_sinhallazgos'),
            'check_caries':               map_to_boolean_local('check_caries'),
            'check_apinamientodental':    map_to_boolean_local('check_apinamientodental'),
            'check_retenciondental':      map_to_boolean_local('check_retenciondental'),
            'check_frenillolingual':      map_to_boolean_local('check_frenillolingual'),
            'check_hipertrofia':          map_to_boolean_local('check_hipertrofia'),
            'altura':         get_form_field_value('altura', request.form, return_none_if_empty=True),
            'peso':           get_form_field_value('peso', request.form, return_none_if_empty=True),
            'imc':            get_form_field_value('imc', request.form, return_none_if_empty=True) or None,
            'clasificacion_imc': get_form_field_value('clasificacion_imc', request.form, return_none_if_empty=True) or None,
        })

    elif form_type == 'neurologia':
        update_data.update({
            'estado_general': get_form_field_value('estado', request.form),
            'diagnostico':    get_form_field_value('diagnostico', request.form),
            'derivaciones':   get_form_field_value('derivaciones', request.form),
        })

    elif form_type == 'informe_neurologico':
        update_data.update({
            'motivo_consulta':       get_form_field_value('motivo_consulta', request.form),
            'observaciones':         get_form_field_value('observaciones', request.form),
            'observacion_neurologia': get_form_field_value('observacion_neurologia', request.form),
            'diagnostico':           get_form_field_value('diagnostico', request.form),
            'indicaciones':          get_form_field_value('indicaciones', request.form),
        })

    try:
        # No sobreescribir campos numéricos/clasificación con None si llegaron vacíos
        SKIP_IF_NONE = {'imc', 'clasificacion_imc', 'clasificacion', 'altura', 'peso'}
        clean_data = {k: v for k, v in update_data.items()
                      if not (k in SKIP_IF_NONE and v is None)}
        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina?id=eq.{estudiante_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json=clean_data
        )
        if res.status_code >= 400:
            return jsonify({"success": False, "message": f"Error al guardar: {res.text}"}), 500
        return jsonify({"success": True, "message": "Datos guardados correctamente."})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  Marcar notificación de corrección como vista por la coordinadora
#  POST /api/correcciones/marcar_vista
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/correcciones/marcar_vista', methods=['POST'])
def marcar_correccion_vista():
    if 'usuario' not in session:
        return jsonify({"success": False}), 401
    try:
        solicitud_id = (request.get_json() or {}).get('solicitud_id')
        if not solicitud_id:
            return jsonify({"success": False, "message": "Falta solicitud_id"}), 400
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?id=eq.{solicitud_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json={"notificacion_vista": True}
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  Actualizar mensaje de respuesta en corrección ya resuelta
#  POST /api/correcciones/actualizar_respuesta
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/correcciones/actualizar_respuesta', methods=['POST'])
def actualizar_respuesta_correccion():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    try:
        data         = request.get_json() or {}
        solicitud_id = data.get('solicitud_id')
        respuesta    = (data.get('respuesta_admin') or '').strip() or None
        if not solicitud_id:
            return jsonify({"success": False, "message": "Falta solicitud_id"}), 400
        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion?id=eq.{solicitud_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json={"respuesta_admin": respuesta, "notificacion_vista": False}
        )
        res.raise_for_status()
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  AGENTE IA — Analizar solicitud de corrección
#  POST /api/agente/analizar_correccion
#  Body: { solicitud_id }
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/agente/analizar_correccion', methods=['POST'])
def agente_analizar_correccion():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    if not ANTHROPIC_API_KEY:
        return jsonify({"success": False, "message": "API key de Anthropic no configurada"}), 500
    try:
        data         = request.get_json() or {}
        solicitud_id = data.get('solicitud_id')
        if not solicitud_id:
            return jsonify({"success": False, "message": "Falta solicitud_id"}), 400

        # 1. Obtener la solicitud
        res_sol = requests.get(
            f"{SUPABASE_URL}/rest/v1/solicitudes_correccion"
            f"?id=eq.{solicitud_id}"
            f"&select=*,estudiantes_nomina(nombre,rut,fecha_nacimiento,sexo,"
            f"diagnostico_1,diagnostico_2,diagnostico_complementario,clasificacion_imc,"
            f"derivaciones,observacion_1,observacion_2,observacion_3,observacion_4,"
            f"observacion_5,observacion_6,observacion_7,fecha_evaluacion,fecha_relleno,"
            f"evaluado_flag,nomina_id,nacionalidad)",
            headers=SUPABASE_SERVICE_HEADERS)
        sol_list = res_sol.json() if res_sol.ok else []
        if not sol_list:
            return jsonify({"success": False, "message": "Solicitud no encontrada"}), 404
        sol  = sol_list[0]
        est  = sol.get('estudiantes_nomina') or {}

        # 2. Obtener nombre del colegio
        colegio = ''
        if est.get('nomina_id'):
            res_nom = requests.get(
                f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                f"?id=eq.{est['nomina_id']}&select=nombre_colegio",
                headers=SUPABASE_SERVICE_HEADERS)
            if res_nom.ok and res_nom.json():
                colegio = res_nom.json()[0].get('nombre_colegio', '')

        # 3. Construir contexto para el agente
        alumno_ctx = f"""
DATOS ACTUALES DEL ALUMNO EN PLATAFORMA:
- Nombre: {est.get('nombre', 'N/A')}
- RUT: {est.get('rut', 'N/A')}
- Fecha nacimiento: {est.get('fecha_nacimiento', 'N/A')}
- Sexo: {est.get('sexo', 'N/A')}
- Nacionalidad: {est.get('nacionalidad', 'N/A')}
- Colegio: {colegio}
- Evaluado: {'Sí' if est.get('evaluado_flag') else 'No'}
- Fecha evaluación: {est.get('fecha_evaluacion', 'N/A')}
- Diagnóstico PIE: {est.get('diagnostico_1', 'N/A')}
- Diagnóstico complementario: {est.get('diagnostico_complementario', 'N/A')}
- Clasificación IMC: {est.get('clasificacion_imc', 'N/A')}
- Derivaciones: {est.get('derivaciones', 'N/A')}
- Observaciones: {' | '.join(filter(None, [est.get(f'observacion_{i}','') for i in range(1,8)]))}

SOLICITUD DE LA COORDINADORA:
{sol.get('detalles', 'Sin detalle')}
"""

        system_prompt = """Eres un asistente experto del programa PIE (Programa de Integración Escolar) de Chile.
Tu rol es analizar solicitudes de corrección de coordinadoras de escuela sobre evaluaciones médicas de estudiantes.

Debes responder SIEMPRE en formato JSON con esta estructura exacta:
{
  "procedencia": "PROCEDE" | "NO PROCEDE" | "REQUIERE REVISIÓN",
  "justificacion": "Explicación breve en 1-2 oraciones de por qué procede o no",
  "cambios_sugeridos": [
    {"campo": "nombre_del_campo", "valor_actual": "...", "valor_sugerido": "..."}
  ],
  "mensaje_coordinadora": "Mensaje claro y amable para enviar a la coordinadora explicando la decisión",
  "nivel_urgencia": "ALTA" | "MEDIA" | "BAJA"
}

Reglas:
- PROCEDE si el error es claro (dato incorrecto, typo, campo equivocado)
- NO PROCEDE si la solicitud pide algo clínicamente incorrecto o fuera de protocolo PIE
- REQUIERE REVISIÓN si necesitas más información o hay ambigüedad
- cambios_sugeridos puede estar vacío si no procede
- Sé directo, claro y profesional. Usa lenguaje simple."""

        # 4. Llamar a Claude
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key":         ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type":      "application/json",
            },
            json={
                "model":      ANTHROPIC_MODEL,
                "max_tokens": 1024,
                "system":     system_prompt,
                "messages":   [{"role": "user", "content": alumno_ctx}]
            },
            timeout=30
        )
        resp.raise_for_status()
        resp_data = resp.json()

        # 5. Parsear respuesta JSON del agente
        raw_text = resp_data['content'][0]['text'].strip()
        # Limpiar posibles markdown fences
        raw_text = re.sub(r'^```json\s*', '', raw_text)
        raw_text = re.sub(r'\s*```$',     '', raw_text)
        analisis = json.loads(raw_text)

        return jsonify({"success": True, "analisis": analisis, "alumno": {
            "nombre": est.get('nombre',''),
            "rut":    est.get('rut',''),
            "colegio": colegio,
        }})

    except json.JSONDecodeError as e:
        return jsonify({"success": False, "message": f"Error parseando respuesta del agente: {e}"}), 500
    except Exception as e:
        print(f"ERROR agente_analizar_correccion: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  AGENTE IA — Revisión de nómina completa por la DOCTORA
#  POST /api/doctora/revisar_nomina
#  Body: { nomina_id }
#  Solo barre alumnos evaluados (evaluado_flag=true, activo o extra)
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/doctora/revisar_nomina', methods=['POST'])
def doctora_revisar_nomina():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    try:
        data      = request.get_json() or {}
        nomina_id = data.get('nomina_id')
        if not nomina_id:
            return jsonify({"success": False, "message": "Falta nomina_id"}), 400

        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&evaluado_flag=eq.true"
            f"&estado_asistencia=in.(activo,extra)"
            f"&select=id,nombre,rut,fecha_nacimiento,edad,nacionalidad,sexo,"
            f"diagnostico_1,clasificacion_imc,altura,peso,imc,"
            f"check_cesarea,check_atermino,check_vaginal,check_prematuro,"
            f"check_acorde,check_retraso,check_retrasogeneralizado,"
            f"check_esquemac,check_esquemai,check_alergiano,check_alergiasi,"
            f"check_cirugiano,check_cirugiasi,check_visionsinalteracion,check_visionrefraccion,"
            f"check_audicionnormal,check_hipoacusia,check_tapondecerumen,"
            f"check_sinhallazgos,check_caries,check_apinamientodental,"
            f"check_retenciondental,check_frenillolingual,check_hipertrofia,"
            f"observacion_1,observacion_2,observacion_3,observacion_4,"
            f"observacion_5,observacion_6,observacion_7,fecha_evaluacion"
            f"&order=nombre.asc",
            headers=SUPABASE_SERVICE_HEADERS
        )
        alumnos = res.json() if res.ok else []

        if not alumnos:
            return jsonify({"success": True, "resultados": [], "total": 0})

        def obs_no_informado(t): return 'NO INFORMADO' in (t or '').upper()

        GRUPOS = [
            ("Antecedentes Perinatales",   ["check_cesarea","check_atermino","check_vaginal","check_prematuro"],                                                              "observacion_1"),
            ("Desarrollo Psicomotor (DSM)",["check_acorde","check_retraso","check_retrasogeneralizado"],                                                                      "observacion_2"),
            ("Vacunas",                    ["check_esquemac","check_esquemai"],                                                                                               "observacion_3"),
            ("Alergias",                   ["check_alergiano","check_alergiasi"],                                                                                             "observacion_3"),
            ("Cirugías/Hospitalizaciones", ["check_cirugiano","check_cirugiasi"],                                                                                             "observacion_6"),
            ("Visión",                     ["check_visionsinalteracion","check_visionrefraccion"],                                                                             "observacion_7"),
            ("Audición",                   ["check_audicionnormal","check_hipoacusia","check_tapondecerumen"],                                                                 None),
            ("Salud Bucodental",           ["check_sinhallazgos","check_caries","check_apinamientodental","check_retenciondental","check_frenillolingual","check_hipertrofia"], None),
        ]

        resultados = []
        for a in alumnos:
            errores = []

            if not a.get('rut'):              errores.append({"campo": "RUT",              "descripcion": "RUT vacío",                        "tipo": "error"})
            if not a.get('sexo'):             errores.append({"campo": "Sexo",             "descripcion": "Sexo no registrado",               "tipo": "error"})
            if not a.get('nacionalidad'):     errores.append({"campo": "Nacionalidad",     "descripcion": "Nacionalidad vacía",               "tipo": "error"})
            if not a.get('fecha_evaluacion'): errores.append({"campo": "Fecha evaluación", "descripcion": "Fecha de evaluación vacía",        "tipo": "error"})
            if not a.get('diagnostico_1'):    errores.append({"campo": "Diagnóstico PIE",  "descripcion": "Diagnóstico PIE vacío",            "tipo": "error"})
            if not a.get('altura'):           errores.append({"campo": "Altura",           "descripcion": "Altura no registrada",             "tipo": "advertencia"})
            if not a.get('peso'):             errores.append({"campo": "Peso",             "descripcion": "Peso no registrado",               "tipo": "advertencia"})
            if not a.get('imc'):              errores.append({"campo": "IMC",              "descripcion": "IMC no calculado",                 "tipo": "advertencia"})
            if not a.get('clasificacion_imc'):errores.append({"campo": "Clasificación IMC","descripcion": "Clasificación IMC vacía",          "tipo": "advertencia"})

            if a.get('fecha_nacimiento'):
                try:
                    from datetime import date as dt_date
                    fn   = dt_date.fromisoformat(a['fecha_nacimiento'])
                    diff = (dt_date.today() - fn).days / 365.25
                    if diff < 0:
                        errores.append({"campo": "Fecha de nacimiento", "descripcion": f"Fecha en el futuro ({a['fecha_nacimiento']})", "tipo": "error"})
                    elif diff > 25:
                        errores.append({"campo": "Fecha de nacimiento", "descripcion": f"Edad inusualmente alta ({diff:.1f} años)", "tipo": "advertencia"})
                except:
                    errores.append({"campo": "Fecha de nacimiento", "descripcion": "Formato de fecha inválido", "tipo": "error"})
            else:
                errores.append({"campo": "Fecha de nacimiento", "descripcion": "Fecha de nacimiento vacía", "tipo": "error"})

            for grupo, campos, obs_key in GRUPOS:
                if any(a.get(c) for c in campos):
                    continue
                if obs_key and obs_no_informado(a.get(obs_key)):
                    continue
                if obs_key:
                    errores.append({"campo": grupo, "descripcion": "Ningún check marcado y observación sin 'NO INFORMADO'", "tipo": "error"})
                else:
                    errores.append({"campo": grupo, "descripcion": "Debe tener al menos un check marcado", "tipo": "error"})

            resultados.append({"id": a.get('id'), "nombre": a.get('nombre', 'N/A'), "errores": errores})

        return jsonify({"success": True, "resultados": resultados, "total": len(resultados)})

    except Exception as e:
        print(f"ERROR doctora_revisar_nomina: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  AGENTE IA — Barrido de nómina completa
#  POST /api/agente/barrer_nomina
#  Body: { nomina_id }
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/agente/barrer_nomina', methods=['POST'])
def agente_barrer_nomina():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False, "message": "Acceso denegado"}), 403
    if not ANTHROPIC_API_KEY:
        return jsonify({"success": False, "message": "API key de Anthropic no configurada"}), 500
    try:
        data      = request.get_json() or {}
        nomina_id = data.get('nomina_id')
        if not nomina_id:
            return jsonify({"success": False, "message": "Falta nomina_id"}), 400

        # 1. Obtener datos de la nómina
        res_nom = requests.get(
            f"{SUPABASE_URL}/rest/v1/nominas_medicas"
            f"?id=eq.{nomina_id}&select=nombre_nomina,nombre_colegio,form_type",
            headers=SUPABASE_SERVICE_HEADERS)
        nom_data = res_nom.json()[0] if res_nom.ok and res_nom.json() else {}

        # 2. Obtener todos los alumnos evaluados de la nómina
        res_est = requests.get(
            f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
            f"?nomina_id=eq.{nomina_id}"
            f"&evaluado_flag=eq.true"
            f"&estado_asistencia=in.(activo,extra)"
            f"&select=id,nombre,rut,fecha_nacimiento,edad,nacionalidad,sexo,"
            f"diagnostico_1,diagnostico_complementario,clasificacion_imc,"
            f"derivaciones,check_cesarea,check_atermino,check_vaginal,check_prematuro,"
            f"check_acorde,check_retraso,check_retrasogeneralizado,"
            f"check_esquemac,check_esquemai,check_alergiano,check_alergiasi,"
            f"check_cirugiano,check_cirugiasi,check_visionsinalteracion,check_visionrefraccion,"
            f"check_audicionnormal,check_hipoacusia,check_tapondecerumen,"
            f"check_sinhallazgos,check_caries,check_apinamientodental,"
            f"check_retenciondental,check_frenillolingual,check_hipertrofia,"
            f"altura,peso,imc,clasificacion_imc,observacion_1,observacion_2,observacion_3,"
            f"observacion_4,observacion_5,observacion_6,observacion_7,fecha_evaluacion",
            headers=SUPABASE_SERVICE_HEADERS)
        alumnos = res_est.json() if res_est.ok else []

        if not alumnos:
            return jsonify({"success": True, "resumen": "No hay alumnos evaluados en esta nómina.", "errores": [], "total": 0, "con_errores": 0})

        # 3. Análisis local de errores (rápido, sin llamar a Claude)
        errores_por_alumno = []
        for a in alumnos:
            errores = []
            nombre = a.get('nombre', 'N/A')

            # ── Datos personales ──────────────────────────────────────────────
            if not a.get('rut'):             errores.append("RUT vacío")
            if not a.get('sexo'):            errores.append("Sexo no registrado")
            if not a.get('nacionalidad'):    errores.append("Nacionalidad vacía")
            if not a.get('fecha_evaluacion'):errores.append("Fecha de evaluación vacía")

            # ── Edad / Fecha de nacimiento ────────────────────────────────────
            edad_val = a.get('edad')
            if edad_val is not None:
                try:
                    edad_num = float(str(edad_val).replace(' años','').replace(' meses','').strip().split()[0])
                    if edad_num < 0:
                        errores.append(f"Edad negativa ({edad_val}) — fecha de nacimiento incorrecta")
                    elif edad_num > 25:
                        errores.append(f"Edad inusualmente alta ({edad_val}) — verificar fecha de nacimiento")
                except:
                    pass
            if a.get('fecha_nacimiento'):
                try:
                    from datetime import date as dt_date
                    fn = dt_date.fromisoformat(a['fecha_nacimiento'])
                    today = dt_date.today()
                    diff_years = (today - fn).days / 365.25
                    if diff_years < 0:
                        errores.append(f"Fecha de nacimiento en el futuro ({a['fecha_nacimiento']})")
                    elif diff_years > 25:
                        errores.append(f"Fecha de nacimiento inusual — edad calculada {diff_years:.1f} años")
                except:
                    errores.append("Formato de fecha de nacimiento inválido")
            else:
                errores.append("Fecha de nacimiento vacía")

            # ── Diagnóstico PIE (obligatorio) ─────────────────────────────────
            if not a.get('diagnostico_1'):
                errores.append("Diagnóstico PIE vacío — campo obligatorio")

            # ── IMC / Medidas antropométricas ─────────────────────────────────
            if not a.get('altura'):          errores.append("Altura no registrada")
            if not a.get('peso'):            errores.append("Peso no registrado")
            if not a.get('imc'):             errores.append("IMC no calculado")
            if not a.get('clasificacion_imc'): errores.append("Clasificación IMC vacía")

            # ── Helper: observacion contiene "NO INFORMADO" en cualquier parte ──
            def obs_no_informado(obs_text):
                return 'NO INFORMADO' in (obs_text or '').upper()

            # ── Grupos de checks con su observación de respaldo ────────────────
            # Regla: si ningún check del grupo está marcado, la observación
            # correspondiente debe contener "NO INFORMADO". Si no hay observación
            # asociada (Audición, Boca/Dental), el check es SIEMPRE obligatorio.
            GRUPOS = [
                # (nombre, campos_check, obs_respaldo_key)
                ("Antecedentes Perinatales", ["check_cesarea","check_atermino","check_vaginal","check_prematuro"], "observacion_1"),
                ("DSM",                      ["check_acorde","check_retraso","check_retrasogeneralizado"],         "observacion_2"),
                ("Vacunas",                  ["check_esquemac","check_esquemai"],                                  "observacion_3"),
                ("Alergias",                 ["check_alergiano","check_alergiasi"],                                "observacion_3"),
                ("Cirugías/Hospitalizaciones",["check_cirugiano","check_cirugiasi"],                               "observacion_6"),
                ("Visión",                   ["check_visionsinalteracion","check_visionrefraccion"],               "observacion_7"),
                # Sin observación de respaldo — check obligatorio siempre
                ("Audición",     ["check_audicionnormal","check_hipoacusia","check_tapondecerumen"],    None),
                ("Salud Bucodental", ["check_sinhallazgos","check_caries","check_apinamientodental",
                                     "check_retenciondental","check_frenillolingual","check_hipertrofia"], None),
            ]
            for grupo_nombre, campos, obs_key in GRUPOS:
                tiene_check = any(a.get(c) for c in campos)
                if tiene_check:
                    continue  # OK — tiene al menos un check marcado
                # No tiene check — verificar si hay observación con NO INFORMADO
                if obs_key and obs_no_informado(a.get(obs_key)):
                    continue  # OK — tiene NO INFORMADO en la observación
                # Error real
                if obs_key:
                    errores.append(f"{grupo_nombre}: ningún check marcado y observación sin 'NO INFORMADO'")
                else:
                    errores.append(f"{grupo_nombre}: debe tener al menos un check marcado")

            # Derivaciones vs Checks
            # Valores exactos del formulario: OFTALMÓLOGO, OTORRINO, DENTISTA,
            # NUTRICIONISTA — separados por ' - '
            import unicodedata as _ud
            def _nrm(s):
                return _ud.normalize('NFD', (s or '').upper()).encode('ascii','ignore').decode().strip()
            raw_deriv  = a.get('derivaciones') or ''
            drv_items  = [_nrm(d) for d in raw_deriv.replace('\n',' - ').split(' - ') if d.strip()]
            OFTALMO    = {'OFTALMOLOGO','OFTALMO','OFTALM','OFTALMOLOGIA'}
            OTORR      = {'OTORRINO','OTORRINOLARINGOLOGO','ORL','OTORRIN'}
            DENTIST    = {'DENTISTA','ODONTOLOGO','ODONTOLOGIA','DENTAL'}
            NUTRI      = {'NUTRICIONISTA','NUTRIOLOGO','NUTRICION'}
            def tiene_esp(keys):
                for item in drv_items:
                    for k in keys:
                        if k in item or item in k:
                            return True
                return False
            if a.get('check_visionrefraccion'):
                if not tiene_esp(OFTALMO):
                    errores.append('Refraccion visual marcada pero sin derivacion a Oftalmologo')
            if a.get('check_hipoacusia') or a.get('check_tapondecerumen') or a.get('check_hipertrofia'):
                if not tiene_esp(OTORR):
                    errores.append('Hallazgo auditivo marcado pero sin derivacion a Otorrino')
            if a.get('check_caries') or a.get('check_apinamientodental') or a.get('check_frenillolingual'):
                if not tiene_esp(DENTIST):
                    errores.append('Hallazgo dental marcado pero sin derivacion a Dentista')
            clasif = (a.get('clasificacion_imc') or '').lower()
            if 'obesidad' in clasif or 'bajo peso' in clasif:
                if not tiene_esp(NUTRI):
                    errores.append('IMC ' + str(a.get('clasificacion_imc')) + ' sin derivacion a Nutricionista')

            if errores:
                errores_por_alumno.append({
                    "alumno_id": str(a.get('id','')),
                    "nombre":    nombre,
                    "rut":       a.get('rut','N/A'),
                    "nomina_id": nomina_id,
                    "errores":   errores,
                    "total_errores": len(errores)
                })

        total        = len(alumnos)
        con_errores  = len(errores_por_alumno)
        sin_errores  = total - con_errores

        # 4. Llamar a Claude para generar resumen inteligente
        ctx = f"""Eres el asistente del sistema CardioHome del programa PIE chileno.
Acabas de hacer un barrido de la nómina "{nom_data.get('nombre_colegio','')}" con {total} alumnos evaluados.
Resultados: {con_errores} alumnos con errores, {sin_errores} sin errores.

Errores encontrados:
{json.dumps(errores_por_alumno, ensure_ascii=False, indent=2) if errores_por_alumno else 'Ninguno — todo correcto.'}

Genera un resumen ejecutivo breve (máximo 4 oraciones) para el administrador, mencionando los errores más críticos y si es urgente corregirlos antes de cerrar la jornada. Sé directo y claro. No uses markdown."""

        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": ANTHROPIC_MODEL,
                "max_tokens": 300,
                "messages": [{"role": "user", "content": ctx}]
            },
            timeout=20
        )
        resumen_ia = resp.json()['content'][0]['text'].strip() if resp.ok else "Barrido completado."

        return jsonify({
            "success":      True,
            "total":        total,
            "con_errores":  con_errores,
            "sin_errores":  sin_errores,
            "resumen_ia":   resumen_ia,
            "errores":      errores_por_alumno,
            "colegio":      nom_data.get('nombre_colegio',''),
        })

    except Exception as e:
        print(f"ERROR agente_barrer_nomina: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ─────────────────────────────────────────────────────────────────────────────
#  AGENTE IA — Chat flotante (pregunta libre sobre la plataforma)
#  POST /api/agente/chat
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/agente/chat', methods=['POST'])
def agente_chat():
    if 'usuario' not in session:
        return jsonify({"success": False, "message": "No autorizado"}), 401
    if not ANTHROPIC_API_KEY:
        return jsonify({"success": False, "message": "API key no configurada"}), 500
    try:
        data     = request.get_json() or {}
        mensaje  = data.get('mensaje', '').strip()
        historial = data.get('historial', [])
        if not mensaje:
            return jsonify({"success": False, "message": "Mensaje vacío"}), 400

        system = """Eres el asistente IA de CardioHome, plataforma de evaluaciones médicas del Programa PIE (Programa de Integración Escolar) de Chile.
Ayudas a administradores, coordinadoras y doctoras a usar la plataforma.
Conoces el flujo completo: carga de nóminas Excel, evaluaciones médicas (medicina familiar y neurología), diagnósticos PIE, derivaciones automáticas, sistema de correcciones y roles de usuario.
Responde siempre en español, de forma clara, breve y conversacional.
IMPORTANTE: NO uses Markdown bajo ninguna circunstancia. Nada de asteriscos, nada de ##, nada de guiones como listas, nada de negrita ni cursiva. Escribe en texto plano como si fuera una conversación normal.
Máximo 3 oraciones salvo que se pida más detalle. Si necesitas listar cosas, sepáralas con coma o punto y coma, nunca con guiones ni asteriscos."""

        messages = historial[-6:] + [{"role": "user", "content": mensaje}]

        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={"model": ANTHROPIC_MODEL, "max_tokens": 400, "system": system, "messages": messages},
            timeout=20
        )
        resp.raise_for_status()
        respuesta = resp.json()['content'][0]['text'].strip()
        return jsonify({"success": True, "respuesta": respuesta})

    except Exception as e:
        print(f"ERROR agente_chat: {e}")
        return jsonify({"success": False, "message": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
#  CHAT SOPORTE — Chat en vivo coordinadora ↔ admin
#  Tabla Supabase: chat_soporte_sesiones + chat_soporte_mensajes
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/soporte/solicitar', methods=['POST'])
def soporte_solicitar():
    """Coordinadora solicita chat con un admin disponible."""
    rol = session.get('usuario')
    if rol not in ('coordinador_escuela', 'coordinadora', 'doctora'):
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        data      = request.get_json() or {}
        nombre    = (data.get('nombre') or session.get('nombre') or '').strip()
        escuela   = (data.get('escuela') or '').strip()
        es_general = (rol in ('coordinadora', 'doctora'))  # doctora y coord general = prioridad alta

        if not nombre:
            return jsonify({"success": False, "message": "Ingresa tu nombre"}), 400
        if not es_general and not escuela:
            return jsonify({"success": False, "message": "Ingresa el establecimiento"}), 400

        # Token único por pestaña del navegador (enviado desde el JS frontend)
        # Esto aísla coordinadores que comparten el mismo usuario_id
        import uuid as _uuid
        tab_token = (data.get('tab_token') or '').strip()
        if not tab_token:
            tab_token = str(_uuid.uuid4())
        # Guardar en sesión Flask también para logout
        session['soporte_session_token'] = tab_token
        session_token  = tab_token
        solicitante_id = str(session.get('usuario_id', ''))

        # Crear sesión de chat
        sesion = {
            "solicitante_id":     solicitante_id,
            "session_token":      session_token,
            "solicitante_nombre": nombre,
            "solicitante_escuela": escuela if not es_general else "Coordinación General",
            "solicitante_rol":    rol,
            "estado":             "esperando",   # esperando | activo | cerrado
            "prioridad":          "alta" if es_general else "normal",
            "admin_id":           None,
            "admin_nombre":       None,
        }
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json=sesion
        )
        if not res.ok:
            return jsonify({"success": False, "message": res.text}), 500

        nueva_sesion = res.json()[0]
        return jsonify({"success": True, "sesion": nueva_sesion})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/sesiones_pendientes', methods=['GET'])
def soporte_sesiones_pendientes():
    """Admin: ver solicitudes de chat pendientes y activas propias."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        admin_id = str(session.get('usuario_id', ''))
        # Pendientes (cualquier admin puede tomar)
        res_pend = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones"
            f"?estado=eq.esperando&order=created_at.asc"
            f"&select=id,solicitante_nombre,solicitante_escuela,solicitante_rol,prioridad,created_at",
            headers=SUPABASE_SERVICE_HEADERS
        )
        # Activas de este admin
        res_activa = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones"
            f"?estado=eq.activo&admin_id=eq.{admin_id}"
            f"&select=id,solicitante_nombre,solicitante_escuela,solicitante_rol,prioridad,created_at",
            headers=SUPABASE_SERVICE_HEADERS
        )
        return jsonify({
            "success":   True,
            "pendientes": res_pend.json() if res_pend.ok else [],
            "activas":    res_activa.json() if res_activa.ok else [],
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/aceptar', methods=['POST'])
def soporte_aceptar():
    """Admin acepta una solicitud de chat."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        data     = request.get_json() or {}
        sesion_id = data.get('sesion_id')
        admin_id  = str(session.get('usuario_id', ''))
        admin_nombre = session.get('nombre') or session.get('usuario') or 'Admin'
        # Si nombre no está en sesión, buscarlo en BD
        if not session.get('nombre'):
            try:
                res_n = requests.get(f"{SUPABASE_URL}/rest/v1/doctoras?id=eq.{admin_id}&select=nombre", headers=SUPABASE_SERVICE_HEADERS)
                if res_n.ok and res_n.json():
                    admin_nombre = res_n.json()[0].get('nombre') or admin_nombre
            except: pass

        # Verificar que aún esté esperando
        res_check = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones?id=eq.{sesion_id}&select=estado",
            headers=SUPABASE_SERVICE_HEADERS
        )
        if not res_check.ok or not res_check.json():
            return jsonify({"success": False, "message": "Sesión no encontrada"}), 404
        if res_check.json()[0]['estado'] != 'esperando':
            return jsonify({"success": False, "message": "Esta solicitud ya fue tomada por otro admin"}), 409

        res = requests.patch(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones?id=eq.{sesion_id}",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json={"estado": "activo", "admin_id": admin_id, "admin_nombre": admin_nombre}
        )
        if not res.ok:
            return jsonify({"success": False, "message": res.text}), 500

        # Mensaje de bienvenida automático
        requests.post(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_mensajes",
            headers=SUPABASE_SERVICE_HEADERS,
            json={
                "sesion_id": sesion_id,
                "autor_rol": "admin",
                "autor_nombre": admin_nombre,
                "mensaje": f"Hola, soy {admin_nombre}. ¿En qué te puedo ayudar?"
            }
        )
        return jsonify({"success": True, "sesion": res.json()[0]})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/mensajes/<sesion_id>', methods=['GET'])
def soporte_mensajes(sesion_id):
    """Obtener mensajes de una sesión + estado de la sesión."""
    rol = session.get('usuario')
    if rol not in ('admin', 'coordinador_escuela', 'coordinadora', 'doctora'):
        return jsonify({"success": False}), 403
    try:
        since = request.args.get('since', '')
        url = (
            f"{SUPABASE_URL}/rest/v1/chat_soporte_mensajes"
            f"?sesion_id=eq.{sesion_id}"
            f"&order=created_at.asc"
            f"&limit=100"
        )
        if since:
            # Supabase acepta ISO timestamp directamente con gt
            url += f"&created_at=gt.{since}"
        else:
            # Sin since: traer últimos 50 mensajes
            url = url.replace("&limit=100", "&limit=50")

        res_msgs = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        res_ses  = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones?id=eq.{sesion_id}&select=estado,admin_nombre,admin_id,solicitante_nombre",
            headers=SUPABASE_SERVICE_HEADERS
        )
        sesion_data = res_ses.json()[0] if res_ses.ok and res_ses.json() else {}
        return jsonify({
            "success":  True,
            "mensajes": res_msgs.json() if res_msgs.ok else [],
            "sesion":   sesion_data,
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/enviar', methods=['POST'])
def soporte_enviar():
    """Enviar mensaje en una sesión activa."""
    rol = session.get('usuario')
    if rol not in ('admin', 'coordinador_escuela', 'coordinadora', 'doctora'):
        return jsonify({"success": False}), 403
    try:
        data      = request.get_json() or {}
        sesion_id = data.get('sesion_id')
        mensaje   = (data.get('mensaje') or '').strip()
        if not mensaje or not sesion_id:
            return jsonify({"success": False, "message": "Datos incompletos"}), 400

        nombre = session.get('nombre') or session.get('usuario') or rol
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_mensajes",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json={"sesion_id": sesion_id, "autor_rol": rol, "autor_nombre": nombre, "mensaje": mensaje}
        )
        if not res.ok:
            return jsonify({"success": False, "message": res.text}), 500
        return jsonify({"success": True, "mensaje": res.json()[0]})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/cerrar', methods=['POST'])
def soporte_cerrar():
    """Cerrar una sesión de chat."""
    rol = session.get('usuario')
    if rol not in ('admin', 'coordinador_escuela', 'coordinadora', 'doctora'):
        return jsonify({"success": False}), 403
    try:
        data      = request.get_json() or {}
        sesion_id = data.get('sesion_id')
        requests.patch(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones?id=eq.{sesion_id}",
            headers=SUPABASE_SERVICE_HEADERS,
            json={"estado": "cerrado"}
        )
        return jsonify({"success": True})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/mi_sesion', methods=['GET'])
def soporte_mi_sesion():
    """Coordinadora: verificar si tiene sesión activa o esperando."""
    rol = session.get('usuario')
    if rol not in ('coordinador_escuela', 'coordinadora', 'doctora'):
        return jsonify({"success": False}), 403
    try:
        solicitante_id = str(session.get('usuario_id', ''))
        session_token = session.get('soporte_session_token', '')
        filtro_id = f'session_token=eq.{session_token}' if session_token else f'solicitante_id=eq.{solicitante_id}'
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones"
            f"?{filtro_id}"
            f"&estado=in.(esperando,activo)"
            f"&order=created_at.desc&limit=1"
            f"&select=id,estado,admin_nombre,created_at,prioridad",
            headers=SUPABASE_SERVICE_HEADERS
        )
        sesiones = res.json() if res.ok else []
        return jsonify({"success": True, "sesion": sesiones[0] if sesiones else None})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/chat_admin_mensajes', methods=['GET'])
def soporte_chat_admin_mensajes():
    """Admin: traer mensajes nuevos de su sesión activa."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        admin_id = str(session.get('usuario_id', ''))
        since    = request.args.get('since', '')
        # Buscar sesión activa de este admin
        res_ses = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones"
            f"?admin_id=eq.{admin_id}&estado=eq.activo&select=id",
            headers=SUPABASE_SERVICE_HEADERS
        )
        sesiones = res_ses.json() if res_ses.ok else []
        if not sesiones:
            return jsonify({"success": True, "mensajes": [], "sesion_id": None})
        sesion_id = sesiones[0]['id']
        url = (
            f"{SUPABASE_URL}/rest/v1/chat_soporte_mensajes"
            f"?sesion_id=eq.{sesion_id}&order=created_at.asc"
        )
        if since: url += f"&created_at=gt.{since}"
        res_msgs = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        return jsonify({
            "success": True,
            "mensajes": res_msgs.json() if res_msgs.ok else [],
            "sesion_id": sesion_id
        })
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/chat_admin/mensajes', methods=['GET'])
def chat_admin_get():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        since = request.args.get('since', '')
        url = f"{SUPABASE_URL}/rest/v1/chat_admin?select=id,autor_id,autor_nombre,mensaje,created_at&order=created_at.asc"
        if since:  url += f"&created_at=gt.{since}"
        else:      url += f"&limit=60"
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        return jsonify({"success": True, "mensajes": res.json() if res.ok else []})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/chat_admin/enviar', methods=['POST'])
def chat_admin_enviar():
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        data    = request.get_json() or {}
        mensaje = (data.get('mensaje') or '').strip()
        if not mensaje: return jsonify({"success": False}), 400
        autor_id     = session.get('usuario_id')
        autor_nombre = session.get('nombre') or session.get('usuario') or 'Admin'
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/chat_admin",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json={"autor_id": str(autor_id), "autor_nombre": autor_nombre, "mensaje": mensaje}
        )
        if not res.ok: return jsonify({"success": False}), 500
        return jsonify({"success": True, "mensaje": res.json()[0]})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/soporte/historial', methods=['GET'])
def soporte_historial():
    """Admin: últimas 30 sesiones de soporte (atendidas y no atendidas)."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/chat_soporte_sesiones"
            f"?order=created_at.desc&limit=30"
            f"&select=id,solicitante_nombre,solicitante_escuela,solicitante_rol,"
            f"prioridad,estado,admin_id,admin_nombre,created_at",
            headers=SUPABASE_SERVICE_HEADERS
        )
        sesiones = res.json() if res.ok else []
        return jsonify({"success": True, "sesiones": sesiones})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


# ═══════════════════════════════════════════════════════════════════════════════
#  CHAT PRIVADO ENTRE ADMINS
# ═══════════════════════════════════════════════════════════════════════════════

@app.route('/api/chat_admin/admins', methods=['GET'])
def chat_admin_lista_admins():
    """Lista todos los usuarios con rol admin."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        res = requests.get(
            f"{SUPABASE_URL}/rest/v1/doctoras?rol=eq.admin&select=id,nombre,usuario",
            headers=SUPABASE_SERVICE_HEADERS
        )
        return jsonify({"success": True, "admins": res.json() if res.ok else []})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/chat_admin/mensajes_privados/<dest_id>', methods=['GET'])
def chat_admin_mensajes_privados(dest_id):
    """Mensajes del hilo privado entre el admin actual y dest_id."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        mi_id = str(session.get('usuario_id', ''))
        since = request.args.get('since', '')
        # Mensajes donde (autor=yo, dest=ellos) OR (autor=ellos, dest=yo)
        url = (
            f"{SUPABASE_URL}/rest/v1/chat_admin_privado"
            f"?or=(and(autor_id.eq.{mi_id},dest_id.eq.{dest_id}),and(autor_id.eq.{dest_id},dest_id.eq.{mi_id}))"
            f"&order=created_at.asc&limit=80"
        )
        if since:
            url += f"&created_at=gt.{since}"
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        return jsonify({"success": True, "mensajes": res.json() if res.ok else []})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/chat_admin/enviar_privado', methods=['POST'])
def chat_admin_enviar_privado():
    """Envía un mensaje privado a otro admin."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        data    = request.get_json() or {}
        dest_id = data.get('dest_id')
        mensaje = (data.get('mensaje') or '').strip()
        if not mensaje or not dest_id:
            return jsonify({"success": False, "message": "Datos incompletos"}), 400
        autor_id     = str(session.get('usuario_id', ''))
        autor_nombre = session.get('nombre') or session.get('usuario') or 'Admin'
        res = requests.post(
            f"{SUPABASE_URL}/rest/v1/chat_admin_privado",
            headers={**SUPABASE_SERVICE_HEADERS, "Prefer": "return=representation"},
            json={"autor_id": autor_id, "dest_id": dest_id,
                  "autor_nombre": autor_nombre, "mensaje": mensaje}
        )
        if not res.ok:
            return jsonify({"success": False, "message": res.text}), 500
        return jsonify({"success": True, "mensaje": res.json()[0]})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500


@app.route('/api/chat_admin/mensajes_recibidos', methods=['GET'])
def chat_admin_mensajes_recibidos():
    """Mensajes privados recibidos por el admin actual desde cualquier otro admin."""
    if session.get('usuario') != 'admin':
        return jsonify({"success": False}), 403
    try:
        mi_id = str(session.get('usuario_id', ''))
        since = request.args.get('since', '')
        url   = (
            f"{SUPABASE_URL}/rest/v1/chat_admin_privado"
            f"?dest_id=eq.{mi_id}"
            f"&order=created_at.asc&limit=20"
        )
        if since:
            url += f"&created_at=gt.{since}"
        res = requests.get(url, headers=SUPABASE_SERVICE_HEADERS)
        return jsonify({"success": True, "mensajes": res.json() if res.ok else []})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Coordinadora — Exportar listado alumnos (Excel .xlsx)
#  GET /api/coordinadora/export_listado?tipo=evaluados|pendientes
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/coordinadora/export_listado', methods=['GET'])
def api_coordinadora_export_listado():
    if session.get('usuario') not in ('coordinadora', 'admin'):
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO

        tipo      = request.args.get('tipo', 'evaluados')
        user_role = session.get('usuario')
        user_id   = session.get('usuario_id')

        # Nóminas según rol
        if user_role == 'coordinadora':
            url_n = (f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                     f"?coord_general_id=eq.{user_id}"
                     f"&select=id,nombre_colegio,nombre_nomina,tipo_nomina,doctora_id")
        else:
            url_n = (f"{SUPABASE_URL}/rest/v1/nominas_medicas"
                     f"?select=id,nombre_colegio,nombre_nomina,tipo_nomina,doctora_id")
        nominas = requests.get(url_n, headers=SUPABASE_SERVICE_HEADERS).json() or []
        if not nominas:
            return jsonify({"success": False, "message": "Sin nóminas"}), 404

        nom_map  = {n['id']: n for n in nominas}
        nom_ids  = list(nom_map.keys())

        # Nombres doctoras
        res_d = requests.get(f"{SUPABASE_URL}/rest/v1/doctoras?select=id,nombre,usuario",
                             headers=SUPABASE_SERVICE_HEADERS)
        docs_map = {d['id']: (d.get('nombre') or d.get('usuario','')) for d in (res_d.json() if res_d.ok else [])}

        # Alumnos — iterar por nomina con paginación (evita límite 1000 Supabase)
        ev_flag = 'true' if tipo == 'evaluados' else 'false'
        alumnos = []
        for nom_id_e in nom_ids:
            page_size = 1000
            offset    = 0
            while True:
                url_a = (f"{SUPABASE_URL}/rest/v1/estudiantes_nomina"
                         f"?nomina_id=eq.{nom_id_e}"
                         f"&evaluado_flag=eq.{ev_flag}"
                         f"&estado_asistencia=in.(activo,extra)"
                         f"&select=nombre,rut,fecha_evaluacion,nomina_id"
                         f"&order=nombre.asc"
                         f"&limit={page_size}&offset={offset}")
                res_a = requests.get(url_a, headers=SUPABASE_SERVICE_HEADERS)
                if not res_a.ok:
                    break
                page = res_a.json()
                alumnos.extend(page)
                if len(page) < page_size:
                    break
                offset += page_size

        # Crear Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Listado" if tipo == 'evaluados' else "Pendientes"

        hdr_fill  = PatternFill("solid", fgColor="0F3460")
        hdr_font  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        body_font = Font(name="Calibri", size=10)
        alt_fill  = PatternFill("solid", fgColor="EEF4FF")
        thin      = Side(style='thin', color='D1D5DB')
        border    = Border(left=thin, right=thin, top=thin, bottom=thin)
        center    = Alignment(horizontal='center', vertical='center')

        titulo = "ALUMNOS EVALUADOS" if tipo == 'evaluados' else "ALUMNOS PENDIENTES"
        ws.merge_cells('A1:F1')
        ws['A1'] = f"REPORTE PIE — {titulo}"
        ws['A1'].font = Font(name="Calibri", bold=True, color="FFFFFF", size=13)
        ws['A1'].fill = PatternFill("solid", fgColor="0F3460")
        ws['A1'].alignment = center
        ws.row_dimensions[1].height = 28

        headers = ['N°','Nombre Alumno','RUT','Establecimiento','Tipo Nómina','Doctora','Fecha Evaluación','Estado']
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=2, column=col, value=h)
            c.font = hdr_font; c.fill = hdr_fill
            c.alignment = center; c.border = border

        for idx, al in enumerate(alumnos, 1):
            nom = nom_map.get(al.get('nomina_id'), {})
            row = idx + 2
            vals = [
                idx,
                al.get('nombre',''),
                al.get('rut',''),
                nom.get('nombre_colegio') or nom.get('nombre_nomina',''),
                nom.get('tipo_nomina',''),
                docs_map.get(nom.get('doctora_id',''),''),
                al.get('fecha_evaluacion',''),
                'Evaluado' if tipo == 'evaluados' else 'Pendiente'
            ]
            fill = alt_fill if idx % 2 == 0 else None
            for col, v in enumerate(vals, 1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = body_font; c.border = border
                if fill: c.fill = fill
                if col == 1: c.alignment = center

        # Anchos de columna
        for w, col in zip([6,32,14,30,18,22,16,12], range(1,9)):
            ws.column_dimensions[chr(64+col)].width = w

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"PIE_{tipo}_{__import__('datetime').date.today().isoformat()}.xlsx"
        return buf.read(), 200, {
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': f'attachment; filename="{filename}"'
        }

    except Exception as e:
        print(f"ERROR export_listado: {e}")
        return jsonify({"success": False, "message": str(e)}), 500



# ─────────────────────────────────────────────────────────────────────────────
# RUTA: Coordinadora — Exportar Reporte PDF (descarga directa con xhtml2pdf)
#  GET /api/coordinadora/export_pdf
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/api/coordinadora/export_pdf', methods=['GET'])
def api_coordinadora_export_pdf():
    if session.get('usuario') not in ('coordinadora', 'admin'):
        return jsonify({"success": False, "message": "No autorizado"}), 403
    try:
        from xhtml2pdf import pisa
        from io import BytesIO
        from datetime import date as date_cls

        user_role = session.get('usuario')
        user_id   = session.get('usuario_id')

        # Nominas
        if user_role == 'coordinadora':
            url_n = (SUPABASE_URL + "/rest/v1/nominas_medicas"
                     "?coord_general_id=eq." + str(user_id) +
                     "&select=id,nombre_colegio,nombre_nomina,form_type,tipo_nomina,doctora_id")
        else:
            url_n = (SUPABASE_URL + "/rest/v1/nominas_medicas"
                     "?select=id,nombre_colegio,nombre_nomina,form_type,tipo_nomina,doctora_id")
        nominas = requests.get(url_n, headers=SUPABASE_SERVICE_HEADERS).json() or []
        if not nominas:
            return jsonify({"success": False, "message": "Sin nominas"}), 404

        nom_map = {n['id']: n for n in nominas}

        # Doctoras
        res_d = requests.get(SUPABASE_URL + "/rest/v1/doctoras?select=id,nombre,usuario",
                             headers=SUPABASE_SERVICE_HEADERS)
        docs_map = {d['id']: (d.get('nombre') or d.get('usuario',''))
                    for d in (res_d.json() if res_d.ok else [])}

        # Totales
        total_global = 0; completed_global = 0
        neuro_total = 0;  neuro_comp = 0
        fam_total = 0;    fam_comp = 0
        ranking_dict = {}
        est_dict = {}

        for nom in nominas:
            nid   = nom['id']
            ftype = (nom.get('form_type') or '').lower()
            tipo  = (nom.get('tipo_nomina') or '').lower()
            t = get_supabase_count("nomina_id=eq." + str(nid) + "&estado_asistencia=in.(activo,extra)")
            c = get_supabase_count("nomina_id=eq." + str(nid) + "&evaluado_flag=eq.true&estado_asistencia=in.(activo,extra)")
            total_global += t; completed_global += c

            if 'neuro' in ftype or 'neuro' in tipo:
                neuro_total += t; neuro_comp += c
            elif 'familiar' in ftype or 'medicina' in ftype or 'familiar' in tipo:
                fam_total += t; fam_comp += c

            did = str(nom.get('doctora_id', '') or '')
            if did:
                if did not in ranking_dict:
                    ranking_dict[did] = {'nombre': docs_map.get(did,''), 'completados': 0, 'total': 0}
                ranking_dict[did]['completados'] += c
                ranking_dict[did]['total'] += t

            colegio = nom.get('nombre_colegio') or nom.get('nombre_nomina','Sin nombre')
            if colegio not in est_dict:
                est_dict[colegio] = {'completados': 0, 'total': 0}
            est_dict[colegio]['completados'] += c
            est_dict[colegio]['total'] += t

        pending_global = total_global - completed_global
        pct_global = round(completed_global / total_global * 100, 1) if total_global > 0 else 0
        neuro_pct = round(neuro_comp / neuro_total * 100) if neuro_total > 0 else 0
        fam_pct   = round(fam_comp / fam_total * 100) if fam_total > 0 else 0

        # Ranking rows
        ranking = sorted(ranking_dict.values(), key=lambda x: x['completados'], reverse=True)[:8]
        medals = ['1.', '2.', '3.']
        rank_rows = ''
        for i, r in enumerate(ranking):
            p = round(r['completados']/r['total']*100) if r['total'] > 0 else 0
            bar_w = min(p, 100)
            medal = medals[i] if i < 3 else str(i+1) + '.'
            rank_rows += (
                '<tr>'
                '<td style="text-align:center;">' + medal + '</td>'
                '<td>' + (r['nombre'] or 'Sin nombre') + '</td>'
                '<td style="text-align:center;color:#059669;font-weight:700;">' + str(r['completados']) + '</td>'
                '<td style="text-align:center;">' + str(r['total']) + '</td>'
                '<td>'
                '<div style="display:inline-block;width:80px;height:7px;background:#e2e8f0;border-radius:4px;overflow:hidden;vertical-align:middle;">'
                '<div style="width:' + str(bar_w) + '%;height:100%;background:#1c67a3;"></div>'
                '</div>'
                ' <span style="font-size:8px;font-weight:700;color:#1c67a3;">' + str(p) + '%</span>'
                '</td>'
                '</tr>'
            )

        # Establecimientos rows
        establecimientos = sorted(est_dict.items(), key=lambda x: x[1]['completados'], reverse=True)[:8]
        est_rows = ''
        for nom_e, vals_e in establecimientos:
            p = round(vals_e['completados']/vals_e['total']*100) if vals_e['total'] > 0 else 0
            col = '#059669' if p >= 80 else '#d97706' if p >= 50 else '#dc2626'
            est_rows += (
                '<tr>'
                '<td>' + nom_e + '</td>'
                '<td style="text-align:center;">' + str(vals_e['total']) + '</td>'
                '<td style="text-align:center;color:#059669;font-weight:700;">' + str(vals_e['completados']) + '</td>'
                '<td style="text-align:center;font-weight:800;color:' + col + ';">' + str(p) + '%</td>'
                '</tr>'
            )

        hoy_str = date_cls.today().strftime('%d/%m/%Y')
        pct_str = str(pct_global) + '%'

        html_parts = [
            '<!DOCTYPE html><html lang="es"><head><meta charset="UTF-8">',
            '<style>',
            '* { box-sizing:border-box; margin:0; padding:0; }',
            'body { font-family:Arial,sans-serif; font-size:10px; color:#1e293b; padding:16px; }',
            '.header { background:#0f3460; color:white; border-radius:8px; padding:16px 20px; margin-bottom:14px; }',
            '.header h1 { font-size:14px; font-weight:900; margin-bottom:3px; color:white; }',
            '.header p { font-size:9px; opacity:.75; color:white; }',
            '.header-pct { float:right; font-size:28px; font-weight:900; color:white; }',
            '.kpi-table { width:100%; margin-bottom:12px; }',
            '.kpi-cell { background:#f8faff; border:1px solid #e2e8f0; border-radius:8px; padding:8px; text-align:center; width:20%; }',
            '.kpi-v { font-size:16px; font-weight:900; color:#0f3460; }',
            '.kpi-l { font-size:8px; color:#94a3b8; text-transform:uppercase; letter-spacing:.5px; margin-top:2px; }',
            '.section { border:1px solid #e8f0f6; border-radius:8px; margin-bottom:12px; }',
            '.sec-head { background:#f8faff; padding:8px 12px; font-size:8px; font-weight:800; text-transform:uppercase; letter-spacing:.5px; color:#64748b; border-bottom:1px solid #e8f0f6; }',
            '.sec-body { padding:12px; }',
            'table { width:100%; border-collapse:collapse; font-size:9px; }',
            'th { padding:5px 8px; background:#0f3460; color:white; text-align:left; font-size:8px; font-weight:700; }',
            'td { padding:5px 8px; border-bottom:1px solid #f1f5f9; }',
            '.bar-outer { display:inline-block; width:80px; height:7px; background:#e2e8f0; border-radius:4px; overflow:hidden; vertical-align:middle; }',
            '.footer { margin-top:10px; font-size:8px; color:#94a3b8; text-align:center; border-top:1px solid #f1f5f9; padding-top:8px; }',
            '@page { size:A4; margin:15mm 12mm; }',
            '</style></head><body>',
            '<div class="header">',
            '<span class="header-pct">' + pct_str + '</span>',
            '<h1>Reporte de Avance PIE</h1>',
            '<p>Coordinacion General &middot; ' + hoy_str + '</p>',
            '</div>',
            '<table class="kpi-table"><tr>',
            '<td class="kpi-cell"><div class="kpi-v">' + str(total_global) + '</div><div class="kpi-l">Total</div></td>',
            '<td class="kpi-cell"><div class="kpi-v" style="color:#059669;">' + str(completed_global) + '</div><div class="kpi-l">Evaluados</div></td>',
            '<td class="kpi-cell"><div class="kpi-v" style="color:#f97316;">' + str(pending_global) + '</div><div class="kpi-l">Pendientes</div></td>',
            '<td class="kpi-cell"><div class="kpi-v" style="color:#8b5cf6;">' + str(neuro_comp) + '/' + str(neuro_total) + '</div><div class="kpi-l">Neurologia</div></td>',
            '<td class="kpi-cell"><div class="kpi-v" style="color:#1c67a3;">' + str(fam_comp) + '/' + str(fam_total) + '</div><div class="kpi-l">Med. Familiar</div></td>',
            '</tr></table>',
            '<div class="section"><div class="sec-head">Progreso por Especialidad</div><div class="sec-body">',
            '<table><tr>',
            '<td style="width:50%;padding:10px;text-align:center;">',
            '<div style="font-size:22px;font-weight:900;color:#8b5cf6;">' + str(neuro_pct) + '%</div>',
            '<div style="font-size:9px;color:#64748b;margin:4px 0;">Neurologia &mdash; ' + str(neuro_comp) + '/' + str(neuro_total) + ' evaluados</div>',
            '<div style="width:100%;height:10px;background:#f1f5f9;border-radius:5px;overflow:hidden;">',
            '<div style="width:' + str(neuro_pct) + '%;height:100%;background:#8b5cf6;border-radius:5px;"></div></div>',
            '</td>',
            '<td style="width:50%;padding:10px;text-align:center;">',
            '<div style="font-size:22px;font-weight:900;color:#f97316;">' + str(fam_pct) + '%</div>',
            '<div style="font-size:9px;color:#64748b;margin:4px 0;">Med. Familiar &mdash; ' + str(fam_comp) + '/' + str(fam_total) + ' evaluados</div>',
            '<div style="width:100%;height:10px;background:#f1f5f9;border-radius:5px;overflow:hidden;">',
            '<div style="width:' + str(fam_pct) + '%;height:100%;background:#f97316;border-radius:5px;"></div></div>',
            '</td>',
            '</tr></table>',
            '</div></div>',
            '<div class="section"><div class="sec-head">Ranking de Doctoras</div><div class="sec-body">',
            '<table><thead><tr>',
            '<th style="width:28px;">#</th><th>Doctora</th>',
            '<th style="width:60px;text-align:center;">Evaluados</th>',
            '<th style="width:50px;text-align:center;">Total</th>',
            '<th style="width:120px;">Avance</th>',
            '</tr></thead><tbody>' + rank_rows + '</tbody></table>',
            '</div></div>',
            '<div class="section"><div class="sec-head">Top Establecimientos</div><div class="sec-body">',
            '<table><thead><tr>',
            '<th>Establecimiento</th>',
            '<th style="width:50px;text-align:center;">Total</th>',
            '<th style="width:60px;text-align:center;">Evaluados</th>',
            '<th style="width:40px;text-align:center;">%</th>',
            '</tr></thead><tbody>' + est_rows + '</tbody></table>',
            '</div></div>',
            '<div class="footer">CardioHome &middot; Sistema PIE &middot; ' + hoy_str + '</div>',
            '</body></html>'
        ]
        html = ''.join(html_parts)

        buf = BytesIO()
        result = pisa.CreatePDF(html, dest=buf)
        if result.err:
            return jsonify({"success": False, "message": "Error al generar PDF"}), 500
        buf.seek(0)
        filename = "Reporte_PIE_" + date_cls.today().isoformat() + ".pdf"
        return buf.read(), 200, {
            'Content-Type': 'application/pdf',
            'Content-Disposition': 'attachment; filename="' + filename + '"'
        }

    except Exception as e:
        print("ERROR export_pdf: " + str(e))
        return jsonify({"success": False, "message": str(e)}), 500


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 8080)))
